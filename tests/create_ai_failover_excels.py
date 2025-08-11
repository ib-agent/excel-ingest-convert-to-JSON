"""
Create Excel workbooks designed to trigger AI failover based on complexity
metrics in `converter/excel_complexity_analyzer.py`.

Files generated into `tests/fixtures/excel/`:
- ai_failover_multi_level_headers.xlsx
- ai_failover_complex_merges.xlsx
- ai_failover_sparse_clusters.xlsx
- ai_failover_irregular_boundaries.xlsx
- ai_failover_dynamic_formulas.xlsx
- ai_failover_nonstandard_layout.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

TARGET_DIR = Path(__file__).parent / "test_excel"


def save_book(wb: Workbook, filename: str) -> None:
    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET_DIR / filename)


def make_multi_level_headers() -> None:
    """3+ header rows with inconsistent patterns to raise header complexity."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MultiLevelHeaders"

    # Header rows (3 levels) with some gaps
    ws["A1"] = "Region"; ws["B1"] = "Q1"; ws["C1"] = "Q2"; ws["D1"] = "Q3"; ws["E1"] = "Q4"
    ws["A2"] = "Region"; ws["B2"] = "Jan"; ws["C2"] = "Feb"; ws["D2"] = "Mar"  # E2 left blank
    ws["A3"] = "Region"; ws["B3"] = "01/31/2025"; ws["C3"] = "02/28/2025"; ws["D3"] = "03/31/2025"; ws["E3"] = "04/30/2025"

    for cell in ["A1","B1","C1","D1","E1","A2","B2","C2","D2","A3","B3","C3","D3","E3"]:
        ws[cell].font = Font(bold=True)

    # Data rows with mixed content to increase column inconsistency
    ws["A4"] = "North"; ws["B4"] = 100; ws["C4"] = 120; ws["D4"] = 110; ws["E4"] = 95
    ws["A5"] = "South"; ws["B5"] = "N/A"; ws["C5"] = 130; ws["D5"] = "-"; ws["E5"] = 105
    ws["A6"] = "East"; ws["B6"] = 90; ws["C6"] = "128"; ws["D6"] = 115; ws["E6"] = 100

    # Inflate dimensions to increase sparsity: write a far-away cell
    ws["BJ173"] = "dim"

    save_book(wb, "ai_failover_multi_level_headers.xlsx")


def make_complex_merges() -> None:
    """Create extensive merged cells across multiple rows/cols to raise merge complexity."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ComplexMerges"

    # Large merged header spanning rows and columns
    ws.merge_cells("A1:E2")
    ws["A1"] = "Consolidated Statement"
    ws["A1"].font = Font(bold=True)

    # Section headers merged down and across
    ws.merge_cells("A3:A5"); ws["A3"] = "Assets"
    ws.merge_cells("A6:A8"); ws["A6"] = "Liabilities"
    ws.merge_cells("B3:C3"); ws["B3"] = "Current"
    ws.merge_cells("D3:E3"); ws["D3"] = "Non-Current"

    # Data rows scatter
    ws["B4"] = 1000; ws["C4"] = 1200; ws["D4"] = 500; ws["E4"] = 700
    ws["B5"] = 800; ws["C5"] = 900; ws["D5"] = 400; ws["E5"] = 600

    # Inflate dimensions for sparsity
    ws["BJ173"] = 1

    save_book(wb, "ai_failover_complex_merges.xlsx")


def make_sparse_clusters() -> None:
    """Sparse sheet with several clusters to confuse boundary detection and raise sparsity complexity."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SparseClusters"

    # Cluster 1
    ws["B2"] = "Header"; ws["C2"] = "Jan"; ws["D2"] = "Feb"
    ws["B3"] = "A"; ws["C3"] = 10; ws["D3"] = 12

    # Cluster 2 far away
    ws["K10"] = "Header"; ws["L10"] = "Mar"; ws["M10"] = "Apr"
    ws["K11"] = "B"; ws["L11"] = 8; ws["M11"] = 9

    # Inflate dimensions
    ws["BL200"] = "x"

    save_book(wb, "ai_failover_sparse_clusters.xlsx")


def make_irregular_boundaries() -> None:
    """Inconsistent column counts and mixed types across rows to raise boundary ambiguity/inconsistency."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IrregularBoundaries"

    # Header-like rows with gaps
    ws["A1"] = "Item"; ws["C1"] = "Q1"; ws["E1"] = "Q2"
    ws["A2"] = "Item"; ws["B2"] = "Jan"; ws["D2"] = "Feb"

    # Irregular data rows
    ws["A3"] = "X"; ws["B3"] = 1; ws["C3"] = 2
    ws["A4"] = "Y"; ws["D4"] = 3; ws["E4"] = 4
    ws["A5"] = "Z"; ws["B5"] = "3"; ws["E5"] = "4"

    # Inflate dimensions
    ws["BO210"] = 0

    save_book(wb, "ai_failover_irregular_boundaries.xlsx")


def make_dynamic_formulas() -> None:
    """Formula-heavy area to raise formula complexity metric."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DynamicFormulas"

    # Simple header
    ws["A1"] = "Metric"; ws["B1"] = "Value1"; ws["C1"] = "Value2"; ws["D1"] = "Calc"

    # Data + formulas (include complex functions)
    ws["A2"] = "Alpha"; ws["B2"] = 10; ws["C2"] = 5; ws["D2"] = "=IF(B2>5, SUMPRODUCT(B2:C2), INDEX(B2:C2,1))"
    ws["A3"] = "Beta"; ws["B3"] = 2; ws["C3"] = 7; ws["D3"] = "=IF(B3<3, VLOOKUP(2, B2:C3, 2, FALSE), MATCH(7, C2:C3, 0))"
    ws["A4"] = "Gamma"; ws["B4"] = 9; ws["C4"] = 1; ws["D4"] = "=SUMPRODUCT(B4:C4)"

    # Inflate dimensions
    ws["BH180"] = "tail"

    save_book(wb, "ai_failover_dynamic_formulas.xlsx")


def make_nonstandard_layout() -> None:
    """Non-standard layout: nested/side-by-side blocks and a section header-only row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NonStandard"

    # Section header row (label in first column only)
    ws["A2"] = "Operating Metrics"
    ws["A2"].font = Font(bold=True)

    # Side-by-side blocks beneath
    ws["A3"] = "Metric"; ws["B3"] = "Jan"; ws["C3"] = "Feb"; ws["D3"] = "Mar"
    ws["A4"] = "Units"; ws["B4"] = 10; ws["C4"] = 11; ws["D4"] = 12

    ws["F3"] = "Metric"; ws["G3"] = "Q1"; ws["H3"] = "Q2"
    ws["F4"] = "Revenue"; ws["G4"] = 100; ws["H4"] = 110

    # Inflate dimensions
    ws["BP220"] = 1

    save_book(wb, "ai_failover_nonstandard_layout.xlsx")


def main() -> None:
    make_multi_level_headers()
    make_complex_merges()
    make_sparse_clusters()
    make_irregular_boundaries()
    make_dynamic_formulas()
    make_nonstandard_layout()
    print(f"Created AI failover Excel test files in: {TARGET_DIR}")


if __name__ == "__main__":
    main()


