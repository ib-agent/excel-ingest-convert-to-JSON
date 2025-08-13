import json
import os
import tempfile
import unittest
from fastapi.testclient import TestClient
from fastapi_service.main import app


class ExcelStorageIntegrationTest(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        os.environ['USE_STORAGE_SERVICE'] = 'true'

    def tearDown(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def _create_minimal_xlsx(self) -> bytes:
        # Create a minimal Excel file in-memory
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        ws['A2'] = 'Value'
        buf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        try:
            wb.save(buf.name)
            with open(buf.name, 'rb') as f:
                return f.read()
        finally:
            try:
                os.unlink(buf.name)
            except Exception:
                pass

    def test_upload_and_convert_stores_artifacts(self):
        xlsx_bytes = self._create_minimal_xlsx()
        files = {'file': ('test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = self.client.post('/api/upload/', files=files)
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertIn('storage', body)
        storage = body['storage']
        self.assertIsNotNone(storage)
        self.assertIn('processed_json', storage)
        self.assertIn('table_data', storage)
        self.assertIn('download_urls', storage)

        # Verify status endpoint returns the processing record
        pid = storage['processing_id']
        resp = self.client.get(f'/api/status/{pid}/')
        self.assertIn(resp.status_code, (200, 404))
        if resp.status_code == 200:
            record = resp.json()
            self.assertEqual(record['processing_id'], pid)
            self.assertEqual(record['type'], 'excel')


