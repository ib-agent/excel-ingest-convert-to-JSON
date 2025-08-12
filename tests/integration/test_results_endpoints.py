import os
import tempfile
from fastapi.testclient import TestClient

from fastapi_service.main import app


class TestResultsEndpoints:
    def setup_method(self):
        self.client = TestClient(app)
        # For deterministic local storage
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        # Force off to exercise cache path in one test
        os.environ['USE_STORAGE_SERVICE'] = 'false'

    def teardown_method(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def _create_minimal_xlsx(self) -> bytes:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws['A1'] = 'H1'; ws['B1'] = 'H2'
        ws.append(['v1', 1])
        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        try:
            wb.save(f.name)
            with open(f.name, 'rb') as r:
                return r.read()
        finally:
            try:
                os.unlink(f.name)
            except Exception:
                pass

    def _test_excel_flow_and_fetch_results(self, use_storage: bool):
        os.environ['USE_STORAGE_SERVICE'] = 'true' if use_storage else 'false'
        xlsx_bytes = self._create_minimal_xlsx()
        files = {'file': ('mini.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = self.client.post('/api/upload/', files=files)
        assert resp.status_code == 200
        body = resp.json()
        processing_id = body['processing_id']

        # Health of status record
        status_resp = self.client.get(f'/api/status/{processing_id}/')
        assert status_resp.status_code == 200
        status_body = status_resp.json()
        assert status_body['processing_id'] == processing_id

        # Retrieve results via new endpoints
        full_resp = self.client.get(f'/api/results/{processing_id}/full')
        assert full_resp.status_code == 200
        table_resp = self.client.get(f'/api/results/{processing_id}/table')
        assert table_resp.status_code == 200

    def test_results_endpoints_with_cache(self):
        self._test_excel_flow_and_fetch_results(use_storage=False)

    def test_results_endpoints_with_storage(self):
        self._test_excel_flow_and_fetch_results(use_storage=True)

    def test_pdf_flow_populates_processing_and_fetch(self):
        os.environ['USE_STORAGE_SERVICE'] = 'false'
        repo_root = os.path.dirname(os.path.dirname(__file__))
        pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers.pdf')
        if not os.path.exists(pdf_path):
            pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers_with_before_and_after_paragraphs.pdf')
        with open(pdf_path, 'rb') as f:
            files = {'file': ('mini.pdf', f.read(), 'application/pdf')}
        resp = self.client.post('/api/pdf/upload/', files=files)
        assert resp.status_code == 200
        processing_id = resp.json().get('processing_id')
        assert processing_id is not None
        # status should be present
        status_resp = self.client.get(f'/api/status/{processing_id}/')
        assert status_resp.status_code == 200


