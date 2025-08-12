import os
import tempfile
from fastapi.testclient import TestClient

# Import the FastAPI app
from fastapi_service.main import app


class TestFastAPIBasic:
    def setup_method(self):
        self.client = TestClient(app)
        # Ensure storage uses local temp path
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        os.environ['USE_STORAGE_SERVICE'] = 'true'

    def teardown_method(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def _create_minimal_xlsx(self) -> bytes:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws['A1'] = 'Header'; ws['A2'] = 'Value'
        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        try:
            wb.save(f.name)
            with open(f.name, 'rb') as r:
                return r.read()
        finally:
            try:
                os.unlink(f.name)
            except Exception:
                pass

    def test_health(self):
        resp = self.client.get('/api/health/')
        assert resp.status_code == 200
        assert resp.json()['status'] == 'healthy'

    def test_ui_pages(self):
        assert self.client.get('/').status_code == 200
        assert self.client.get('/excel/').status_code == 200
        assert self.client.get('/pdf/').status_code == 200

    def test_storage_endpoints(self):
        # Store JSON
        payload = {
            'storage_type': 'processed',
            'key_prefix': 'fastapi_test',
            'data': {'hello': 'world'}
        }
        resp = self.client.post('/api/storage/store-json/', json=payload)
        assert resp.status_code == 201
        key = resp.json()['reference']['key']

        # Get JSON
        resp = self.client.get(f'/api/storage/get-json/?key={key}')
        assert resp.status_code == 200
        assert resp.json()['data'] == {'hello': 'world'}

        # List
        prefix = key.split('/')[0]
        resp = self.client.get(f'/api/storage/list/?prefix={prefix}')
        assert resp.status_code == 200
        assert any(item['key'] == key for item in resp.json()['items'])

        # Delete
        resp = self.client.delete(f'/api/storage/delete/?key={key}')
        assert resp.status_code == 200
        assert resp.json()['deleted'] is True

    def test_excel_upload(self):
        xlsx_bytes = self._create_minimal_xlsx()
        files = {'file': ('test.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = self.client.post('/api/upload/', files=files)
        assert resp.status_code == 200
        body = resp.json()
        assert body['success'] is True
        assert body['format'] == 'compact'
        assert body['filename'] == 'test.xlsx'
        assert 'processing_id' in body

    def test_excel_download_cache_fallback(self):
        # Seed the in-memory cache used by Django impl and our FastAPI wrapper
        from converter import models as django_like_models
        file_id = 'test123'
        django_like_models.processed_data_cache[file_id] = {
            'full_data': {'hello': 'world'},
            'table_data': {'tables': []},
            'filename': 'seed.xlsx',
            'format': 'compact'
        }
        resp = self.client.get(f'/api/download/?type=full&file_id={file_id}')
        assert resp.status_code == 200
        assert resp.headers.get('content-disposition') is not None
        assert resp.json()['hello'] == 'world'

    def test_pdf_upload(self):
        # Use existing small test PDF
        repo_root = os.path.dirname(os.path.dirname(__file__))
        pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers.pdf')
        if not os.path.exists(pdf_path):
            # Fallback
            pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers_with_before_and_after_paragraphs.pdf')
        with open(pdf_path, 'rb') as f:
            files = {'file': ('test.pdf', f.read(), 'application/pdf')}
        resp = self.client.post('/api/pdf/upload/', files=files)
        assert resp.status_code == 200
        body = resp.json()
        assert body['success'] is True
        assert body['filename'] == 'test.pdf'
        assert 'result' in body

    def test_pdf_process_with_options(self):
        repo_root = os.path.dirname(os.path.dirname(__file__))
        pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers.pdf')
        if not os.path.exists(pdf_path):
            pdf_path = os.path.join(repo_root, 'fixtures', 'pdfs', 'Test_PDF_Table_9_numbers_with_before_and_after_paragraphs.pdf')
        payload = {
            'file_path': pdf_path,
            'extract_tables': True,
            'extract_text': True,
            'extract_numbers': True
        }
        resp = self.client.post('/api/pdf/process/', json=payload)
        assert resp.status_code == 200
        body = resp.json()
        assert body['success'] is True


