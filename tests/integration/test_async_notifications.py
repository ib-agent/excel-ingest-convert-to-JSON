import os
import tempfile
import threading
import http.server
import socketserver
import json
import time
from fastapi.testclient import TestClient

from fastapi_service.main import app


class _WebhookHandler(http.server.BaseHTTPRequestHandler):
    last_payload = None
    def do_POST(self):
        length = int(self.headers.get('content-length', '0'))
        body = self.rfile.read(length)
        try:
            _WebhookHandler.last_payload = json.loads(body.decode('utf-8'))
        except Exception:
            _WebhookHandler.last_payload = None
        self.send_response(200)
        self.end_headers()
    def log_message(self, format, *args):
        return


def _start_webhook_server():
    # Bind random available port
    httpd = socketserver.TCPServer(("127.0.0.1", 0), _WebhookHandler)
    port = httpd.server_address[1]
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    return httpd, port


class TestAsyncNotifications:
    def setup_method(self):
        self.client = TestClient(app)
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        os.environ['USE_STORAGE_SERVICE'] = 'false'

    def teardown_method(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def _create_minimal_xlsx(self) -> bytes:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws['A1'] = 'H1'; ws['B1'] = 'H2'
        ws.append(['v1', 1])
        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        try:
            wb.save(f.name)
            with open(f.name, 'rb') as r:
                return r.read()
        finally:
            try:
                os.unlink(f.name)
            except Exception:
                pass

    def test_excel_async_with_webhook(self):
        # Start webhook server
        httpd, port = _start_webhook_server()
        try:
            xlsx_bytes = self._create_minimal_xlsx()
            files = {'file': ('mini.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp = self.client.post('/api/upload/', files=files, data={'async_mode': 'true', 'callback_url': f'http://127.0.0.1:{port}/cb'})
            assert resp.status_code == 202
            processing_id = resp.json()['processing_id']

            # Poll status until completed or timeout
            for _ in range(60):
                st = self.client.get(f'/api/status/{processing_id}/').json()
                if st.get('status') == 'completed':
                    break
                time.sleep(0.2)
            assert st.get('status') == 'completed'

            # Webhook should have been hit with payload containing endpoints
            for _ in range(30):
                if _WebhookHandler.last_payload:
                    break
                time.sleep(0.1)
            assert _WebhookHandler.last_payload is not None
            payload = _WebhookHandler.last_payload
            assert payload.get('processing_id') == processing_id
            assert 'results_endpoints' in payload

            # Try fetching results via endpoint
            full_resp = self.client.get(f"/api/results/{processing_id}/full")
            assert full_resp.status_code == 200
        finally:
            try:
                httpd.shutdown()
            except Exception:
                pass

    def test_pdf_async_with_status(self):
        repo_root = os.path.dirname(os.path.dirname(__file__))
        pdf_path = os.path.join(repo_root, 'tests', 'test_pdfs', 'Test_PDF_Table_9_numbers.pdf')
        if not os.path.exists(pdf_path):
            pdf_path = os.path.join(repo_root, 'tests', 'test_pdfs', 'Test_PDF_Table_9_numbers_with_before_and_after_paragraphs.pdf')
        with open(pdf_path, 'rb') as f:
            files = {'file': ('mini.pdf', f.read(), 'application/pdf')}
        resp = self.client.post('/api/pdf/upload/', files=files, data={'async_mode': 'true'})
        assert resp.status_code == 202
        processing_id = resp.json()['processing_id']
        for _ in range(60):
            st = self.client.get(f'/api/status/{processing_id}/').json()
            if st.get('status') == 'completed':
                break
            time.sleep(0.2)
        assert st.get('status') == 'completed'
        # Fetch via results endpoint
        tr = self.client.get(f'/api/results/{processing_id}/table')
        assert tr.status_code == 200


