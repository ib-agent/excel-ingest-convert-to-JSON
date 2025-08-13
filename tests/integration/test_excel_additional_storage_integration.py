import os
import tempfile
import unittest
from fastapi.testclient import TestClient
from fastapi_service.main import app


class ExcelAdditionalStorageIntegrationTest(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        os.environ['USE_STORAGE_SERVICE'] = 'true'

    def tearDown(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def _create_minimal_xlsx(self) -> bytes:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws['A1'] = 'Header'; ws['A2'] = 'Value'
        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(f.name)
        with open(f.name, 'rb') as r:
            data = r.read()
        os.unlink(f.name)
        return data

    def test_analyze_excel_complexity_stores_artifacts(self):
        xlsx_bytes = self._create_minimal_xlsx()
        files = {'file': ('complexity.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = self.client.post('/api/excel/analyze-complexity/', files=files)
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertIn('storage', body)
        storage = body['storage']
        self.assertIsNotNone(storage)
        self.assertIn('processing_id', storage)
        self.assertIn('original_file', storage)
        self.assertIn('complexity_results', storage)
        self.assertIn('download_urls', storage)

        pid = storage['processing_id']
        resp = self.client.get(f'/api/status/{pid}/')
        self.assertEqual(resp.status_code, 200)
        record = resp.json()
        self.assertEqual(record['processing_id'], pid)
        self.assertEqual(record['type'], 'excel_complexity')

    def test_excel_comparison_analysis_stores_artifacts(self):
        xlsx_bytes = self._create_minimal_xlsx()
        files = {'file': ('comparison.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = self.client.post('/api/excel/comparison-analysis/', files=files)
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertIn('storage', body)
        storage = body['storage']
        self.assertIsNotNone(storage)
        self.assertIn('processing_id', storage)
        self.assertIn('original_file', storage)
        self.assertIn('comparison_results', storage)
        self.assertIn('download_urls', storage)

        pid = storage['processing_id']
        resp = self.client.get(f'/api/status/{pid}/')
        self.assertEqual(resp.status_code, 200)
        record = resp.json()
        self.assertEqual(record['processing_id'], pid)
        self.assertEqual(record['type'], 'excel_comparison')


