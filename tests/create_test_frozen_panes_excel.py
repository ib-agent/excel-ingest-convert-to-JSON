#!/usr/bin/env python3
"""
Create a test Excel file with frozen panes and data patterns that would
normally trigger table detection to verify the frozen panes override behavior
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def create_test_frozen_panes_excel():
    """Create a test Excel file with frozen panes and multiple data sections"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Frozen Panes Test"
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    section_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Add main header (frozen rows)
    ws['A1'] = "Main Header"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    
    for col in range(2, 9):
        ws[f'{get_column_letter(col)}1'] = f"Column {col-1}"
        ws[f'{get_column_letter(col)}1'].font = header_font
        ws[f'{get_column_letter(col)}1'].fill = header_fill
    
    # Add sub-header (frozen rows)
    ws['A2'] = "Sub Header"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    
    for col in range(2, 9):
        ws[f'{get_column_letter(col)}2'] = f"Sub {col-1}"
        ws[f'{get_column_letter(col)}2'].font = header_font
        ws[f'{get_column_letter(col)}2'].fill = header_fill
    
    # Add data rows
    for row in range(3, 6):
        ws[f'A{row}'] = f"Row {row-2}"
        for col in range(2, 9):
            ws[f'{get_column_letter(col)}{row}'] = f"Data {row-2}-{col-1}"
    
    # Add large gap (rows 6-10 empty) - this would normally trigger table detection
    
    # Add second section with different formatting
    ws['A11'] = "Second Section"
    ws['A11'].font = header_font
    ws['A11'].fill = section_fill
    
    for col in range(2, 6):
        ws[f'{get_column_letter(col)}11'] = f"Sec {col-1}"
        ws[f'{get_column_letter(col)}11'].font = header_font
        ws[f'{get_column_letter(col)}11'].fill = section_fill
    
    for row in range(12, 15):
        ws[f'A{row}'] = f"Sec Row {row-11}"
        for col in range(2, 6):
            ws[f'{get_column_letter(col)}{row}'] = f"Sec Data {row-11}-{col-1}"
    
    # Add another gap (rows 15-17 empty)
    
    # Add third section
    ws['A18'] = "Third Section"
    ws['A18'].font = header_font
    
    for col in range(2, 8):
        ws[f'{get_column_letter(col)}18'] = f"Third {col-1}"
        ws[f'{get_column_letter(col)}18'].font = header_font
    
    for row in range(19, 22):
        ws[f'A{row}'] = f"Third Row {row-18}"
        for col in range(2, 8):
            ws[f'{get_column_letter(col)}{row}'] = f"Third Data {row-18}-{col-1}"
    
    # Add final data extending to the end
    for row in range(22, 25):
        ws[f'A{row}'] = f"Final Row {row-21}"
        for col in range(2, 9):
            ws[f'{get_column_letter(col)}{row}'] = f"Final Data {row-21}-{col-1}"
    
    # Set frozen panes to freeze first 2 rows and first 1 column (should be "B3")
    ws.freeze_panes = "B3"
    
    # Save the file
    filename = "test_frozen_panes_override.xlsx"
    wb.save(filename)
    print(f"Created test file: {filename}")
    print(f"Frozen panes set to: {ws.freeze_panes}")
    print("This file has:")
    print("- 2 frozen rows and 1 frozen column")
    print("- Large gaps that would normally trigger table detection")
    print("- Multiple sections with different formatting")
    print("- Data extending to the end")
    print("Expected behavior: Should create exactly ONE table covering the entire sheet")


if __name__ == "__main__":
    create_test_frozen_panes_excel() 