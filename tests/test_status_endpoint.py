import json
import os
import tempfile
import django
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'excel_converter.settings')
django.setup()


class StatusEndpointTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tmpdir = tempfile.TemporaryDirectory()
        os.environ['STORAGE_BACKEND'] = 'local'
        os.environ['LOCAL_STORAGE_PATH'] = self.tmpdir.name
        os.environ['USE_STORAGE_SERVICE'] = 'true'

    def tearDown(self):
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass

    def test_status_returns_record_for_excel(self):
        # Create small xlsx
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws['A1'] = 'H'; ws['A2'] = 'V'
        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(f.name)
        with open(f.name, 'rb') as r:
            xlsx_bytes = r.read()
        os.unlink(f.name)

        upload = SimpleUploadedFile('t.xlsx', xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp = self.client.post('/api/upload/', data={'file': upload})
        self.assertEqual(resp.status_code, 200)
        storage = resp.json().get('storage')
        self.assertIsNotNone(storage)
        pid = storage['processing_id']

        resp = self.client.get(f'/api/status/{pid}/')
        self.assertEqual(resp.status_code, 200)
        record = resp.json()
        self.assertEqual(record['processing_id'], pid)
        self.assertIn('storage', record)


