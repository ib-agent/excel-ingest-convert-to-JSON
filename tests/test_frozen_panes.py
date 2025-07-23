#!/usr/bin/env python3
"""
Test script to verify frozen panes parsing and preservation
"""

import sys
import os
import unittest
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from converter.excel_processor import ExcelProcessor
from converter.table_processor import TableProcessor


class TestFrozenPanes(unittest.TestCase):
    """Test cases for frozen panes functionality"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.excel_processor = ExcelProcessor()
        self.table_processor = TableProcessor()
    
    def test_frozen_panes_parsing(self):
        """Test that frozen panes are correctly parsed from Excel files"""
        # Create a test Excel file with frozen panes
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # Create a workbook with frozen panes
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some data
        for row in range(1, 6):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.value = f"Data {row}-{col}"
        
        # Set frozen panes to freeze first 2 rows and first 1 column (should be "B3")
        ws.freeze_panes = "B3"
        
        # Save the file
        test_file = "test_frozen_panes.xlsx"
        wb.save(test_file)
        
        try:
            # Process the file
            result = self.excel_processor.process_file(test_file)
            
            # Check that frozen panes are correctly extracted
            self.assertIn('workbook', result)
            self.assertIn('sheets', result['workbook'])
            self.assertGreater(len(result['workbook']['sheets']), 0)
            
            sheet = result['workbook']['sheets'][0]
            self.assertIn('frozen_panes', sheet)
            
            frozen_panes = sheet['frozen_panes']
            self.assertIn('frozen_rows', frozen_panes)
            self.assertIn('frozen_cols', frozen_panes)
            
            # "B3" should result in frozen_rows=2, frozen_cols=1
            self.assertEqual(frozen_panes['frozen_rows'], 2, 
                           f"Expected frozen_rows=2, got {frozen_panes['frozen_rows']}")
            self.assertEqual(frozen_panes['frozen_cols'], 1, 
                           f"Expected frozen_cols=1, got {frozen_panes['frozen_cols']}")
            
            # Check top_left_cell
            self.assertIn('top_left_cell', frozen_panes)
            self.assertEqual(frozen_panes['top_left_cell'], 'B3')
            
        finally:
            # Clean up
            import os
            if os.path.exists(test_file):
                os.remove(test_file)
    
    def test_frozen_panes_preservation(self):
        """Test that frozen panes are preserved during table transformation"""
        # Test data with frozen panes
        test_data = {
            'workbook': {
                'sheets': [{
                    'name': 'Test Sheet',
                    'dimensions': {
                        'min_row': 1,
                        'max_row': 5,
                        'min_col': 1,
                        'max_col': 4
                    },
                    'frozen_panes': {
                        'frozen_rows': 2,
                        'frozen_cols': 1,
                        'top_left_cell': 'B3'
                    },
                    'cells': {
                        'A1': {'value': 'Header1'},
                        'B1': {'value': 'Header2'},
                        'C1': {'value': 'Header3'},
                        'D1': {'value': 'Header4'},
                        'A2': {'value': 'SubHeader1'},
                        'B2': {'value': 'SubHeader2'},
                        'C2': {'value': 'SubHeader3'},
                        'D2': {'value': 'SubHeader4'},
                        'A3': {'value': 'Data1'},
                        'B3': {'value': 'Data2'},
                        'C3': {'value': 'Data3'},
                        'D3': {'value': 'Data4'},
                        'A4': {'value': 'Data5'},
                        'B4': {'value': 'Data6'},
                        'C4': {'value': 'Data7'},
                        'D4': {'value': 'Data8'},
                        'A5': {'value': 'Data9'},
                        'B5': {'value': 'Data10'},
                        'C5': {'value': 'Data11'},
                        'D5': {'value': 'Data12'}
                    },
                    'tables': []
                }]
            }
        }
        
        # Process the test data
        result = self.table_processor.transform_to_table_format(test_data)
        
        # Verify that frozen panes are preserved at sheet level
        sheet = result['workbook']['sheets'][0]
        self.assertIn('frozen_panes', sheet, "Frozen panes not preserved at sheet level")
        
        frozen_panes = sheet['frozen_panes']
        self.assertEqual(frozen_panes['frozen_rows'], 2, 
                        "Frozen rows not preserved correctly")
        self.assertEqual(frozen_panes['frozen_cols'], 1, 
                        "Frozen columns not preserved correctly")
        self.assertEqual(frozen_panes['top_left_cell'], 'B3', 
                        "Top left cell not preserved correctly")
        
        # Verify that tables are created
        self.assertIn('tables', sheet)
        self.assertGreater(len(sheet['tables']), 0, "No tables created")
        
        # Verify that frozen panes are NOT in table regions
        table = sheet['tables'][0]
        self.assertNotIn('frozen_rows', table['region'], 
                        "Frozen rows should not be in table region")
        self.assertNotIn('frozen_cols', table['region'], 
                        "Frozen columns should not be in table region")
    
    def test_different_frozen_pane_formats(self):
        """Test parsing of different frozen pane formats"""
        # Test data with different frozen pane configurations
        test_cases = [
            {
                'frozen_panes': {'frozen_rows': 1, 'frozen_cols': 0, 'top_left_cell': 'A2'},
                'description': 'Only first row frozen'
            },
            {
                'frozen_panes': {'frozen_rows': 0, 'frozen_cols': 1, 'top_left_cell': 'B1'},
                'description': 'Only first column frozen'
            },
            {
                'frozen_panes': {'frozen_rows': 3, 'frozen_cols': 2, 'top_left_cell': 'C4'},
                'description': 'Multiple rows and columns frozen'
            },
            {
                'frozen_panes': {'frozen_rows': 0, 'frozen_cols': 0, 'top_left_cell': None},
                'description': 'No frozen panes'
            }
        ]
        
        for test_case in test_cases:
            with self.subTest(test_case['description']):
                test_data = {
                    'workbook': {
                        'sheets': [{
                            'name': 'Test Sheet',
                            'dimensions': {'min_row': 1, 'max_row': 5, 'min_col': 1, 'max_col': 4},
                            'frozen_panes': test_case['frozen_panes'],
                            'cells': {'A1': {'value': 'Test'}},
                            'tables': []
                        }]
                    }
                }
                
                # Process the test data
                result = self.table_processor.transform_to_table_format(test_data)
                
                # Verify frozen panes are preserved
                sheet = result['workbook']['sheets'][0]
                self.assertIn('frozen_panes', sheet)
                
                frozen_panes = sheet['frozen_panes']
                expected = test_case['frozen_panes']
                
                self.assertEqual(frozen_panes['frozen_rows'], expected['frozen_rows'],
                               f"Frozen rows mismatch for {test_case['description']}")
                self.assertEqual(frozen_panes['frozen_cols'], expected['frozen_cols'],
                               f"Frozen columns mismatch for {test_case['description']}")
                
                if expected['top_left_cell']:
                    self.assertEqual(frozen_panes['top_left_cell'], expected['top_left_cell'],
                                   f"Top left cell mismatch for {test_case['description']}")


def run_tests():
    """Run the test suite"""
    print("=== RUNNING FROZEN PANES TESTS ===")
    unittest.main(verbosity=2, exit=False)


if __name__ == "__main__":
    run_tests() 