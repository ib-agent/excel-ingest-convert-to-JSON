#!/usr/bin/env python3
"""
Create a test Excel file WITHOUT frozen panes to verify that other
table detection methods are used when frozen panes are not present
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def create_test_no_frozen_panes_excel():
    """Create a test Excel file without frozen panes and multiple data sections"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "No Frozen Panes Test"
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    section_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # First table
    ws['A1'] = "Table 1 Header"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    
    for col in range(2, 6):
        ws[f'{get_column_letter(col)}1'] = f"Col {col-1}"
        ws[f'{get_column_letter(col)}1'].font = header_font
        ws[f'{get_column_letter(col)}1'].fill = header_fill
    
    for row in range(2, 5):
        ws[f'A{row}'] = f"Row {row-1}"
        for col in range(2, 6):
            ws[f'{get_column_letter(col)}{row}'] = f"Data {row-1}-{col-1}"
    
    # Large gap (rows 5-8 empty) - this should trigger table detection
    
    # Second table with different formatting
    ws['A9'] = "Table 2 Header"
    ws['A9'].font = header_font
    ws['A9'].fill = section_fill
    
    for col in range(2, 7):
        ws[f'{get_column_letter(col)}9'] = f"T2 Col {col-1}"
        ws[f'{get_column_letter(col)}9'].font = header_font
        ws[f'{get_column_letter(col)}9'].fill = section_fill
    
    for row in range(10, 13):
        ws[f'A{row}'] = f"T2 Row {row-9}"
        for col in range(2, 7):
            ws[f'{get_column_letter(col)}{row}'] = f"T2 Data {row-9}-{col-1}"
    
    # Another gap (rows 13-15 empty)
    
    # Third table
    ws['A16'] = "Table 3 Header"
    ws['A16'].font = header_font
    
    for col in range(2, 5):
        ws[f'{get_column_letter(col)}16'] = f"T3 Col {col-1}"
        ws[f'{get_column_letter(col)}16'].font = header_font
    
    for row in range(17, 20):
        ws[f'A{row}'] = f"T3 Row {row-16}"
        for col in range(2, 5):
            ws[f'{get_column_letter(col)}{row}'] = f"T3 Data {row-16}-{col-1}"
    
    # Save the file (NO frozen panes set)
    filename = "test_no_frozen_panes.xlsx"
    wb.save(filename)
    print(f"Created test file: {filename}")
    print("This file has:")
    print("- NO frozen panes")
    print("- Large gaps that should trigger table detection")
    print("- Multiple sections with different formatting")
    print("Expected behavior: Should create MULTIPLE tables (3 tables)")


if __name__ == "__main__":
    create_test_no_frozen_panes_excel() 