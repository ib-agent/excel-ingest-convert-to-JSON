"""
Complexity-Preserving Compact Excel Processor

This processor extends the compact Excel processor to preserve data needed
for complexity analysis while maintaining excellent compression through RLE.
"""

import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from .compact_excel_processor import CompactExcelProcessor
from .excel_processor import ExcelProcessor


class ComplexityPreservingCompactProcessor(CompactExcelProcessor):
    """
    Enhanced compact processor that preserves complexity-relevant data
    while maintaining RLE compression benefits.
    """
    
    def __init__(self, 
                 enable_rle: bool = True,
                 rle_min_run_length: int = 3,
                 rle_max_row_width: int = 20,
                 rle_aggressive_none: bool = True,
                 preserve_complexity_data: bool = True):
        """
        Initialize with complexity preservation options
        
        Args:
            preserve_complexity_data: Whether to preserve data needed for complexity analysis
            Other args: Same as parent CompactExcelProcessor
        """
        super().__init__(enable_rle, rle_min_run_length, rle_max_row_width, rle_aggressive_none)
        self.preserve_complexity_data = preserve_complexity_data
        self.complexity_metadata = {}
    
    def process_file(self, file_path: str, 
                    filter_empty_trailing: bool = True,
                    include_complexity_metadata: bool = True) -> Dict[str, Any]:
        """
        Process Excel file with complexity preservation
        
        Args:
            file_path: Path to Excel file
            filter_empty_trailing: Apply selective filtering
            include_complexity_metadata: Include metadata for complexity analysis
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # First, get complexity metadata using full processor if needed
        if include_complexity_metadata and self.preserve_complexity_data:
            self.complexity_metadata = self._extract_complexity_metadata(file_path)
        
        # Process normally with compact processor
        result = super().process_file(file_path, filter_empty_trailing)
        
        # Enhance result with complexity metadata
        if include_complexity_metadata and self.preserve_complexity_data:
            result['complexity_metadata'] = self.complexity_metadata
        
        return result
    
    def _extract_complexity_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata needed for complexity analysis using full processor
        
        This preserves key complexity indicators without storing all the data.
        """
        print("🔍 Extracting complexity metadata...")
        
        # Use full processor to get complete data for analysis
        full_processor = ExcelProcessor()
        full_data = full_processor.process_file(file_path)
        
        complexity_metadata = {
            'extraction_timestamp': datetime.now().isoformat(),
            'sheets': {}
        }
        
        for sheet in full_data.get('workbook', {}).get('sheets', []):
            sheet_name = sheet.get('name', 'Unknown')
            sheet_metadata = self._extract_sheet_complexity_metadata(sheet)
            complexity_metadata['sheets'][sheet_name] = sheet_metadata
        
        print(f"✅ Extracted complexity metadata for {len(complexity_metadata['sheets'])} sheets")
        return complexity_metadata
    
    def _extract_sheet_complexity_metadata(self, sheet: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract complexity-relevant metadata from a single sheet
        
        This captures the essence of complexity without storing all cell data.
        """
        cells = sheet.get('cells', {})
        dimensions = sheet.get('dimensions', {})
        merged_cells = sheet.get('merged_cells', [])
        
        # 1. Merged Cell Analysis
        merged_metadata = {
            'count': len(merged_cells),
            'complex_merges': 0,
            'patterns': []
        }
        
        for merge in merged_cells:
            start_row = merge.get('start_row', 0) or 0
            end_row = merge.get('end_row', 0) or 0
            start_col = merge.get('start_column', 0) or 0
            end_col = merge.get('end_column', 0) or 0
            
            rows_spanned = end_row - start_row + 1
            cols_spanned = end_col - start_col + 1
            
            if rows_spanned > 2 or cols_spanned > 2:
                merged_metadata['complex_merges'] += 1
            
            # Store pattern for analysis
            merged_metadata['patterns'].append({
                'start_row': start_row,
                'start_col': start_col,
                'rows_spanned': rows_spanned,
                'cols_spanned': cols_spanned
            })
        
        # 2. Header Structure Analysis
        header_metadata = self._analyze_header_patterns(cells, dimensions)
        
        # 3. Data Distribution Analysis
        distribution_metadata = self._analyze_data_distribution(cells, dimensions)
        
        # 4. Formula Analysis
        formula_metadata = self._analyze_formula_patterns(cells)
        
        # 5. Column Structure Analysis
        column_metadata = self._analyze_column_patterns(cells, dimensions)
        
        return {
            'dimensions': dimensions,
            'cell_count': len(cells),
            'merged_cells': merged_metadata,
            'header_structure': header_metadata,
            'data_distribution': distribution_metadata,
            'formulas': formula_metadata,
            'columns': column_metadata
        }
    
    def _analyze_header_patterns(self, cells: Dict[str, Any], 
                                dimensions: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze header patterns for complexity assessment"""
        max_row_check = min(dimensions.get('max_row', 10), 10)
        min_col = dimensions.get('min_col', 1)
        max_col = dimensions.get('max_col', 1)
        
        header_rows = []
        inconsistencies = 0
        
        # Check first few rows for header patterns
        for row in range(1, max_row_check + 1):
            row_cells = []
            gaps = 0
            
            for col in range(min_col, max_col + 1):
                from openpyxl.utils import get_column_letter
                coord = f"{get_column_letter(col)}{row}"
                
                if coord in cells:
                    value = cells[coord].get('value')
                    if value is not None:
                        row_cells.append({
                            'col': col,
                            'value': str(value),
                            'type': type(value).__name__
                        })
                else:
                    gaps += 1
            
            if row_cells:
                # Analyze if this looks like a header row
                text_ratio = sum(1 for cell in row_cells 
                               if isinstance(cell.get('value'), str)) / len(row_cells)
                
                header_rows.append({
                    'row': row,
                    'cell_count': len(row_cells),
                    'gaps': gaps,
                    'text_ratio': text_ratio,
                    'is_header_like': text_ratio > 0.7
                })
                
                # Count inconsistencies
                if gaps > 0:
                    gap_ratio = gaps / (max_col - min_col + 1)
                    inconsistencies += gap_ratio
        
        # Detect header levels
        header_levels = len([hr for hr in header_rows if hr['is_header_like']])
        
        return {
            'detected_levels': header_levels,
            'inconsistency_score': min(inconsistencies, 1.0),
            'header_rows': header_rows[:5]  # Keep first 5 for analysis
        }
    
    def _analyze_data_distribution(self, cells: Dict[str, Any], 
                                  dimensions: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze data distribution patterns"""
        if not cells or not dimensions:
            return {'sparsity': 0.0, 'clusters': []}
        
        # Calculate sparsity
        total_addressable = (
            (dimensions.get('max_row', 1) - dimensions.get('min_row', 1) + 1) *
            (dimensions.get('max_col', 1) - dimensions.get('min_col', 1) + 1)
        )
        
        data_cells = len([cell for cell in cells.values() 
                         if cell.get('value') is not None and 
                         str(cell.get('value', '')).strip()])
        
        sparsity = 1.0 - (data_cells / total_addressable) if total_addressable > 0 else 0.0
        
        # Analyze data clustering (simplified)
        row_distribution = {}
        col_distribution = {}
        
        for coord, cell in cells.items():
            if cell.get('value') is not None:
                # Extract row and column
                import re
                row_match = re.search(r'\d+', coord)
                col_match = re.search(r'[A-Z]+', coord)
                
                if row_match:
                    row_num = int(row_match.group())
                    row_distribution[row_num] = row_distribution.get(row_num, 0) + 1
                
                if col_match:
                    col_str = col_match.group()
                    col_distribution[col_str] = col_distribution.get(col_str, 0) + 1
        
        # Find largest gaps (indicating sparse regions)
        row_gaps = self._find_gaps(sorted(row_distribution.keys()))
        col_gaps = self._find_gaps_alpha(sorted(col_distribution.keys()))
        
        return {
            'sparsity': sparsity,
            'total_addressable': total_addressable,
            'data_cells': data_cells,
            'row_gaps': row_gaps[:5],  # Top 5 largest gaps
            'col_gaps': col_gaps[:5],
            'clustered': len(row_gaps) > 3 or len(col_gaps) > 3  # Multiple gaps suggest clustering
        }
    
    def _analyze_formula_patterns(self, cells: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze formula complexity patterns"""
        formula_count = 0
        complex_formula_count = 0
        formula_types = {}
        
        for cell in cells.values():
            formula = cell.get('formula')
            if formula:
                formula_count += 1
                
                # Categorize formula complexity
                if self._is_complex_formula(formula):
                    complex_formula_count += 1
                
                # Track formula types
                formula_type = self._categorize_formula(formula)
                formula_types[formula_type] = formula_types.get(formula_type, 0) + 1
        
        total_cells = len(cells)
        
        return {
            'formula_count': formula_count,
            'complex_formula_count': complex_formula_count,
            'formula_ratio': formula_count / total_cells if total_cells > 0 else 0,
            'complex_ratio': complex_formula_count / formula_count if formula_count > 0 else 0,
            'formula_types': formula_types
        }
    
    def _analyze_column_patterns(self, cells: Dict[str, Any], 
                                dimensions: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze column consistency patterns"""
        min_col = dimensions.get('min_col', 1)
        max_col = dimensions.get('max_col', 1)
        min_row = dimensions.get('min_row', 1)
        max_row = dimensions.get('max_row', 1)
        
        column_analysis = {}
        
        from openpyxl.utils import get_column_letter
        
        for col in range(min_col, min(max_col + 1, min_col + 20)):  # Limit to first 20 cols
            col_letter = get_column_letter(col)
            col_values = []
            
            for row in range(min_row, max_row + 1):
                coord = f"{col_letter}{row}"
                if coord in cells:
                    value = cells[coord].get('value')
                    if value is not None:
                        col_values.append(value)
            
            if len(col_values) > 1:
                inconsistency = self._calculate_column_inconsistency(col_values)
                column_analysis[col_letter] = {
                    'value_count': len(col_values),
                    'inconsistency': inconsistency,
                    'types': self._get_value_types(col_values)
                }
        
        # Calculate overall inconsistency
        inconsistencies = [col['inconsistency'] for col in column_analysis.values()]
        avg_inconsistency = sum(inconsistencies) / len(inconsistencies) if inconsistencies else 0
        
        return {
            'column_count': len(column_analysis),
            'average_inconsistency': avg_inconsistency,
            'high_inconsistency_columns': len([i for i in inconsistencies if i > 0.3])
        }
    
    def _find_gaps(self, sorted_numbers: List[int]) -> List[int]:
        """Find gaps in a sorted list of numbers"""
        gaps = []
        for i in range(1, len(sorted_numbers)):
            gap = sorted_numbers[i] - sorted_numbers[i-1] - 1
            if gap > 1:
                gaps.append(gap)
        return sorted(gaps, reverse=True)
    
    def _find_gaps_alpha(self, sorted_cols: List[str]) -> List[int]:
        """Find gaps in alphabetical column sequence"""
        # Simplified - just count missing columns
        if len(sorted_cols) < 2:
            return []
        
        from openpyxl.utils import column_index_from_string
        try:
            indices = [column_index_from_string(col) for col in sorted_cols]
            return self._find_gaps(sorted(indices))
        except:
            return []
    
    def _is_complex_formula(self, formula: str) -> bool:
        """Check if formula is complex (same as in complexity analyzer)"""
        if not formula:
            return False
        
        import re
        complex_patterns = [
            r'IF.*IF',      # Nested IFs
            r'VLOOKUP',     # Lookup functions
            r'INDEX.*MATCH', # Index/Match combinations
            r'SUMPRODUCT',  # Array functions
            r'\$[A-Z]+\$\d+:\$[A-Z]+\$\d+', # Absolute ranges
        ]
        
        for pattern in complex_patterns:
            if re.search(pattern, formula, re.IGNORECASE):
                return True
        
        return False
    
    def _categorize_formula(self, formula: str) -> str:
        """Categorize formula type"""
        if not formula:
            return 'empty'
        
        formula_upper = formula.upper()
        
        if 'SUM' in formula_upper:
            return 'sum'
        elif 'IF' in formula_upper:
            return 'conditional'
        elif any(lookup in formula_upper for lookup in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
            return 'lookup'
        elif any(ref in formula_upper for ref in ['$', ':']):
            return 'reference'
        else:
            return 'other'
    
    def _calculate_column_inconsistency(self, col_values: List[Any]) -> float:
        """Calculate inconsistency score for column values"""
        if len(col_values) <= 1:
            return 0.0
        
        # Analyze data types
        type_counts = {}
        for value in col_values:
            value_type = type(value).__name__
            type_counts[value_type] = type_counts.get(value_type, 0) + 1
        
        # Calculate inconsistency based on type diversity
        if len(type_counts) == 1:
            return 0.0  # All same type = consistent
        
        # More types = more inconsistency
        inconsistency = (len(type_counts) - 1) / len(col_values)
        return min(inconsistency, 1.0)
    
    def _get_value_types(self, values: List[Any]) -> Dict[str, int]:
        """Get count of value types in a list"""
        type_counts = {}
        for value in values:
            value_type = type(value).__name__
            type_counts[value_type] = type_counts.get(value_type, 0) + 1
        return type_counts
    
    def _filter_empty_trailing_areas(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Enhanced filtering that preserves complexity-relevant data
        
        This is more selective than the original filtering - it only removes
        truly empty trailing areas while preserving data needed for complexity analysis.
        """
        if not self.preserve_complexity_data:
            # Use original filtering if complexity preservation is disabled
            return super()._filter_empty_trailing_areas(result)
        
        # Apply selective filtering that preserves complexity indicators
        return self._selective_filter_empty_areas(result)
    
    def _selective_filter_empty_areas(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Selectively filter empty areas while preserving complexity indicators
        
        This method is more conservative than the original filtering.
        """
        if 'workbook' not in result or 'sheets' not in result['workbook']:
            return result
        
        filtered_result = result.copy()
        filtered_sheets = []
        
        for sheet in result['workbook']['sheets']:
            # Check if we have complexity metadata for this sheet
            sheet_name = sheet.get('name', 'Unknown')
            has_complexity_metadata = (
                'complexity_metadata' in result and 
                'sheets' in result['complexity_metadata'] and
                sheet_name in result['complexity_metadata']['sheets']
            )
            
            if has_complexity_metadata:
                # Use metadata to inform filtering decisions
                complexity_data = result['complexity_metadata']['sheets'][sheet_name]
                filtered_sheet = self._selective_filter_sheet(sheet, complexity_data)
            else:
                # Conservative filtering without metadata
                filtered_sheet = self._conservative_filter_sheet(sheet)
            
            filtered_sheets.append(filtered_sheet)
        
        filtered_result['workbook']['sheets'] = filtered_sheets
        return filtered_result
    
    def _selective_filter_sheet(self, sheet: Dict[str, Any], 
                               complexity_data: Dict[str, Any]) -> Dict[str, Any]:
        """Filter sheet using complexity metadata to guide decisions"""
        
        # If the sheet has high sparsity or complex patterns, be more conservative
        distribution = complexity_data.get('data_distribution', {})
        sparsity = distribution.get('sparsity', 0.0)
        is_clustered = distribution.get('clustered', False)
        
        if sparsity > 0.7 or is_clustered:
            # High sparsity or clustering - very conservative filtering
            return self._conservative_filter_sheet(sheet)
        else:
            # Lower sparsity - can apply more aggressive filtering
            return self._moderate_filter_sheet(sheet)
    
    def _conservative_filter_sheet(self, sheet: Dict[str, Any]) -> Dict[str, Any]:
        """Very conservative filtering - only remove completely empty trailing rows/cols"""
        if 'rows' not in sheet:
            return sheet
        
        rows = sheet['rows']
        if not rows:
            return sheet
        
        # Only filter if there are truly empty trailing areas (no data at all)
        # Find last row and column with any data
        max_data_row = 0
        max_data_col = 0
        
        for row in rows:
            row_num = row.get('r', 0)
            cells = row.get('cells', [])
            
            if cells:  # If row has any cells at all, keep it
                max_data_row = max(max_data_row, row_num)
                
                for cell in cells:
                    if len(cell) >= 1:
                        if self._is_rle_cell(cell):
                            start_col = cell[0]
                            run_length = cell[-1]
                            end_col = start_col + run_length - 1
                            max_data_col = max(max_data_col, end_col)
                        else:
                            max_data_col = max(max_data_col, cell[0])
        
        # Very minimal filtering - only remove completely empty areas
        filtered_rows = []
        for row in rows:
            row_num = row.get('r', 0)
            if row_num <= max_data_row + 5:  # Keep extra buffer
                filtered_rows.append(row)
        
        filtered_sheet = sheet.copy()
        filtered_sheet['rows'] = filtered_rows
        
        # Update dimensions conservatively
        if max_data_row > 0 and max_data_col > 0:
            original_dimensions = sheet.get('dimensions', [1, 1, 1, 1])
            filtered_sheet['dimensions'] = [
                original_dimensions[0],
                original_dimensions[1],
                max(max_data_row + 5, original_dimensions[2]),  # Extra buffer
                max(max_data_col + 5, original_dimensions[3])   # Extra buffer
            ]
        
        return filtered_sheet
    
    def _moderate_filter_sheet(self, sheet: Dict[str, Any]) -> Dict[str, Any]:
        """Moderate filtering - original approach but with some preservation"""
        # Use the original filtering method but with less aggressive trimming
        return super()._filter_sheet_empty_trailing_areas(sheet)
