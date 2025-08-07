import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


class CompactExcelProcessor:
    """Compact Excel to JSON converter with Run-Length Encoding support"""
    
    def __init__(self, 
                 enable_rle: bool = True,
                 rle_min_run_length: int = 3,
                 rle_max_row_width: int = 20,
                 rle_aggressive_none: bool = True):
        """
        Initialize processor with RLE configuration
        
        Args:
            enable_rle: Enable run-length encoding for repeated values
            rle_min_run_length: Minimum consecutive cells to apply RLE (default: 3)
            rle_max_row_width: Only apply RLE to rows with more than this many cells
            rle_aggressive_none: Use more aggressive compression for None values (min_run=2)
        """
        self.supported_extensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']
        self.style_registry = {}  # Central style dictionary
        self.next_style_id = 1
        
        # RLE Configuration
        self.enable_rle = enable_rle
        self.rle_min_run_length = rle_min_run_length
        self.rle_max_row_width = rle_max_row_width
        self.rle_aggressive_none = rle_aggressive_none
        
        # RLE Statistics
        self.rle_stats = {
            "rows_compressed": 0,
            "cells_before_rle": 0,
            "cells_after_rle": 0,
            "runs_created": 0
        }
    
    def process_file(self, file_path: str, filter_empty_trailing: bool = True) -> Dict[str, Any]:
        """
        Process an Excel file and return compact JSON representation
        
        Args:
            file_path: Path to the Excel file
            filter_empty_trailing: Whether to filter out empty trailing rows and columns
            
        Returns:
            Dictionary containing the compact Excel structure
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension not in self.supported_extensions:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
        # Reset statistics
        self.rle_stats = {
            "rows_compressed": 0,
            "cells_before_rle": 0,
            "cells_after_rle": 0,
            "runs_created": 0
        }
        
        try:
            # First pass: get all data including formulas (data_only=False)
            workbook_with_formulas = load_workbook(file_path, data_only=False, keep_vba=True)
            
            # Second pass: get calculated values (data_only=True)  
            workbook_with_values = load_workbook(file_path, data_only=True, keep_vba=True)
            
            result = self._process_workbook(workbook_with_formulas, workbook_with_values, file_path)
            
            # Apply empty trailing filtering if requested
            if filter_empty_trailing:
                result = self._filter_empty_trailing_areas(result)
            
            # Add RLE statistics to result if compression was applied
            if self.enable_rle and self.rle_stats["runs_created"] > 0:
                result["rle_compression"] = self.rle_stats.copy()
                compression_ratio = (
                    (self.rle_stats["cells_before_rle"] - self.rle_stats["cells_after_rle"]) 
                    / max(self.rle_stats["cells_before_rle"], 1) * 100
                )
                result["rle_compression"]["compression_ratio_percent"] = round(compression_ratio, 2)
            
            return result
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def _process_workbook(self, workbook_with_formulas: Workbook, workbook_with_values: Workbook, file_path: str) -> Dict[str, Any]:
        """Process the entire workbook with both formula and value data"""
        file_stats = os.stat(file_path)
        
        # Reset style registry for each workbook
        self.style_registry = {}
        self.next_style_id = 1
        
        result = {
            "workbook": {
                "meta": self._extract_compact_metadata(workbook_with_formulas, file_path, file_stats),
                "styles": {},  # Will be populated during processing
                "sheets": []   # Array of sheet objects
            }
        }
        
        # Process each worksheet
        for sheet_name in workbook_with_formulas.sheetnames:
            worksheet_with_formulas = workbook_with_formulas[sheet_name]
            worksheet_with_values = workbook_with_values[sheet_name]
            
            # Check if it's a worksheet (not a chart sheet)
            if hasattr(worksheet_with_formulas, 'title') and not hasattr(worksheet_with_formulas, 'chart'):
                sheet_data = self._process_worksheet(worksheet_with_formulas, worksheet_with_values)
                result["workbook"]["sheets"].append(sheet_data)
        
        # Add the style registry to the result
        result["workbook"]["styles"] = self.style_registry
        
        # Generate table_data for compatibility with UI
        result["table_data"] = self._generate_table_data(result["workbook"])
        
        return result
    
    def _extract_compact_metadata(self, workbook: Workbook, file_path: str, file_stats) -> Dict[str, Any]:
        """Extract essential workbook metadata only"""
        props = workbook.properties
        
        metadata = {}
        
        # Only include non-empty/meaningful metadata
        if props.creator:
            metadata["creator"] = props.creator
        if props.created:
            metadata["created"] = self._format_datetime(props.created)
        if props.modified:
            metadata["modified"] = self._format_datetime(props.modified)
        
        metadata["filename"] = os.path.basename(file_path)
        
        return metadata
    
    def _process_worksheet(self, worksheet_with_formulas: Worksheet, worksheet_with_values: Worksheet) -> Dict[str, Any]:
        """Process a single worksheet with compact format"""
        sheet_data = {
            "name": worksheet_with_formulas.title
        }
        
        # Add optional properties only if they have meaningful values
        dimensions = self._extract_dimensions(worksheet_with_formulas)
        if dimensions:
            sheet_data["dimensions"] = [
                dimensions["min_row"],
                dimensions["min_col"], 
                dimensions["max_row"],
                dimensions["max_col"]
            ]
        
        frozen_panes = self._extract_frozen_panes(worksheet_with_formulas)
        if frozen_panes["frozen_rows"] > 0 or frozen_panes["frozen_cols"] > 0:
            sheet_data["frozen"] = [frozen_panes["frozen_rows"], frozen_panes["frozen_cols"]]
        
        merged_cells = self._extract_compact_merged_cells(worksheet_with_formulas)
        if merged_cells:
            sheet_data["merged"] = merged_cells
        
        # Process cells in row-based format with optional RLE compression
        rows = self._extract_compact_cell_data_with_rle(worksheet_with_formulas, worksheet_with_values)
        if rows:
            sheet_data["rows"] = rows
        
        return sheet_data
    
    def _extract_compact_cell_data_with_rle(self, worksheet_with_formulas: Worksheet, worksheet_with_values: Worksheet) -> List[Dict[str, Any]]:
        """Extract cell data in compact row-based format with optional RLE compression"""
        rows_data = {}  # Dictionary to organize by row number
        
        for row in worksheet_with_formulas.iter_rows():
            for cell in row:
                # Get formula from the worksheet with formulas
                formula = cell.value if getattr(cell, 'data_type', None) == 'f' else None
                
                # Get calculated value from the worksheet with values
                value_cell = worksheet_with_values[cell.coordinate]
                calculated_value = self._extract_cell_value(value_cell) if value_cell.value is not None else None
                
                # Only process cells that have content
                if calculated_value is not None or formula or cell.comment:
                    row_num = cell.row
                    col_num = cell.column
                    
                    if row_num not in rows_data:
                        rows_data[row_num] = {"r": row_num, "cells": []}
                    
                    # Create compact cell format: [col, value, style_ref?, formula?]
                    cell_array = [col_num]
                    
                    # Add value
                    cell_array.append(calculated_value)
                    
                    # Get style reference
                    style_ref = self._get_style_reference(cell)
                    if style_ref:
                        cell_array.append(style_ref)
                    elif formula:
                        cell_array.append(None)  # Placeholder for style if formula follows
                    
                    # Add formula if present
                    if formula:
                        # Ensure we have a style slot
                        while len(cell_array) < 3:
                            cell_array.append(None)
                        cell_array.append(str(formula))
                    
                    rows_data[row_num]["cells"].append(cell_array)
        
        # Apply RLE compression to each row
        for row_num, row_data in rows_data.items():
            # Apply RLE to all rows if enabled, with more aggressive settings for wide rows
            if self.enable_rle:
                original_cell_count = len(row_data["cells"])
                
                # For extremely wide rows (>1000 columns), use more aggressive compression
                if original_cell_count > 1000:
                    # Temporarily reduce thresholds for very wide sheets
                    original_min_run = self.rle_min_run_length
                    self.rle_min_run_length = 2  # More aggressive for wide sheets
                    
                    row_data["cells"] = self._apply_rle_to_row(row_data["cells"])
                    
                    # Restore original threshold
                    self.rle_min_run_length = original_min_run
                elif original_cell_count > self.rle_max_row_width:
                    row_data["cells"] = self._apply_rle_to_row(row_data["cells"])
                
                compressed_cell_count = len(row_data["cells"])
                
                if compressed_cell_count < original_cell_count:
                    self.rle_stats["rows_compressed"] += 1
                
                self.rle_stats["cells_before_rle"] += original_cell_count
                self.rle_stats["cells_after_rle"] += compressed_cell_count
        
        # Check for row-level shared styles
        for row_num, row_data in rows_data.items():
            shared_style = self._detect_shared_row_style(row_data["cells"])
            if shared_style:
                row_data["style"] = shared_style
                # Remove style references from individual cells if they match the row style
                self._remove_redundant_cell_styles(row_data["cells"], shared_style)
        
        # Convert to sorted list
        return [rows_data[row_num] for row_num in sorted(rows_data.keys())]
    
    def _apply_rle_to_row(self, cells: List[List[Any]]) -> List[List[Any]]:
        """Apply run-length encoding to a row of cells"""
        if len(cells) < self.rle_min_run_length:
            return cells
        
        # Sort cells by column to ensure proper sequence
        sorted_cells = sorted(cells, key=lambda x: x[0])
        compressed_cells = []
        current_run = []
        
        for cell in sorted_cells:
            if self._can_extend_run(current_run, cell):
                current_run.append(cell)
            else:
                # Process completed run with dynamic thresholds
                min_run_for_value = self._get_min_run_length_for_value(current_run[0] if current_run else None)
                if len(current_run) >= min_run_for_value:
                    rle_cell = self._create_rle_cell(current_run)
                    if rle_cell:
                        compressed_cells.append(rle_cell)
                        self.rle_stats["runs_created"] += 1
                    else:
                        compressed_cells.extend(current_run)  # Keep as individual cells
                else:
                    compressed_cells.extend(current_run)  # Keep as individual cells
                current_run = [cell]
        
        # Handle final run with dynamic thresholds
        if current_run:
            min_run_for_value = self._get_min_run_length_for_value(current_run[0])
            if len(current_run) >= min_run_for_value:
                rle_cell = self._create_rle_cell(current_run)
                if rle_cell:
                    compressed_cells.append(rle_cell)
                    self.rle_stats["runs_created"] += 1
                else:
                    compressed_cells.extend(current_run)
            else:
                compressed_cells.extend(current_run)
        
        return compressed_cells
    
    def _get_min_run_length_for_value(self, cell: Optional[List[Any]]) -> int:
        """Get the minimum run length required for a specific cell value"""
        if not cell or len(cell) < 2:
            return self.rle_min_run_length
        
        value = cell[1]
        
        # Use aggressive compression for None values (very common in wide sheets)
        if self.rle_aggressive_none and value is None:
            return 2  # Compress even 2 consecutive None values
        
        return self.rle_min_run_length
    
    def _can_extend_run(self, current_run: List[List[Any]], new_cell: List[Any]) -> bool:
        """Check if a new cell can extend the current run"""
        if not current_run:
            return True
        
        last_cell = current_run[-1]
        
        # Check column continuity (must be consecutive)
        if new_cell[0] != last_cell[0] + 1:
            return False
        
        # Check value equality
        if new_cell[1] != last_cell[1]:
            return False
        
        # Check style compatibility
        last_style = last_cell[2] if len(last_cell) > 2 else None
        new_style = new_cell[2] if len(new_cell) > 2 else None
        if last_style != new_style:
            return False
        
        # Check formula compatibility
        last_formula = last_cell[3] if len(last_cell) > 3 else None
        new_formula = new_cell[3] if len(new_cell) > 3 else None
        if last_formula != new_formula:
            return False
        
        return True
    
    def _create_rle_cell(self, run_cells: List[List[Any]]) -> Optional[List[Any]]:
        """Create a run-length encoded cell from a sequence of identical cells"""
        if not run_cells:
            return None
        
        first_cell = run_cells[0]
        
        # Verify all cells are truly identical (safety check)
        base_value = first_cell[1]
        base_style = first_cell[2] if len(first_cell) > 2 else None
        base_formula = first_cell[3] if len(first_cell) > 3 else None
        
        for cell in run_cells[1:]:
            if (cell[1] != base_value or 
                (len(cell) > 2 and cell[2] != base_style) or
                (len(cell) > 3 and cell[3] != base_formula)):
                # Cells are not identical, cannot create RLE
                return None
        
        start_col = first_cell[0]
        run_length = len(run_cells)
        
        # Create RLE cell: [start_col, value, style, formula, run_length]
        rle_cell = [start_col, base_value]
        
        # Add style if present
        if base_style is not None:
            rle_cell.append(base_style)
        elif base_formula is not None:
            rle_cell.append(None)  # Style placeholder if formula exists
        
        # Add formula if present
        if base_formula is not None:
            while len(rle_cell) < 4:
                rle_cell.append(None)
            rle_cell[3] = base_formula
        
        # Add run length as final element (this is the RLE marker)
        rle_cell.append(run_length)
        
        return rle_cell
    
    def expand_rle_cells(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Expand RLE cells back to individual cells (utility method)"""
        if "cells" not in row_data:
            return row_data
        
        expanded_cells = []
        
        for cell in row_data["cells"]:
            if self._is_rle_cell(cell):
                # This is an RLE cell, expand it
                start_col, value = cell[0], cell[1]
                style = cell[2] if len(cell) > 2 else None
                formula = cell[3] if len(cell) > 3 else None
                run_length = cell[-1]  # Last element is run length
                
                # Expand the run
                for i in range(run_length):
                    expanded_cell = [start_col + i, value]
                    if style is not None:
                        expanded_cell.append(style)
                    if formula is not None:
                        while len(expanded_cell) < 4:
                            expanded_cell.append(None)
                        expanded_cell[3] = formula
                    expanded_cells.append(expanded_cell)
            else:
                # Normal cell, keep as is
                expanded_cells.append(cell)
        
        return {"r": row_data["r"], "cells": expanded_cells}
    
    def _is_rle_cell(self, cell: List[Any]) -> bool:
        """Check if a cell is RLE encoded"""
        return (len(cell) >= 5 and 
                isinstance(cell[-1], int) and 
                cell[-1] > 1)  # Run length > 1
    
    # The following methods are inherited from the original CompactExcelProcessor
    # with minimal modifications for compatibility
    
    def _extract_dimensions(self, worksheet: Worksheet) -> Dict[str, int]:
        """Extract worksheet dimensions"""
        return {
            "min_row": worksheet.min_row or 1,
            "max_row": worksheet.max_row or 1,
            "min_col": worksheet.min_column or 1,
            "max_col": worksheet.max_column or 1
        }
    
    def _extract_frozen_panes(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract frozen pane information"""
        if not worksheet.freeze_panes:
            return {"frozen_rows": 0, "frozen_cols": 0}
        
        frozen_cell = worksheet.freeze_panes
        if isinstance(frozen_cell, str):
            try:
                if ':' in frozen_cell:
                    parts = frozen_cell.split(':')
                    if len(parts) == 2:
                        col_part = parts[0]
                        row_part = parts[1]
                        col_idx = column_index_from_string(col_part)
                        row_idx = int(row_part)
                    else:
                        col_idx = 1
                        row_idx = 1
                else:
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', frozen_cell)
                    if match:
                        col_letter, row_num = match.groups()
                        col_idx = column_index_from_string(col_letter)
                        row_idx = int(row_num)
                    else:
                        col_idx = column_index_from_string(frozen_cell)
                        row_idx = 1
            except:
                col_idx = 1
                row_idx = 1
        else:
            col_idx, row_idx = frozen_cell
        
        return {
            "frozen_rows": row_idx - 1 if row_idx > 0 else 0,
            "frozen_cols": col_idx - 1 if col_idx > 0 else 0
        }
    
    def _extract_compact_merged_cells(self, worksheet: Worksheet) -> List[List[int]]:
        """Extract merged cells in compact format"""
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            start_cell = getattr(merged_range, 'start_cell', None)
            end_cell = getattr(merged_range, 'end_cell', None)
            if start_cell and end_cell:
                merged_cells.append([
                    getattr(start_cell, 'row', 1),
                    getattr(start_cell, 'column', 1),
                    getattr(end_cell, 'row', 1),
                    getattr(end_cell, 'column', 1)
                ])
        return merged_cells
    
    def _extract_cell_value(self, cell) -> Any:
        """Extract cell value with proper formatting"""
        if cell.value is None:
            return None
        
        if cell.is_date:
            return self._format_datetime(cell.value)
        
        return cell.value
    
    def _get_style_reference(self, cell) -> Optional[str]:
        """Get or create a style reference for the cell"""
        style_dict = self._extract_cell_style_dict(cell)
        
        # Skip empty styles
        if not style_dict:
            return None
        
        # Create a hashable representation of the style
        style_key = self._create_style_key(style_dict)
        
        # Check if this style already exists
        for style_id, existing_style in self.style_registry.items():
            if self._create_style_key(existing_style) == style_key:
                return style_id
        
        # Create new style reference
        style_id = f"s{self.next_style_id}"
        self.next_style_id += 1
        self.style_registry[style_id] = style_dict
        
        return style_id
    
    def _extract_cell_style_dict(self, cell) -> Dict[str, Any]:
        """Extract cell style as a dictionary"""
        style = {}
        
        # Font styling
        font_style = self._extract_font_style(cell.font)
        if font_style:
            style["font"] = font_style
        
        # Fill styling
        fill_style = self._extract_fill_style(cell.fill)
        if fill_style:
            style["fill"] = fill_style
        
        # Border styling
        border_style = self._extract_border_style(cell.border)
        if border_style:
            style["border"] = border_style
        
        # Alignment styling
        alignment_style = self._extract_alignment_style(cell.alignment)
        if alignment_style:
            style["alignment"] = alignment_style
        
        # Number format
        if cell.number_format and cell.number_format != "General":
            style["number_format"] = cell.number_format
        
        return style
    
    def _extract_font_style(self, font: Font) -> Dict[str, Any]:
        """Extract font styling"""
        if not font:
            return {}
        
        style = {}
        if font.name and font.name != "Calibri":
            style["name"] = font.name
        if font.size and font.size != 11:
            style["size"] = font.size
        if font.bold:
            style["bold"] = True
        if font.italic:
            style["italic"] = True
        if font.underline and font.underline != "none":
            style["underline"] = font.underline
        if font.strike:
            style["strike"] = True
        
        color = self._extract_color(font.color)
        if color:
            style["color"] = color
        
        return style
    
    def _extract_fill_style(self, fill: Fill) -> Dict[str, Any]:
        """Extract fill styling"""
        if not fill or fill.fill_type == "none":
            return {}
        
        style = {}
        if fill.fill_type:
            style["type"] = fill.fill_type
        
        start_color = self._extract_color(fill.start_color)
        if start_color:
            style["color"] = start_color
        
        return style
    
    def _extract_border_style(self, border: Border) -> Dict[str, Any]:
        """Extract border styling"""
        if not border:
            return {}
        
        style = {}
        for side_name in ["left", "right", "top", "bottom"]:
            side = getattr(border, side_name)
            if side and side.style:
                style[side_name] = side.style
        
        return style
    
    def _extract_alignment_style(self, alignment: Alignment) -> Dict[str, Any]:
        """Extract alignment styling"""
        if not alignment:
            return {}
        
        style = {}
        if alignment.horizontal and alignment.horizontal != "general":
            style["horizontal"] = alignment.horizontal
        if alignment.vertical and alignment.vertical != "bottom":
            style["vertical"] = alignment.vertical
        if alignment.wrap_text:
            style["wrap_text"] = True
        
        return style
    
    def _extract_color(self, color) -> Optional[str]:
        """Extract color in compact format"""
        if not color:
            return None
        
        # Handle RGB objects properly
        rgb = getattr(color, 'rgb', None)
        if rgb:
            # Convert RGB object to string if it's not already a string
            if hasattr(rgb, 'rgb'):  # openpyxl RGB object
                rgb_str = rgb.rgb if rgb.rgb else None
            elif isinstance(rgb, str):
                rgb_str = rgb
            else:
                # Try to convert to string representation
                rgb_str = str(rgb) if rgb else None
            
            if rgb_str and rgb_str != "00000000":
                return rgb_str
        
        theme = getattr(color, 'theme', None)
        if theme is not None:
            return f"theme{theme}"
        
        return None
    
    def _create_style_key(self, style_dict: Dict[str, Any]) -> str:
        """Create a hashable key for a style dictionary"""
        return json.dumps(style_dict, sort_keys=True)
    
    def _detect_shared_row_style(self, cells: List[List[Any]]) -> Optional[str]:
        """Detect if all cells in a row share the same style"""
        if not cells:
            return None
        
        # Get style references from cells, handling RLE cells
        style_refs = []
        for cell in cells:
            if self._is_rle_cell(cell):
                # RLE cell - style is at position 2
                if len(cell) > 2 and cell[2] is not None:
                    style_refs.append(cell[2])
            else:
                # Normal cell - style is at position 2
                if len(cell) > 2 and cell[2] is not None:
                    style_refs.append(cell[2])
        
        # If all cells have the same style reference, return it
        if style_refs and all(ref == style_refs[0] for ref in style_refs):
            return style_refs[0]
        
        return None
    
    def _remove_redundant_cell_styles(self, cells: List[List[Any]], shared_style: str):
        """Remove redundant style references from cells when they match the row style"""
        for cell in cells:
            if self._is_rle_cell(cell):
                # RLE cell handling
                if len(cell) > 2 and cell[2] == shared_style:
                    cell[2] = None  # Keep slot structure for RLE
            else:
                # Normal cell handling
                if len(cell) > 2 and cell[2] == shared_style:
                    # Remove the style reference
                    if len(cell) == 3:
                        cell.pop(2)  # Remove style, no formula
                    else:
                        cell[2] = None  # Keep slot for formula
    
    def _format_datetime(self, dt) -> Optional[str]:
        """Format datetime objects to ISO 8601 string"""
        if dt is None:
            return None
        
        if isinstance(dt, datetime):
            return dt.isoformat()
        
        return str(dt)
    
    def _generate_table_data(self, workbook_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate table_data structure for UI compatibility"""
        table_data = {
            "workbook": {
                "sheets": []
            }
        }
        
        # Convert each sheet to table format
        for sheet in workbook_data.get("sheets", []):
            table_sheet = {
                "name": sheet.get("name", "Unknown"),
                "tables": []
            }
            
            # Create a single table from the sheet data
            if "rows" in sheet and sheet["rows"]:
                # Expand RLE cells for table display
                expanded_rows = []
                for row in sheet["rows"]:
                    expanded_row = self.expand_rle_cells(row)
                    expanded_rows.append(expanded_row)
                
                table = {
                    "table_id": 1,
                    "name": f"{sheet['name']} - Table",
                    "range": f"A1:{self._get_last_cell_ref_from_expanded_rows(expanded_rows)}",
                    "headers": self._extract_headers_from_expanded_rows(expanded_rows),
                    "data": self._convert_expanded_rows_to_table_data(expanded_rows),
                    "columns": self._create_columns_for_expanded_rows(expanded_rows)
                }
                table_sheet["tables"].append(table)
            
            table_data["workbook"]["sheets"].append(table_sheet)
        
        return table_data
    
    def _get_last_cell_ref_from_expanded_rows(self, expanded_rows: List[Dict[str, Any]]) -> str:
        """Get the last cell reference from expanded rows"""
        max_row = 0
        max_col = 0
        
        for row in expanded_rows:
            max_row = max(max_row, row.get("r", 0))
            for cell in row.get("cells", []):
                max_col = max(max_col, cell[0])
        
        # Convert to Excel reference
        col_str = ""
        temp_col = max_col
        while temp_col > 0:
            temp_col -= 1
            col_str = chr(65 + (temp_col % 26)) + col_str
            temp_col //= 26
        
        return f"{col_str}{max_row}"
    
    def _extract_headers_from_expanded_rows(self, expanded_rows: List[Dict[str, Any]]) -> List[str]:
        """Extract headers from the first expanded row"""
        if not expanded_rows:
            return []
        
        first_row = expanded_rows[0]
        headers = []
        
        for cell in first_row.get("cells", []):
            headers.append(str(cell[1]) if cell[1] is not None else "")
        
        return headers
    
    def _convert_expanded_rows_to_table_data(self, expanded_rows: List[Dict[str, Any]]) -> List[List[Any]]:
        """Convert expanded rows format to table data format"""
        table_data = []
        
        for row in expanded_rows[1:]:  # Skip header row
            row_data = []
            cells_dict = {cell[0]: cell[1] for cell in row.get("cells", [])}
            
            # Find max column to ensure consistent row length
            max_col = max(cells_dict.keys()) if cells_dict else 0
            
            for col in range(1, max_col + 1):
                row_data.append(cells_dict.get(col, ""))
            
            table_data.append(row_data)
        
        return table_data
    
    def _create_columns_for_expanded_rows(self, expanded_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Create columns structure for expanded rows"""
        if not expanded_rows:
            return []
        
        # Get headers from first row
        headers = []
        first_row = expanded_rows[0]
        for cell in first_row.get("cells", []):
            headers.append({
                "col": cell[0],
                "value": str(cell[1]) if cell[1] is not None else ""
            })
        
        columns = []
        
        # Create a column for each header
        for header in headers:
            column = {
                "column_id": header["col"],
                "header": header["value"],
                "cells": {}
            }
            
            # Add cells from all rows for this column
            for row_idx, row in enumerate(expanded_rows[1:], start=2):  # Skip header row
                cells_dict = {cell[0]: cell for cell in row.get("cells", [])}
                
                if header["col"] in cells_dict:
                    cell_data = cells_dict[header["col"]]
                    cell_ref = f"{chr(64 + header['col'])}{row_idx}"  # Convert to A1, B2, etc.
                    
                    column["cells"][cell_ref] = {
                        "coordinate": cell_ref,
                        "value": cell_data[1],
                        "headers": {
                            "column_headers": [{"value": header["value"]}],
                            "row_headers": [{"value": str(cells_dict.get(1, [None, ""])[1])}] if 1 in cells_dict else []
                        }
                    }
            
            columns.append(column)
        
        return columns
    
    def _filter_empty_trailing_areas(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Filter out empty trailing rows and columns from each sheet.
        Only removes rows/columns where there's no data and no data after them.
        """
        if 'workbook' not in result or 'sheets' not in result['workbook']:
            return result
        
        filtered_result = result.copy()
        filtered_sheets = []
        
        for sheet in result['workbook']['sheets']:
            filtered_sheet = self._filter_sheet_empty_trailing_areas(sheet)
            filtered_sheets.append(filtered_sheet)
        
        filtered_result['workbook']['sheets'] = filtered_sheets
        return filtered_result
    
    def _filter_sheet_empty_trailing_areas(self, sheet: Dict[str, Any]) -> Dict[str, Any]:
        """Filter empty trailing areas from a single sheet"""
        if 'rows' not in sheet:
            return sheet
        
        rows = sheet['rows']
        if not rows:
            return sheet
        
        # Find the actual data extents (considering RLE cells)
        max_data_row = 0
        max_data_col = 0
        
        for row in rows:
            row_num = row.get('r', 0)
            cells = row.get('cells', [])
            
            for cell in cells:
                if len(cell) >= 2 and self._has_meaningful_data(cell[1]):
                    if self._is_rle_cell(cell):
                        # RLE cell: calculate end column
                        start_col = cell[0]
                        run_length = cell[-1]
                        end_col = start_col + run_length - 1
                        col_num = end_col
                    else:
                        # Normal cell
                        col_num = cell[0]
                    
                    if row_num > max_data_row:
                        max_data_row = row_num
                    if col_num > max_data_col:
                        max_data_col = col_num
        
        # Filter rows - keep only rows up to max_data_row
        filtered_rows = []
        for row in rows:
            row_num = row.get('r', 0)
            if row_num <= max_data_row:
                # Filter columns within this row
                filtered_cells = []
                for cell in row.get('cells', []):
                    if self._is_rle_cell(cell):
                        # RLE cell: check if any part is within range
                        start_col = cell[0]
                        run_length = cell[-1]
                        end_col = start_col + run_length - 1
                        if start_col <= max_data_col:
                            # Trim RLE cell if it extends beyond max_data_col
                            if end_col > max_data_col:
                                trimmed_length = max_data_col - start_col + 1
                                trimmed_cell = cell[:-1] + [trimmed_length]  # Update run length
                                filtered_cells.append(trimmed_cell)
                            else:
                                filtered_cells.append(cell)
                    else:
                        # Normal cell
                        if len(cell) >= 1 and cell[0] <= max_data_col:
                            filtered_cells.append(cell)
                
                # Only include row if it has cells after filtering
                if filtered_cells:
                    filtered_row = row.copy()
                    filtered_row['cells'] = filtered_cells
                    filtered_rows.append(filtered_row)
        
        # Update dimensions to reflect actual data extent
        filtered_sheet = sheet.copy()
        filtered_sheet['rows'] = filtered_rows
        
        if max_data_row > 0 and max_data_col > 0:
            original_dimensions = sheet.get('dimensions', [1, 1, 1, 1])
            filtered_sheet['dimensions'] = [
                original_dimensions[0],  # min_row
                original_dimensions[1],  # min_col
                max_data_row,           # max_row (filtered)
                max_data_col            # max_col (filtered)
            ]
        
        return filtered_sheet
    
    def _has_meaningful_data(self, value) -> bool:
        """Check if a cell value represents meaningful data"""
        if value is None:
            return False
        if isinstance(value, str) and value.strip() == '':
            return False
        return True
