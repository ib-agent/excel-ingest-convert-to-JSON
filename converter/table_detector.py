"""
Simplified Table Detection Module

This module provides a unified, simplified approach to detecting tables in Excel data
that eliminates the complexity and inconsistencies in the current implementation.
"""

from typing import Dict, Any, List, Optional, Tuple
from openpyxl.utils import get_column_letter
import re


class TableDetector:
    """
    Unified table detector that provides consistent, accurate detection
    across both regular and compact Excel formats.
    """
    
    def __init__(self):
        self.detection_methods = [
            self._detect_by_frozen_panes,
            self._detect_financial_statement_layout,
            self._detect_by_blank_row_separation,
            self._detect_by_temporal_headers,
            self._detect_by_column_continuity,
            self._detect_by_multirow_headers,
            self._detect_by_gaps,
            self._detect_by_formatting,
            self._detect_by_content_structure
        ]
    
    def detect_tables(self, cell_data: Dict[str, Any], dimensions: Dict[str, int], 
                     options: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        Main entry point for table detection.
        
        Args:
            cell_data: Dictionary of cell data (coordinate -> cell_info)
            dimensions: Sheet dimensions {min_row, max_row, min_col, max_col}
            options: Detection options
            
        Returns:
            List of detected table regions
        """
        options = options or {}
        
        # Normalize input data
        normalized_cells = self._normalize_cell_data(cell_data)
        bounds = self._extract_bounds(dimensions)
        
        # Try detection methods in priority order
        for method in self.detection_methods:
            regions = method(normalized_cells, bounds, options)
            if regions:
                return self._validate_and_clean_regions(regions, bounds)
        
        # Fallback: create single table for entire data area
        return self._create_default_table(normalized_cells, bounds)
    
    def _normalize_cell_data(self, cell_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Normalize cell data to a consistent format for detection.
        
        Args:
            cell_data: Raw cell data (can be from regular or compact format)
            
        Returns:
            Normalized cell data: {coordinate: {'value': value, 'row': int, 'col': int}}
        """
        normalized = {}
        
        if isinstance(cell_data, dict):
            # Handle regular format (coordinate -> cell_data)
            for coord, cell_info in cell_data.items():
                if isinstance(cell_info, dict) and 'value' in cell_info:
                    value = cell_info['value']
                    if value is not None and str(value).strip():
                        normalized[coord] = {
                            'value': value,
                            'row': cell_info.get('row', self._coord_to_row(coord)),
                            'col': cell_info.get('column', self._coord_to_col(coord))
                        }
        elif isinstance(cell_data, list):
            # Handle compact format (list of row objects)
            for row_data in cell_data:
                row_num = row_data.get('r', 1)
                for cell_array in row_data.get('cells', []):
                    if len(cell_array) >= 2:
                        col_num = cell_array[0]
                        value = cell_array[1]
                        if value is not None and str(value).strip():
                            coord = f"{get_column_letter(col_num)}{row_num}"
                            normalized[coord] = {
                                'value': value,
                                'row': row_num,
                                'col': col_num
                            }
        
        return normalized
    
    def _extract_bounds(self, dimensions: Dict[str, int]) -> Tuple[int, int, int, int]:
        """Extract (min_row, min_col, max_row, max_col) from dimensions."""
        if isinstance(dimensions, dict):
            return (
                dimensions.get('min_row', 1),
                dimensions.get('min_col', 1), 
                dimensions.get('max_row', 1),
                dimensions.get('max_col', 1)
            )
        elif isinstance(dimensions, list) and len(dimensions) == 4:
            return tuple(dimensions)
        else:
            return (1, 1, 1, 1)
    
    def _detect_by_frozen_panes(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                               options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables based on frozen panes - highest priority.
        If any panes are frozen, treat as single structured table.
        """
        frozen_info = self._get_frozen_panes_info(options)
        if frozen_info['rows'] > 0 or frozen_info['cols'] > 0:
            min_row, min_col, max_row, max_col = bounds
            return [{
                'start_row': min_row,
                'end_row': max_row,
                'start_col': min_col,
                'end_col': max_col,
                'detection_method': 'frozen_panes',
                'frozen_rows': frozen_info['rows'],
                'frozen_cols': frozen_info['cols']
            }]
        return []
    
    def _detect_financial_statement_layout(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                          options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect financial statement layouts where section headers (with labels only) 
        are part of the same table, not separate tables.
        
        This method specifically handles layouts like balance sheets where you have:
        - Row headers with dates/periods across columns
        - Section headers like "Assets", "Liabilities", "Equity" with only text in first column
        - Data rows with values across columns
        """
        min_row, min_col, max_row, max_col = bounds
        regions = []
        
        # Check if this looks like a financial statement layout
        if not self._looks_like_financial_statement(cells, bounds):
            return []
        
        # For financial statements, treat the entire data area as one table
        # but ensure we have meaningful data
        data_rows = []
        for row in range(min_row, max_row + 1):
            if self._row_has_meaningful_data(cells, row, min_col, max_col):
                data_rows.append(row)
        
        if len(data_rows) >= 3:  # Need at least 3 rows for a table
            # Find actual column bounds (exclude empty trailing columns)
            actual_bounds = self._determine_table_column_bounds(cells, min(data_rows), max(data_rows), min_col, max_col)
            
            regions.append({
                'start_row': min(data_rows),
                'end_row': max(data_rows),
                'start_col': actual_bounds['min_col'],
                'end_col': actual_bounds['max_col'],
                'detection_method': 'financial_statement_layout'
            })
        
        return regions
    
    def _looks_like_financial_statement(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int]) -> bool:
        """
        Check if the data layout looks like a financial statement.
        
        Characteristics:
        - Has section headers (rows with only text in first column)
        - Has consistent column structure across most rows
        - Section headers are common financial terms
        """
        min_row, min_col, max_row, max_col = bounds
        
        section_headers = []
        data_rows = []
        
        # Analyze each row
        for row in range(min_row, max_row + 1):
            first_col_coord = f"{get_column_letter(min_col)}{row}"
            if first_col_coord in cells:
                first_col_value = cells[first_col_coord]['value']
                
                # Count data in other columns
                other_data_count = 0
                for col in range(min_col + 1, max_col + 1):
                    coord = f"{get_column_letter(col)}{row}"
                    if coord in cells and cells[coord]['value'] is not None:
                        value = str(cells[coord]['value']).strip()
                        if value:
                            other_data_count += 1
                
                # Classify the row
                if isinstance(first_col_value, str) and first_col_value.strip():
                    if other_data_count == 0:
                        # Row with only first column - potential section header
                        section_headers.append((row, first_col_value.strip()))
                    elif other_data_count > 2:
                        # Row with data across multiple columns
                        data_rows.append(row)
        
        # Check if we have financial statement characteristics
        if len(section_headers) >= 2 and len(data_rows) >= 3:
            # Check if section headers contain financial terms
            financial_terms = [
                'assets', 'liabilities', 'equity', 'revenue', 'expenses', 'income',
                'cash', 'receivable', 'payable', 'inventory', 'property', 'debt',
                'retained', 'earnings', 'capital', 'current', 'non-current', 'total'
            ]
            
            financial_header_count = 0
            for _, header_text in section_headers:
                header_lower = header_text.lower()
                if any(term in header_lower for term in financial_terms):
                    financial_header_count += 1
            
            # If most section headers are financial terms, this is likely a financial statement
            return financial_header_count >= len(section_headers) * 0.6
        
        return False
    
    def _row_has_meaningful_data(self, cells: Dict[str, Any], row: int, min_col: int, max_col: int) -> bool:
        """Check if a row has meaningful data (not just empty cells)."""
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells and cells[coord]['value'] is not None:
                value = str(cells[coord]['value']).strip()
                if value:  # Non-empty value
                    return True
        return False
    
    def _detect_by_blank_row_separation(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                       options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables separated by blank rows. This method looks for 1-3 row gaps 
        as natural delimiters between blocks of data.
        """
        min_row, min_col, max_row, max_col = bounds
        regions = []
        
        # Find all rows with data
        data_rows = []
        for row in range(min_row, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                data_rows.append(row)
        
        if len(data_rows) < 4:  # Need at least 4 rows for multiple tables
            return []
        
        # Look for gaps that indicate table separations
        # Small gaps (1-3 rows) or larger gaps (4+ rows) can both indicate separate tables
        table_boundaries = [data_rows[0]]  # Start with first data row
        
        for i in range(1, len(data_rows)):
            gap_size = data_rows[i] - data_rows[i-1] - 1
            
            # Any gap of 1 or more rows could indicate table boundaries
            if gap_size >= 1:
                # For small gaps (1-3 rows), check if it's a genuine boundary
                # For larger gaps (4+ rows), assume it's definitely a boundary
                if gap_size >= 4 or self._is_table_boundary(cells, data_rows[i-1], data_rows[i], min_col, max_col):
                    table_boundaries.append(data_rows[i])
        
        # Create table regions from boundaries
        if len(table_boundaries) >= 2:  # At least 2 tables found
            for i in range(len(table_boundaries)):
                start_row = table_boundaries[i]
                
                # Find end row (either next boundary or last data row)
                if i < len(table_boundaries) - 1:
                    # Find last data row before next boundary
                    next_boundary = table_boundaries[i + 1]
                    end_row = start_row
                    for row in data_rows:
                        if start_row <= row < next_boundary:
                            end_row = row
                else:
                    # Last table - goes to end
                    end_row = data_rows[-1]
                
                if end_row >= start_row:  # Valid table
                    # Determine actual column boundaries for this table region
                    actual_bounds = self._determine_table_column_bounds(cells, start_row, end_row, min_col, max_col)
                    
                    regions.append({
                        'start_row': start_row,
                        'end_row': end_row,
                        'start_col': actual_bounds['min_col'],
                        'end_col': actual_bounds['max_col'],
                        'detection_method': 'blank_row_separation'
                    })
        
        return regions
    
    def _determine_table_column_bounds(self, cells: Dict[str, Any], start_row: int, end_row: int, 
                                     min_col: int, max_col: int) -> Dict[str, int]:
        """
        Determine the actual column boundaries for a table region by analyzing data distribution.
        
        Args:
            cells: Normalized cell data
            start_row: Table start row
            end_row: Table end row  
            min_col: Overall minimum column
            max_col: Overall maximum column
            
        Returns:
            Dictionary with actual min_col and max_col for the table
        """
        # Find columns that have data in this row range
        cols_with_data = set()
        
        for row in range(start_row, end_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                if coord in cells and cells[coord]['value'] is not None:
                    cols_with_data.add(col)
        
        if not cols_with_data:
            return {'min_col': min_col, 'max_col': max_col}
        
        # Return the actual range of columns with data
        return {
            'min_col': min(cols_with_data),
            'max_col': max(cols_with_data)
        }
    
    def _is_table_boundary(self, cells: Dict[str, Any], prev_row: int, next_row: int, 
                          min_col: int, max_col: int) -> bool:
        """
        Check if the gap between prev_row and next_row represents a genuine table boundary.
        """
        # Check if there's a potential section header row between prev_row and next_row
        for gap_row in range(prev_row + 1, next_row):
            if self._is_section_header_row(cells, gap_row, min_col, max_col, next_row, max_col):
                # This is likely a section header within the same table, not a boundary
                return False
        
        # Analyze content patterns around the boundary
        prev_content = self._get_row_content_pattern(cells, prev_row, min_col, max_col)
        next_content = self._get_row_content_pattern(cells, next_row, min_col, max_col)
        
        # Different column counts suggest different tables
        if abs(prev_content['col_count'] - next_content['col_count']) > 2:
            return True
        
        # Different content types suggest different tables
        if prev_content['mostly_numeric'] != next_content['mostly_numeric']:
            return True
        
        # Check for header-like content in next row
        if next_content['has_text_labels'] and not prev_content['has_text_labels']:
            return True
        
        return False
    
    def _get_row_content_pattern(self, cells: Dict[str, Any], row: int, 
                                min_col: int, max_col: int) -> Dict[str, Any]:
        """
        Analyze the content pattern of a row.
        """
        values = []
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                values.append(cells[coord]['value'])
        
        if not values:
            return {'col_count': 0, 'mostly_numeric': False, 'has_text_labels': False}
        
        numeric_count = sum(1 for val in values 
                           if isinstance(val, (int, float)) or 
                           (isinstance(val, str) and val.replace('.', '').replace('-', '').isdigit()))
        
        text_labels = [val for val in values 
                      if isinstance(val, str) and len(str(val)) > 3 and 
                      not str(val).replace('.', '').replace('-', '').isdigit()]
        
        return {
            'col_count': len(values),
            'mostly_numeric': numeric_count > len(values) / 2,
            'has_text_labels': len(text_labels) > 2
        }
    
    def _is_section_header_row(self, cells: Dict[str, Any], row: int, min_col: int, max_col: int, next_data_row: int, max_row: int = None) -> bool:
        """
        Check if a row is a section header (label in first column only, no other data).
        
        A row is considered a section header if:
        1. It has text in the first column (typically a label)
        2. It has no or minimal data in other columns  
        3. The row following it is NOT a row of dates/headers (indicating it's part of the same table)
        """
        # Check if this row has data in the first column
        first_col_coord = f"{get_column_letter(min_col)}{row}"
        if first_col_coord not in cells:
            return False
        
        first_col_value = cells[first_col_coord]['value']
        if not isinstance(first_col_value, str) or len(str(first_col_value).strip()) == 0:
            return False
        
        # Count data cells in other columns (excluding first column)
        other_data_count = 0
        for col in range(min_col + 1, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells and cells[coord]['value'] is not None:
                value = str(cells[coord]['value']).strip()
                if value:  # Non-empty value
                    other_data_count += 1
        
        # If this row has significant data in other columns, it's not a section header
        if other_data_count > 2:
            return False
        
        # Check if the next data row contains dates or header-like content
        # If it does, then this might actually be a new table
        if max_row is None or next_data_row <= max_row:
            next_row_pattern = self._get_row_content_pattern(cells, next_data_row, min_col, max_col)
            
            # Check for date patterns in the next row (indicating new table with date headers)
            has_date_headers = self._row_has_date_pattern(cells, next_data_row, min_col, max_col)
            
            # If next row has date headers or significantly different structure, this could be a new table
            if has_date_headers:
                return False  # This is likely a real table boundary
        
        # This appears to be a section header within the same table
        return True
    
    def _row_has_date_pattern(self, cells: Dict[str, Any], row: int, min_col: int, max_col: int) -> bool:
        """
        Check if a row contains date-like patterns that would indicate column headers.
        """
        date_like_count = 0
        total_values = 0
        
        for col in range(min_col + 1, max_col + 1):  # Skip first column (labels)
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells and cells[coord]['value'] is not None:
                value = str(cells[coord]['value']).strip()
                if value:
                    total_values += 1
                    # Check for date-like patterns
                    if self._looks_like_date(value):
                        date_like_count += 1
        
        # If more than 30% of values look like dates, consider it a date header row
        return total_values > 0 and (date_like_count / total_values) > 0.3
    
    def _looks_like_date(self, value: str) -> bool:
        """
        Check if a value looks like a date.
        """
        import re
        
        # Common date patterns
        date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # MM/DD/YYYY, MM-DD-YYYY
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',    # YYYY/MM/DD, YYYY-MM-DD  
            r'\w{3}\s+\d{1,2}[,]?\s+\d{4}',    # Jan 1, 2024
            r'\d{1,2}\s+\w{3}\s+\d{4}',        # 1 Jan 2024
            r'Q[1-4]\s+\d{4}',                 # Q1 2024
            r'\d{4}Q[1-4]',                    # 2024Q1
        ]
        
        for pattern in date_patterns:
            if re.search(pattern, value, re.IGNORECASE):
                return True
        
        return False
    
    def _detect_by_temporal_headers(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                   options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables with temporal column labels (sequential dates).
        If columns contain sequential dates, treat as a time-series table.
        """
        min_row, min_col, max_row, max_col = bounds
        
        # Look for rows with date patterns
        temporal_header_rows = []
        for row in range(min_row, min(min_row + 5, max_row + 1)):  # Check first 5 rows
            if self._has_temporal_headers(cells, row, min_col, max_col):
                temporal_header_rows.append(row)
        
        if not temporal_header_rows:
            return []
        
        # For each temporal header row, create a table region
        regions = []
        for header_row in temporal_header_rows:
            # Find the extent of the temporal table
            table_start = header_row
            table_end = self._find_temporal_table_end(cells, header_row, min_row, max_row, min_col, max_col)
            
            if table_end > table_start:
                regions.append({
                    'start_row': table_start,
                    'end_row': table_end,
                    'start_col': min_col,
                    'end_col': max_col,
                    'detection_method': 'temporal_headers',
                    'temporal_header_row': header_row
                })
        
        return regions
    
    def _has_temporal_headers(self, cells: Dict[str, Any], row: int, 
                             min_col: int, max_col: int) -> bool:
        """
        Check if a row contains temporal headers (dates like 2025-05-31 or Month 1, Month 2).
        """
        date_count = 0
        month_count = 0
        total_cells = 0
        
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                value = str(cells[coord]['value'])
                total_cells += 1
                
                # Check for ISO date patterns (2025-05-31)
                if re.search(r'202[0-9]-\d{2}-\d{2}', value):
                    date_count += 1
                
                # Check for month patterns (Month 1, Month 2, etc.)
                if re.search(r'Month \d+', value, re.IGNORECASE):
                    month_count += 1
        
        # If more than 25% of cells contain dates or month patterns, it's temporal
        if total_cells > 0:
            temporal_ratio = (date_count + month_count) / total_cells
            return temporal_ratio > 0.25
        
        return False
    
    def _find_temporal_table_end(self, cells: Dict[str, Any], header_row: int,
                                min_row: int, max_row: int, min_col: int, max_col: int) -> int:
        """
        Find where a temporal table ends by looking for gaps or next temporal headers.
        """
        last_data_row = header_row
        
        # Look for the next temporal header or significant gap
        for row in range(header_row + 1, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                # Check if this is another temporal header (new table starts)
                if self._has_temporal_headers(cells, row, min_col, max_col):
                    # New temporal table found - stop here
                    break
                else:
                    # Continue with current table
                    last_data_row = row
            else:
                # Empty row - check for gaps
                gap_start = row
                next_data_row = self._find_next_data_row(cells, row + 1, max_row, min_col, max_col)
                if next_data_row:
                    gap_size = next_data_row - gap_start
                    if gap_size > 1:  # Gap of 2+ rows indicates table boundary
                        break
                else:
                    # No more data
                    break
        
        return last_data_row
    
    def _is_part_of_temporal_table(self, cells: Dict[str, Any], row: int, header_row: int,
                                  min_col: int, max_col: int) -> bool:
        """
        Check if a row is part of the same temporal table as the header row.
        """
        # Get column structure of header row
        header_cols = []
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{header_row}"
            if coord in cells:
                header_cols.append(col)
        
        # Get column structure of current row
        row_cols = []
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                row_cols.append(col)
        
        # Similar column structure suggests same table
        if len(header_cols) > 0 and len(row_cols) > 0:
            overlap = len(set(header_cols) & set(row_cols))
            return overlap / max(len(header_cols), len(row_cols)) > 0.5
        
        return False
    
    def _find_next_data_row(self, cells: Dict[str, Any], start_row: int, max_row: int,
                           min_col: int, max_col: int) -> Optional[int]:
        """
        Find the next row with data starting from start_row.
        """
        for row in range(start_row, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                return row
        return None
    
    def _detect_by_column_continuity(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                   options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables based on column continuity and data type patterns.
        If column data types change significantly, it may indicate a new table.
        """
        min_row, min_col, max_row, max_col = bounds
        
        # Find all rows with data
        data_rows = []
        for row in range(min_row, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                data_rows.append(row)
        
        if len(data_rows) < 3:
            return []
        
        # Analyze column patterns for each row
        row_patterns = {}
        for row in data_rows:
            row_patterns[row] = self._analyze_row_column_pattern(cells, row, min_col, max_col)
        
        # Find boundaries where column patterns change significantly
        table_boundaries = [data_rows[0]]  # Start with first row
        
        for i in range(1, len(data_rows)):
            current_row = data_rows[i]
            prev_row = data_rows[i-1]
            
            current_pattern = row_patterns[current_row]
            prev_pattern = row_patterns[prev_row]
            
            # Check if this represents a significant pattern change
            if self._is_significant_column_pattern_change(prev_pattern, current_pattern):
                table_boundaries.append(current_row)
        
        # Create table regions from boundaries
        regions = []
        if len(table_boundaries) >= 2:  # At least 2 tables found
            for i in range(len(table_boundaries)):
                start_row = table_boundaries[i]
                
                # Find end row
                if i < len(table_boundaries) - 1:
                    # Find last data row before next boundary
                    next_boundary = table_boundaries[i + 1]
                    end_row = start_row
                    for row in data_rows:
                        if start_row <= row < next_boundary:
                            end_row = row
                else:
                    # Last table
                    end_row = data_rows[-1]
                
                if end_row >= start_row:
                    regions.append({
                        'start_row': start_row,
                        'end_row': end_row,
                        'start_col': min_col,
                        'end_col': max_col,
                        'detection_method': 'column_continuity'
                    })
        
        return regions
    
    def _analyze_row_column_pattern(self, cells: Dict[str, Any], row: int, 
                                   min_col: int, max_col: int) -> Dict[str, Any]:
        """
        Analyze the column pattern of a row (data types, column coverage, etc.).
        """
        values = []
        occupied_cols = []
        
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                value = cells[coord]['value']
                values.append(value)
                occupied_cols.append(col)
        
        if not values:
            return {
                'col_count': 0,
                'numeric_ratio': 0,
                'text_ratio': 0,
                'column_span': 0,
                'data_density': 0
            }
        
        # Analyze data types
        numeric_count = sum(1 for val in values 
                           if isinstance(val, (int, float)) or 
                           (isinstance(val, str) and self._is_numeric_string(val)))
        
        text_count = sum(1 for val in values 
                        if isinstance(val, str) and not self._is_numeric_string(val))
        
        # Calculate metrics
        total_values = len(values)
        column_span = max(occupied_cols) - min(occupied_cols) + 1 if occupied_cols else 0
        total_possible_cols = max_col - min_col + 1
        
        return {
            'col_count': len(occupied_cols),
            'numeric_ratio': numeric_count / total_values if total_values > 0 else 0,
            'text_ratio': text_count / total_values if total_values > 0 else 0,
            'column_span': column_span,
            'data_density': len(occupied_cols) / total_possible_cols if total_possible_cols > 0 else 0
        }
    
    def _is_numeric_string(self, value: str) -> bool:
        """Check if a string represents a numeric value."""
        if not isinstance(value, str):
            return False
        
        # Remove common formatting characters
        cleaned = value.replace(',', '').replace('$', '').replace('%', '').replace(' ', '')
        
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    
    def _is_significant_column_pattern_change(self, prev_pattern: Dict[str, Any], 
                                            current_pattern: Dict[str, Any]) -> bool:
        """
        Determine if there's a significant change in column patterns between rows.
        """
        # Column count change threshold
        col_count_change = abs(prev_pattern['col_count'] - current_pattern['col_count'])
        if col_count_change > 5:  # More than 5 columns difference
            return True
        
        # Data type ratio changes
        numeric_change = abs(prev_pattern['numeric_ratio'] - current_pattern['numeric_ratio'])
        if numeric_change > 0.4:  # 40% change in numeric ratio
            return True
        
        # Column span changes
        span_change = abs(prev_pattern['column_span'] - current_pattern['column_span'])
        if span_change > 10:  # More than 10 columns span difference
            return True
        
        # Data density changes
        density_change = abs(prev_pattern['data_density'] - current_pattern['data_density'])
        if density_change > 0.3:  # 30% change in data density
            return True
        
        return False
    
    def _detect_by_multirow_headers(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                   options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect multi-level headers stacked in the first 2-3 rows.
        If consecutive rows contain mostly string data and differ from structure below,
        they likely form a header block.
        """
        min_row, min_col, max_row, max_col = bounds
        
        # Look for multi-row header blocks in the first few rows
        header_blocks = self._find_multirow_header_blocks(cells, min_row, min(min_row + 10, max_row), min_col, max_col)
        
        if not header_blocks:
            return []
        
        regions = []
        
        for header_block in header_blocks:
            header_end = header_block['end_row']
            
            # Find the data section after this header block
            data_start = self._find_next_data_row(cells, header_end + 1, max_row, min_col, max_col)
            if data_start:
                # Find where this table ends
                table_end = self._find_multirow_table_end(cells, data_start, max_row, min_col, max_col)
                
                if table_end and table_end >= data_start:
                    regions.append({
                        'start_row': header_block['start_row'],
                        'end_row': table_end,
                        'start_col': min_col,
                        'end_col': max_col,
                        'detection_method': 'multirow_headers',
                        'header_rows': header_block['end_row'] - header_block['start_row'] + 1,
                        'data_start_row': data_start
                    })
        
        return regions
    
    def _find_multirow_header_blocks(self, cells: Dict[str, Any], min_row: int, max_row: int,
                                   min_col: int, max_col: int) -> List[Dict[str, int]]:
        """
        Find blocks of consecutive rows that look like multi-row headers.
        """
        header_blocks = []
        current_block_start = None
        
        for row in range(min_row, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                row_pattern = self._get_row_content_pattern(cells, row, min_col, max_col)
                
                # Check if this row looks like a header row
                is_header_like = (
                    row_pattern['has_text_labels'] and 
                    not row_pattern['mostly_numeric'] and
                    row_pattern['col_count'] >= 3
                )
                
                if is_header_like:
                    if current_block_start is None:
                        current_block_start = row
                else:
                    # End of header block
                    if current_block_start is not None:
                        block_size = row - current_block_start
                        if block_size >= 2:  # At least 2 rows for multi-row header
                            header_blocks.append({
                                'start_row': current_block_start,
                                'end_row': row - 1
                            })
                        current_block_start = None
            else:
                # Empty row - end any current block
                if current_block_start is not None:
                    block_size = row - current_block_start
                    if block_size >= 2:
                        header_blocks.append({
                            'start_row': current_block_start,
                            'end_row': row - 1
                        })
                    current_block_start = None
        
        # Handle block that goes to the end
        if current_block_start is not None:
            block_size = max_row - current_block_start + 1
            if block_size >= 2:
                header_blocks.append({
                    'start_row': current_block_start,
                    'end_row': max_row
                })
        
        return header_blocks
    
    def _find_multirow_table_end(self, cells: Dict[str, Any], data_start: int, max_row: int,
                                min_col: int, max_col: int) -> Optional[int]:
        """
        Find where a table with multirow headers ends.
        """
        last_data_row = data_start
        
        for row in range(data_start + 1, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                # Check if this might be the start of a new header block
                row_pattern = self._get_row_content_pattern(cells, row, min_col, max_col)
                if (row_pattern['has_text_labels'] and 
                    not row_pattern['mostly_numeric'] and
                    row_pattern['col_count'] >= 3):
                    # Might be new header - check next few rows
                    if self._looks_like_new_header_block(cells, row, max_row, min_col, max_col):
                        break
                
                last_data_row = row
            else:
                # Gap found - check if it's significant
                next_data_row = self._find_next_data_row(cells, row + 1, max_row, min_col, max_col)
                if next_data_row and next_data_row - row > 2:
                    break
        
        return last_data_row
    
    def _looks_like_new_header_block(self, cells: Dict[str, Any], start_row: int, max_row: int,
                                    min_col: int, max_col: int) -> bool:
        """
        Check if starting from start_row, it looks like a new header block.
        """
        header_like_rows = 0
        
        for row in range(start_row, min(start_row + 3, max_row + 1)):
            if self._row_has_data(cells, row, min_col, max_col):
                row_pattern = self._get_row_content_pattern(cells, row, min_col, max_col)
                if (row_pattern['has_text_labels'] and 
                    not row_pattern['mostly_numeric']):
                    header_like_rows += 1
        
        return header_like_rows >= 2
    
    def _detect_by_gaps(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                       options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Simplified gap-based detection. Only runs if explicitly requested.
        """
        if not options.get('table_detection', {}).get('use_gaps', False):
            return []
        
        min_row, min_col, max_row, max_col = bounds
        regions = []
        
        # Find rows with data
        data_rows = []
        for row in range(min_row, max_row + 1):
            if self._row_has_data(cells, row, min_col, max_col):
                data_rows.append(row)
        
        if len(data_rows) < 2:
            return []
        
        # Find gaps between data rows
        gap_threshold = options.get('table_detection', {}).get('gap_threshold', 3)
        current_start = data_rows[0]
        
        for i in range(1, len(data_rows)):
            gap_size = data_rows[i] - data_rows[i-1] - 1
            
            if gap_size >= gap_threshold:
                # End current table
                regions.append({
                    'start_row': current_start,
                    'end_row': data_rows[i-1],
                    'start_col': min_col,
                    'end_col': max_col,
                    'detection_method': 'gaps'
                })
                current_start = data_rows[i]
        
        # Add final table
        regions.append({
            'start_row': current_start,
            'end_row': data_rows[-1],
            'start_col': min_col,
            'end_col': max_col,
            'detection_method': 'gaps'
        })
        
        return regions
    
    def _detect_by_formatting(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                             options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables based on formatting patterns (headers, borders, etc.).
        """
        # This would analyze cell styling to identify table boundaries
        # For now, return empty - can be enhanced later
        return []
    
    def _detect_by_content_structure(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int],
                                   options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect tables based on content structure (headers, data patterns).
        """
        min_row, min_col, max_row, max_col = bounds
        
        # Check if this looks like a structured table
        if self._has_structured_layout(cells, bounds):
            return [{
                'start_row': min_row,
                'end_row': max_row,
                'start_col': min_col,
                'end_col': max_col,
                'detection_method': 'content_structure'
            }]
        
        return []
    
    def _has_structured_layout(self, cells: Dict[str, Any], bounds: Tuple[int, int, int, int]) -> bool:
        """
        Check if the data has a structured table layout.
        """
        min_row, min_col, max_row, max_col = bounds
        
        # Count rows and columns with data
        data_rows = sum(1 for row in range(min_row, max_row + 1) 
                       if self._row_has_data(cells, row, min_col, max_col))
        data_cols = sum(1 for col in range(min_col, max_col + 1)
                       if self._col_has_data(cells, col, min_row, max_row))
        
        # Must have reasonable table dimensions
        return data_rows >= 3 and data_cols >= 2
    
    def _row_has_data(self, cells: Dict[str, Any], row: int, min_col: int, max_col: int) -> bool:
        """Check if a row has any data."""
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                return True
        return False
    
    def _col_has_data(self, cells: Dict[str, Any], col: int, min_row: int, max_row: int) -> bool:
        """Check if a column has any data."""
        for row in range(min_row, max_row + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in cells:
                return True
        return False
    
    def _get_frozen_panes_info(self, options: Dict[str, Any]) -> Dict[str, int]:
        """Extract frozen panes information from options."""
        sheet_data = options.get('sheet_data', {})
        
        # Handle different formats
        if 'frozen_panes' in sheet_data:
            frozen = sheet_data['frozen_panes']
            if isinstance(frozen, dict):
                return {
                    'rows': frozen.get('frozen_rows', 0),
                    'cols': frozen.get('frozen_cols', 0)
                }
        
        if 'frozen' in sheet_data:
            frozen = sheet_data['frozen']
            if isinstance(frozen, list) and len(frozen) >= 2:
                return {'rows': frozen[0], 'cols': frozen[1]}
        
        return {'rows': 0, 'cols': 0}
    
    def _validate_and_clean_regions(self, regions: List[Dict[str, Any]], 
                                   bounds: Tuple[int, int, int, int]) -> List[Dict[str, Any]]:
        """
        Validate and clean detected regions.
        """
        min_row, min_col, max_row, max_col = bounds
        cleaned = []
        
        for region in regions:
            # Ensure region is within bounds
            start_row = max(region['start_row'], min_row)
            end_row = min(region['end_row'], max_row)
            start_col = max(region['start_col'], min_col)
            end_col = min(region['end_col'], max_col)
            
            # Must have at least 1x1 size
            if start_row <= end_row and start_col <= end_col:
                cleaned.append({
                    'start_row': start_row,
                    'end_row': end_row,
                    'start_col': start_col,
                    'end_col': end_col,
                    'detection_method': region.get('detection_method', 'unknown')
                })
        
        return cleaned
    
    def _create_default_table(self, cells: Dict[str, Any], 
                             bounds: Tuple[int, int, int, int]) -> List[Dict[str, Any]]:
        """
        Create a default table covering the entire data area.
        """
        if not cells:
            return []
        
        min_row, min_col, max_row, max_col = bounds
        return [{
            'start_row': min_row,
            'end_row': max_row,
            'start_col': min_col,
            'end_col': max_col,
            'detection_method': 'default'
        }]
    
    def _coord_to_row(self, coord: str) -> int:
        """Extract row number from coordinate string."""
        return int(re.search(r'\d+', coord).group())
    
    def _coord_to_col(self, coord: str) -> int:
        """Extract column number from coordinate string."""
        col_letters = re.search(r'[A-Z]+', coord).group()
        result = 0
        for char in col_letters:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result


class HeaderResolver:
    """
    Simplified header resolution that works with detected table regions.
    """
    
    def resolve_headers(self, cells: Dict[str, Any], region: Dict[str, Any], 
                       options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Resolve header information for a detected table region.
        
        Args:
            cells: Normalized cell data
            region: Table region from detector
            options: Resolution options
            
        Returns:
            Header information with rows, cols, and data_start
        """
        options = options or {}
        
        # Get frozen panes info if available
        frozen_info = self._get_frozen_info(options)
        
        # Determine header rows and columns
        header_rows = self._find_header_rows(cells, region, frozen_info)
        header_cols = self._find_header_cols(cells, region, frozen_info)
        
        # Calculate data start position
        data_start_row = region['start_row'] + len(header_rows)
        data_start_col = region['start_col'] + len(header_cols)
        
        return {
            'header_rows': header_rows,
            'header_columns': header_cols,
            'data_start_row': data_start_row,
            'data_start_col': data_start_col
        }
    
    def _get_frozen_info(self, options: Dict[str, Any]) -> Dict[str, int]:
        """Get frozen panes information."""
        sheet_data = options.get('sheet_data', {})
        
        if 'frozen_panes' in sheet_data:
            frozen = sheet_data['frozen_panes']
            return {
                'rows': frozen.get('frozen_rows', 0),
                'cols': frozen.get('frozen_cols', 0)
            }
        
        if 'frozen' in sheet_data:
            frozen = sheet_data['frozen']
            if isinstance(frozen, list) and len(frozen) >= 2:
                return {'rows': frozen[0], 'cols': frozen[1]}
        
        return {'rows': 0, 'cols': 0}
    
    def _find_header_rows(self, cells: Dict[str, Any], region: Dict[str, Any], 
                         frozen_info: Dict[str, int]) -> List[int]:
        """Find header rows for the table."""
        header_rows = []
        start_row = region['start_row']
        end_row = region['end_row']
        
        # Use frozen rows if available
        if frozen_info['rows'] > 0:
            for row in range(start_row, start_row + frozen_info['rows']):
                if row <= end_row:
                    header_rows.append(row)
        
        # If no frozen rows, use first row as header
        if not header_rows and start_row <= end_row:
            header_rows.append(start_row)
        
        return header_rows
    
    def _find_header_cols(self, cells: Dict[str, Any], region: Dict[str, Any],
                         frozen_info: Dict[str, int]) -> List[int]:
        """Find header columns for the table."""
        header_cols = []
        start_col = region['start_col']
        end_col = region['end_col']
        
        # Use frozen columns if available
        if frozen_info['cols'] > 0:
            for col in range(start_col, start_col + frozen_info['cols']):
                if col <= end_col:
                    header_cols.append(col)
        
        # If no frozen columns, use first column as header
        if not header_cols and start_col <= end_col:
            header_cols.append(start_col)
        
        return header_cols 