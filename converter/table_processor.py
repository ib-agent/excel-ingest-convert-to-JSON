import json
from typing import Dict, Any, List, Optional, Tuple
from openpyxl.utils import get_column_letter, column_index_from_string
from .table_detector import TableDetector, HeaderResolver


class TableProcessor:
    """
    Transforms Excel JSON schema into table-oriented representation
    while preserving all original information and adding table structure
    """
    
    def __init__(self):
        self.table_detection_rules = []
        # Add new simplified detection system
        self.detector = TableDetector()
        self.header_resolver = HeaderResolver()
    
    def transform_to_table_format(self, excel_json: dict, options: dict = None) -> dict:
        """
        Transform Excel JSON schema to table-oriented format
        
        Args:
            excel_json: The original Excel JSON schema output
            options: Configuration options for table detection and processing
            
        Returns:
            Dictionary with table-oriented structure plus all original data
        """
        if not excel_json or 'workbook' not in excel_json:
            raise ValueError("Invalid Excel JSON format")
        
        # Start with the original structure
        result = excel_json.copy()
        
        # Process each sheet to add table information and clean cell data
        for sheet in result['workbook']['sheets']:
            # Clean the cells in the sheet
            if 'cells' in sheet:
                cleaned_cells = {}
                for cell_key, cell_data in sheet['cells'].items():
                    cleaned_cell = self._clean_cell_data(cell_data)
                    if cleaned_cell:  # Only include non-empty cells
                        cleaned_cells[cell_key] = cleaned_cell
                sheet['cells'] = cleaned_cells
            
            sheet['tables'] = self._detect_and_process_tables(sheet, options or {})
        
        # Clean the result to remove null/empty values and fix RGB properties
        result = self._clean_table_json(result)
        
        return result
    
    def _clean_table_json(self, data: dict) -> dict:
        """
        Clean the table-oriented JSON by removing null/empty values and fixing RGB properties
        
        Args:
            data: The table-oriented JSON data
            
        Returns:
            Cleaned JSON data
        """
        if isinstance(data, dict):
            cleaned = {}
            for key, value in data.items():
                # Skip null values and empty strings
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                
                # Skip False boolean values for cleaner output
                if isinstance(value, bool) and not value:
                    continue
                
                # Skip error messages and invalid data patterns (but be less aggressive)
                if isinstance(value, str) and self._is_error_message(value) and len(value) < 50:
                    continue
                
                # Recursively clean nested structures
                if isinstance(value, (dict, list)):
                    cleaned_value = self._clean_table_json(value)
                    # Only include if the cleaned value is not empty
                    if cleaned_value is not None and cleaned_value != {} and cleaned_value != []:
                        cleaned[key] = cleaned_value
                else:
                    cleaned[key] = value
            
            return cleaned if cleaned else None
        
        elif isinstance(data, list):
            cleaned_list = []
            for item in data:
                # For string items, check if they are error messages
                if isinstance(item, str) and self._is_error_message(item):
                    continue
                
                cleaned_item = self._clean_table_json(item)
                if cleaned_item is not None:
                    cleaned_list.append(cleaned_item)
            return cleaned_list if cleaned_list else None
        
        else:
            return data
    
    def _is_error_message(self, value: str) -> bool:
        """
        Check if a string value contains error messages or invalid data patterns
        
        Args:
            value: String value to check
            
        Returns:
            True if the value appears to be an error message or invalid data
        """
        if not isinstance(value, str):
            return False
        
        # Convert to lowercase for case-insensitive matching
        value_lower = value.lower()
        
        # Common error message patterns
        error_patterns = [
            'values must be of type',
            'invalid',
            'error',
            'exception',
            'failed',
            'not found',
            'undefined',
            'null reference',
            'type error',
            'validation error',
            'parse error',
            'syntax error',
            'runtime error',
            'missing',
            'required',
            'cannot',
            'unable to',
            'failed to',
            'unexpected',
            'unrecognized',
            'unknown',
            'unsupported',
            'deprecated',
            'obsolete',
            'placeholder',
            'todo',
            'fixme',
            'bug',
            'issue',
            'problem'
        ]
        
        # Check if the value contains any error patterns
        for pattern in error_patterns:
            if pattern in value_lower:
                return True
        
        # Check for incomplete or truncated strings that might be error messages
        if value_lower.endswith(' of type') or value_lower.endswith(' must be') or value_lower == 'of type' or value_lower == 'must be':
            return True
        
        # Check for strings that are just punctuation or very short error indicators
        if len(value.strip()) <= 3 and value.strip() in ['...', '---', '***', '!!!', '???']:
            return True
        
        return False
    
    def _clean_cell_data(self, cell_data: dict) -> dict:
        """
        Clean individual cell data by removing null/empty values and fixing RGB properties
        
        Args:
            cell_data: Cell data dictionary
            
        Returns:
            Cleaned cell data
        """
        if not cell_data:
            return {}
        
        cleaned_cell = {}
        
        # Copy basic cell properties
        for key in ['value', 'formula', 'data_type', 'coordinate', 'row', 'column']:
            if key in cell_data and cell_data[key] is not None:
                cleaned_cell[key] = cell_data[key]
        
        # Clean style information
        if 'style' in cell_data and cell_data['style']:
            cleaned_style = self._clean_style_data(cell_data['style'])
            if cleaned_style:
                cleaned_cell['style'] = cleaned_style
        
        # Clean other properties
        for key in ['comment', 'hyperlink']:
            if key in cell_data and cell_data[key] is not None:
                cleaned_cell[key] = cell_data[key]
        
        return cleaned_cell
    
    def _clean_style_data(self, style_data: dict) -> dict:
        """
        Clean style data by removing null/empty values and fixing RGB properties
        
        Args:
            style_data: Style data dictionary
            
        Returns:
            Cleaned style data
        """
        if not style_data:
            return {}
        
        cleaned_style = {}
        
        # Clean font data
        if 'font' in style_data and style_data['font']:
            cleaned_font = self._clean_font_data(style_data['font'])
            if cleaned_font:
                cleaned_style['font'] = cleaned_font
        
        # Clean fill data
        if 'fill' in style_data and style_data['fill']:
            cleaned_fill = self._clean_fill_data(style_data['fill'])
            if cleaned_fill:
                cleaned_style['fill'] = cleaned_fill
        
        # Clean border data
        if 'border' in style_data and style_data['border']:
            cleaned_border = self._clean_border_data(style_data['border'])
            if cleaned_border:
                cleaned_style['border'] = cleaned_border
        
        # Clean alignment data
        if 'alignment' in style_data and style_data['alignment']:
            cleaned_alignment = self._clean_alignment_data(style_data['alignment'])
            if cleaned_alignment:
                cleaned_style['alignment'] = cleaned_alignment
        
        # Clean protection data
        if 'protection' in style_data and style_data['protection']:
            cleaned_protection = self._clean_protection_data(style_data['protection'])
            if cleaned_protection:
                cleaned_style['protection'] = cleaned_protection
        
        # Clean number format
        if 'number_format' in style_data and style_data['number_format']:
            cleaned_style['number_format'] = style_data['number_format']
        
        return cleaned_style
    
    def _clean_font_data(self, font_data: dict) -> dict:
        """Clean font data"""
        if not font_data:
            return {}
        
        cleaned_font = {}
        
        # Only include non-null, non-empty values
        for key in ['name', 'size', 'bold', 'italic', 'underline', 'strike', 'scheme']:
            if key in font_data and font_data[key] is not None:
                if isinstance(font_data[key], str) and font_data[key].strip() == "":
                    continue
                cleaned_font[key] = font_data[key]
        
        # Clean color data
        if 'color' in font_data and font_data['color']:
            cleaned_color = self._clean_color_data(font_data['color'])
            if cleaned_color:
                cleaned_font['color'] = cleaned_color
        
        return cleaned_font
    
    def _clean_fill_data(self, fill_data: dict) -> dict:
        """Clean fill data"""
        if not fill_data:
            return {}
        
        cleaned_fill = {}
        
        # Only include non-null, non-empty values
        for key in ['fill_type']:
            if key in fill_data and fill_data[key] is not None:
                if isinstance(fill_data[key], str) and fill_data[key].strip() == "":
                    continue
                cleaned_fill[key] = fill_data[key]
        
        # Clean color data
        for color_key in ['start_color', 'end_color']:
            if color_key in fill_data and fill_data[color_key]:
                cleaned_color = self._clean_color_data(fill_data[color_key])
                if cleaned_color:
                    cleaned_fill[color_key] = cleaned_color
        
        # Remove fill sections with default values
        # Check if this is a default fill (00000000 color and 0.0 tint)
        has_default_values = False
        if 'start_color' in cleaned_fill:
            start_color = cleaned_fill['start_color']
            if isinstance(start_color, dict) and start_color.get('rgb') == '00000000':
                tint = start_color.get('tint', 0.0)
                if tint == 0.0:
                    has_default_values = True
        
        if has_default_values:
            return {}
        
        return cleaned_fill
    
    def _clean_border_data(self, border_data: dict) -> dict:
        """Clean border data"""
        if not border_data:
            return {}
        
        cleaned_border = {}
        
        # Clean each border side
        for side in ['left', 'right', 'top', 'bottom', 'diagonal']:
            if side in border_data and border_data[side]:
                cleaned_side = self._clean_side_data(border_data[side])
                if cleaned_side:
                    cleaned_border[side] = cleaned_side
        
        # Clean diagonal direction
        if 'diagonal_direction' in border_data and border_data['diagonal_direction']:
            if isinstance(border_data['diagonal_direction'], str) and border_data['diagonal_direction'].strip() != "":
                cleaned_border['diagonal_direction'] = border_data['diagonal_direction']
        
        return cleaned_border
    
    def _clean_side_data(self, side_data: dict) -> dict:
        """Clean border side data"""
        if not side_data:
            return {}
        
        cleaned_side = {}
        
        # Only include non-null, non-empty values
        for key in ['style']:
            if key in side_data and side_data[key] is not None:
                if isinstance(side_data[key], str) and side_data[key].strip() == "":
                    continue
                cleaned_side[key] = side_data[key]
        
        # Clean color data
        if 'color' in side_data and side_data['color']:
            cleaned_color = self._clean_color_data(side_data['color'])
            if cleaned_color:
                cleaned_side['color'] = cleaned_color
        
        return cleaned_side
    
    def _clean_alignment_data(self, alignment_data: dict) -> dict:
        """Clean alignment data"""
        if not alignment_data:
            return {}
        
        cleaned_alignment = {}
        
        # Only include non-null values that are not default (0 or 0.0)
        for key in ['horizontal', 'vertical', 'text_rotation', 'wrap_text', 'shrink_to_fit', 
                   'indent', 'relative_indent', 'justify_last_line', 'reading_order']:
            if key in alignment_data and alignment_data[key] is not None:
                value = alignment_data[key]
                # Skip default values (0, 0.0, False for boolean fields)
                if value == 0 or value == 0.0 or value is False:
                    continue
                cleaned_alignment[key] = alignment_data[key]
        
        # If no meaningful alignment data remains, return empty dict
        if not cleaned_alignment:
            return {}
        
        return cleaned_alignment
    
    def _clean_protection_data(self, protection_data: dict) -> dict:
        """Clean protection data"""
        if not protection_data:
            return {}
        
        cleaned_protection = {}
        
        # Only include non-null values
        for key in ['locked', 'hidden']:
            if key in protection_data and protection_data[key] is not None:
                cleaned_protection[key] = protection_data[key]
        
        return cleaned_protection
    
    def _clean_color_data(self, color_data: dict) -> dict:
        """Clean color data and fix RGB property types"""
        if not color_data:
            return {}
        
        cleaned_color = {}
        
        # Fix RGB property - ensure it's a string
        if 'rgb' in color_data and color_data['rgb'] is not None:
            rgb_value = color_data['rgb']
            if isinstance(rgb_value, (list, tuple)):
                # Convert RGB list/tuple to string format
                cleaned_color['rgb'] = f"FF{rgb_value[0]:02X}{rgb_value[1]:02X}{rgb_value[2]:02X}"
            elif isinstance(rgb_value, str):
                cleaned_color['rgb'] = rgb_value
            elif isinstance(rgb_value, int):
                # Convert integer RGB to hex string
                cleaned_color['rgb'] = f"FF{(rgb_value >> 16) & 0xFF:02X}{(rgb_value >> 8) & 0xFF:02X}{rgb_value & 0xFF:02X}"
        
        # Clean theme and tint
        for key in ['theme', 'tint']:
            if key in color_data and color_data[key] is not None:
                cleaned_color[key] = color_data[key]
        
        return cleaned_color
    
    def _clean_metadata(self, metadata: dict) -> dict:
        """
        Clean metadata by removing null/empty values
        
        Args:
            metadata: Metadata dictionary
            
        Returns:
            Cleaned metadata dictionary
        """
        if not metadata:
            return {}
        
        cleaned_metadata = {}
        for key, value in metadata.items():
            # Skip null values and empty strings
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            
            # For boolean values, only include if they're True
            if isinstance(value, bool) and not value:
                continue
            
            cleaned_metadata[key] = value
        
        return cleaned_metadata
    
    def _detect_and_process_tables(self, sheet_data: dict, options: dict) -> List[Dict[str, Any]]:
        """
        Detect tables in a sheet and process them into table format
        
        Args:
            sheet_data: Sheet data from the original JSON
            options: Processing options
            
        Returns:
            List of table objects
        """
        tables = []
        
        # Get cells data and dimensions
        cells = sheet_data.get('cells', {})
        dimensions = sheet_data.get('dimensions', {})
        
        # Prepare options for new detection system
        detection_options = options.copy()
        detection_options['sheet_data'] = sheet_data
        
        # Prefer explicit gap-based splitting when requested
        force_gaps = options.get('table_detection', {}).get('use_gaps', False)
        table_regions: List[Dict[str, Any]] = []
        if force_gaps:
            table_regions = self._detect_regions_by_gaps(cells, dimensions, options)
        # Fallback to detector if no regions found
        if not table_regions:
            # Use new simplified detection system
            table_regions = self.detector.detect_tables(cells, dimensions, detection_options)
        
        # Process each detected table region
        for i, region in enumerate(table_regions):
            table = self._process_table_region(
                sheet_data, region, i, detection_options
            )
            if table:
                tables.append(table)
        
        # If no tables detected, create a default table for the entire sheet
        if not tables and cells:
            default_table = self._create_default_table(sheet_data, options)
            if default_table:
                tables.append(default_table)
        
        return tables
    

    

    

    

    

    

    

    

    
    def _process_table_region(self, sheet_data: dict, region: dict, 
                            table_index: int, options: dict) -> Optional[Dict[str, Any]]:
        """
        Process a table region into table format
        
        Args:
            sheet_data: Original sheet data
            region: Table region coordinates
            table_index: Index of the table
            options: Processing options
            
        Returns:
            Table object with columns and rows
        """
        cells = sheet_data.get('cells', {})
        merged_cells = sheet_data.get('merged_cells', [])
        
        # Extract table data
        table_cells = self._extract_table_cells(cells, region)
        if not table_cells:
            return None
        
        # Clean the table cells
        cleaned_table_cells = {}
        for cell_key, cell_data in table_cells.items():
            cleaned_cell = self._clean_cell_data(cell_data)
            if cleaned_cell:  # Only include non-empty cells
                cleaned_table_cells[cell_key] = cleaned_cell
        
        # Determine header rows and columns using new resolver
        normalized_cells = self.detector._normalize_cell_data(cleaned_table_cells)
        header_info = self.header_resolver.resolve_headers(normalized_cells, region, options)
        
        # Create table structure
        table = {
            'table_id': f"table_{table_index + 1}",
            'name': f"Table {table_index + 1}",
            'region': region,
            'header_info': header_info,
            'columns': self._create_columns(cleaned_table_cells, header_info, region),
            'rows': self._create_rows(cleaned_table_cells, header_info, region, sheet_data),
            'metadata': self._clean_metadata({
                'detection_method': region.get('detection_method', 'unknown'),
                'cell_count': len(cleaned_table_cells),
                'has_merged_cells': bool(merged_cells)
            })
        }
        
        return table

    def _detect_regions_by_gaps(self, cells: dict, dimensions: dict, options: dict) -> List[Dict[str, Any]]:
        """Detect table regions by splitting on blank row gaps.
        This is a lightweight splitter used when API callers explicitly request gap detection.
        """
        if not cells:
            return []
        min_row = dimensions.get('min_row', 1)
        max_row = dimensions.get('max_row', min_row)
        min_col = dimensions.get('min_col', 1)
        max_col = dimensions.get('max_col', min_col)

        # Collect data rows present in cells
        # Note: test fixtures often omit explicit 'row'/'column' fields.
        # Fall back to parsing the row index from the coordinate key (e.g., "A12" -> 12).
        data_rows_set = set()
        for coord, cell_info in cells.items():
            row_num = None
            if isinstance(cell_info, dict):
                row_num = cell_info.get('row')
            if row_num is None and isinstance(coord, str):
                # Extract digits from coordinate as the row number
                import re
                match = re.search(r"(\d+)$", coord)
                if match:
                    try:
                        row_num = int(match.group(1))
                    except ValueError:
                        row_num = None
            if row_num is not None:
                data_rows_set.add(row_num)

        data_rows = sorted(data_rows_set)
        if len(data_rows) < 2:
            # Single row of data → single region
            return [{
                'start_row': min_row,
                'end_row': max_row,
                'start_col': min_col,
                'end_col': max_col,
                'detection_method': 'gaps'
            }]

        # Default to requiring at least 2 consecutive blank rows to consider a split.
        # This prevents splitting titles/metadata that are often separated by a single blank row
        # from the actual table header/data, as expected by unit tests.
        gap_threshold = options.get('table_detection', {}).get('gap_threshold', 2)
        regions: List[Dict[str, Any]] = []
        current_start = data_rows[0]

        for i in range(1, len(data_rows)):
            gap_size = data_rows[i] - data_rows[i - 1] - 1
            if gap_size >= gap_threshold:
                regions.append({
                    'start_row': current_start,
                    'end_row': data_rows[i - 1],
                    'start_col': min_col,
                    'end_col': max_col,
                    'detection_method': 'gaps'
                })
                current_start = data_rows[i]

        # Final region to the last data row (or max_row to satisfy downstream assumptions)
        regions.append({
            'start_row': current_start,
            'end_row': max_row,
            'start_col': min_col,
            'end_col': max_col,
            'detection_method': 'gaps'
        })

        return regions
    
    def _extract_table_cells(self, cells: dict, region: dict) -> Dict[str, Any]:
        """Extract cells that belong to the table region"""
        table_cells = {}
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell_key = f"{get_column_letter(col)}{row}"
                if cell_key in cells:
                    table_cells[cell_key] = cells[cell_key]
        
        return table_cells
    
    def _determine_headers(self, table_cells: dict, region: dict, 
                         merged_cells: List[dict], options: dict) -> Dict[str, Any]:
        """Determine header rows and columns for the table"""
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        # Get frozen panes information from sheet data
        sheet_data = options.get('sheet_data', {})
        frozen_panes = sheet_data.get('frozen_panes', {})
        frozen_rows = frozen_panes.get('frozen_rows', 0)
        frozen_cols = frozen_panes.get('frozen_cols', 0)
        
        # Determine header rows: include all frozen rows plus the first non-frozen row
        header_rows = []
        if frozen_rows > 0:
            # Add all frozen rows
            for row in range(start_row, start_row + frozen_rows):
                if row <= end_row:
                    header_rows.append(row)
        
        # If no frozen rows or we need more header rows, add the first row
        if not header_rows and start_row < end_row:
            header_rows.append(start_row)
        
        # Determine header columns: include all frozen columns plus the first non-frozen column
        header_cols = []
        if frozen_cols > 0:
            # Add all frozen columns
            for col in range(start_col, start_col + frozen_cols):
                if col <= end_col:
                    header_cols.append(col)
        
        # If no frozen columns or we need more header columns, add the first column
        if not header_cols and start_col < end_col:
            header_cols.append(start_col)
        
        return {
            'header_rows': header_rows,
            'header_columns': header_cols,
            'data_start_row': start_row + len(header_rows),
            'data_start_col': start_col + len(header_cols)
        }
    
    def _create_columns(self, table_cells: dict, header_info: dict, region: dict) -> List[Dict[str, Any]]:
        """Create column definitions with labels"""
        columns = []
        start_col = region['start_col']
        end_col = region['end_col']
        header_rows = header_info['header_rows']
        
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            
            # Get column label from header rows
            column_label = self._get_column_label(table_cells, col, header_rows)
            
            column_def = {
                'column_index': col,
                'column_letter': col_letter,
                'column_label': column_label,
                'is_header_column': col in header_info['header_columns'],
                'cells': {}
            }
            
            # Only add width and hidden if they have meaningful values
            # (These would be populated from sheet column properties if available)
            
            # Add cells for this column
            for row in range(region['start_row'], region['end_row'] + 1):
                cell_key = f"{col_letter}{row}"
                if cell_key in table_cells:
                    # Cell data is already cleaned in _process_table_region
                    column_def['cells'][cell_key] = table_cells[cell_key]
            
            columns.append(column_def)
        
        return columns
    
    def _create_rows(self, table_cells: dict, header_info: dict, region: dict, 
                    sheet_data: dict = None) -> List[Dict[str, Any]]:
        """Create row definitions with labels"""
        rows = []
        start_row = region['start_row']
        end_row = region['end_row']
        header_cols = header_info['header_columns']
        header_rows = header_info['header_rows']
        
        for row in range(start_row, end_row + 1):
            # Get row label from header columns, considering frozen rows
            row_label = self._get_row_label(table_cells, row, header_cols, header_rows, sheet_data)
            
            row_def = {
                'row_index': row,
                'row_label': row_label,
                'is_header_row': row in header_info['header_rows'],
                'cells': {}
            }
            
            # Only add height and hidden if they have meaningful values
            # (These would be populated from sheet row properties if available)
            
            # Add cells for this row
            for col in range(region['start_col'], region['end_col'] + 1):
                col_letter = get_column_letter(col)
                cell_key = f"{col_letter}{row}"
                if cell_key in table_cells:
                    # Cell data is already cleaned in _process_table_region
                    row_def['cells'][cell_key] = table_cells[cell_key]
            
            rows.append(row_def)
        
        return rows
    
    def _get_column_label(self, table_cells: dict, col: int, header_rows: List[int]) -> str:
        """Get the label for a column from header rows"""
        col_letter = get_column_letter(col)
        labels = []
        
        for header_row in header_rows:
            cell_key = f"{col_letter}{header_row}"
            if cell_key in table_cells:
                cell_value = table_cells[cell_key].get('value')
                if cell_value is not None:
                    labels.append(str(cell_value))
        
        # Reverse the order so most specific/recent value comes first (e.g., "Jan 2023" not "2023 Jan")
        labels.reverse()
        
        return " ".join(labels) if labels else "unlabeled"
    
    def _get_row_label(self, table_cells: dict, row: int, header_cols: List[int], 
                      header_rows: List[int] = None, sheet_data: dict = None) -> str:
        """Get the label for a row from header columns, considering frozen rows"""
        labels = []
        
        # Get frozen panes information if available
        frozen_rows = 0
        if sheet_data:
            frozen_panes = sheet_data.get('frozen_panes', {})
            frozen_rows = frozen_panes.get('frozen_rows', 0)
        
        for header_col in header_cols:
            col_letter = get_column_letter(header_col)
            column_labels = []
            
            # If we have frozen rows and header rows, get values from all frozen rows for this column
            if frozen_rows > 0 and header_rows:
                for header_row in header_rows:
                    cell_key = f"{col_letter}{header_row}"
                    if cell_key in table_cells:
                        cell_value = table_cells[cell_key].get('value')
                        if cell_value is not None:
                            column_labels.append(str(cell_value))
                # Reverse the order so most specific/recent value comes first (e.g., "Jan 2023" not "2023 Jan")
                column_labels.reverse()
            else:
                # Fallback to original logic: just get the value from the current row
                cell_key = f"{col_letter}{row}"
                if cell_key in table_cells:
                    cell_value = table_cells[cell_key].get('value')
                    if cell_value is not None:
                        column_labels.append(str(cell_value))
            
            # Join the column labels with space (e.g., "Jan 2023")
            if column_labels:
                labels.append(" ".join(column_labels))
        
        return " | ".join(labels) if labels else "unlabeled"
    
    def _create_default_table(self, sheet_data: dict, options: dict) -> Optional[Dict[str, Any]]:
        """Create a default table for the entire sheet if no tables detected"""
        cells = sheet_data.get('cells', {})
        if not cells:
            return None
        
        dimensions = sheet_data.get('dimensions', {})
        region = {
            'start_row': dimensions.get('min_row', 1),
            'end_row': dimensions.get('max_row', 1),
            'start_col': dimensions.get('min_col', 1),
            'end_col': dimensions.get('max_col', 1),
            'detection_method': 'default'
        }
        
        return self._process_table_region(sheet_data, region, 0, options) 
    

    
 