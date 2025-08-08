"""
Excel Complexity Analyzer

This module analyzes Excel sheet complexity to determine when traditional heuristics
will struggle and AI-powered analysis should be used instead.
"""

import re
from typing import Dict, Any, List, Optional, Tuple
from openpyxl.utils import get_column_letter


class ExcelComplexityAnalyzer:
    """
    Analyzes Excel sheet complexity to determine processing approach.
    
    Scoring is based on multiple factors:
    - Merged cell complexity (0.0-0.3)
    - Header hierarchy complexity (0.0-0.25) 
    - Data sparsity complexity (0.0-0.2)
    - Column structure inconsistency (0.0-0.15)
    - Formula complexity (0.0-0.1)
    """
    
    def __init__(self, thresholds: Optional[Dict[str, float]] = None):
        self.thresholds = thresholds or self._default_thresholds()
    
    def _default_thresholds(self) -> Dict[str, float]:
        """Default complexity thresholds for processing decisions"""
        return {
            'simple_threshold': 0.3,      # Below this: traditional only
            'complex_threshold': 0.7,     # Above this: AI-first
            'merge_ratio_threshold': 0.1,  # Significant merge complexity
            'sparsity_threshold': 0.5,     # High sparsity indicator
            'header_levels_threshold': 2,  # Multi-level headers
        }
    
    def analyze_sheet_complexity(self, sheet_data: Dict[str, Any], 
                                 complexity_metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Analyze Excel sheet complexity and return comprehensive analysis.
        
        Args:
            sheet_data: Sheet data from Excel processor
            complexity_metadata: Optional metadata for enhanced analysis
            
        Returns:
            Dictionary with complexity analysis results
        """
        # Use metadata if available for more accurate analysis
        if complexity_metadata:
            metrics = self._analyze_with_metadata(sheet_data, complexity_metadata)
        else:
            # Calculate individual complexity metrics (original method)
            metrics = {
                'merged_cell_complexity': self._analyze_merged_cells(sheet_data),
                'header_complexity': self._analyze_header_structure(sheet_data),
                'sparsity_complexity': self._analyze_data_sparsity(sheet_data),
                'column_inconsistency': self._analyze_column_structure(sheet_data),
                'formula_complexity': self._analyze_formula_patterns(sheet_data)
            }
        
        # Calculate overall complexity score
        overall_score = self._calculate_complexity_score(metrics)
        
        # Determine complexity level and processing recommendation
        complexity_level = self._get_complexity_level(overall_score)
        recommendation = self._get_processing_recommendation(overall_score)
        
        # Identify specific failure indicators
        failure_indicators = self._identify_failure_indicators(metrics, sheet_data)
        
        return {
            'complexity_score': overall_score,
            'complexity_level': complexity_level,
            'metrics': metrics,
            'recommendation': recommendation,
            'failure_indicators': failure_indicators,
            'processing_metadata': {
                'sheet_name': sheet_data.get('name', 'Unknown'),
                'dimensions': sheet_data.get('dimensions', {}),
                'analysis_timestamp': self._get_timestamp()
            }
        }
    
    def _analyze_merged_cells(self, sheet_data: Dict[str, Any]) -> float:
        """
        Analyze complexity from merged cell patterns.
        
        Returns score 0.0-1.0 based on:
        - Ratio of merged cells to total data cells
        - Complexity of individual merges (multi-row/column spans)
        - Overlap patterns
        """
        merged_cells = sheet_data.get('merged_cells', [])
        total_data_cells = self._estimate_data_cell_count(sheet_data)
        
        if total_data_cells == 0 or not merged_cells:
            return 0.0
        
        # Base complexity from merge ratio
        merge_ratio = len(merged_cells) / total_data_cells
        complexity_score = min(merge_ratio * 3, 0.3)  # Cap at 0.3
        
        # Additional complexity for complex merge patterns
        complex_merges = 0
        for merge in merged_cells:
            rows_spanned = merge.get('end_row', 0) - merge.get('start_row', 0) + 1
            cols_spanned = merge.get('end_column', 0) - merge.get('start_column', 0) + 1
            
            # Complex if spans more than 2 rows or columns
            if rows_spanned > 2 or cols_spanned > 2:
                complex_merges += 1
            
            # Very complex if spans both multiple rows AND columns
            if rows_spanned > 1 and cols_spanned > 1:
                complex_merges += 0.5
        
        # Add complexity bonus for complex patterns
        if len(merged_cells) > 0:
            complex_ratio = complex_merges / len(merged_cells)
            complexity_score += complex_ratio * 0.2
        
        return min(complexity_score, 1.0)
    
    def _analyze_header_structure(self, sheet_data: Dict[str, Any]) -> float:
        """
        Analyze multi-level header complexity.
        
        Returns score 0.0-1.0 based on:
        - Number of header levels detected
        - Inconsistencies in header patterns
        - Merged headers across levels
        """
        cells = sheet_data.get('cells', {})
        dimensions = sheet_data.get('dimensions', {})
        
        if not cells:
            return 0.0
        
        # Detect header levels by analyzing first few rows
        max_row_to_check = min(dimensions.get('max_row', 10), 10)
        header_levels = self._detect_header_levels(cells, max_row_to_check)
        
        # Check for header inconsistencies
        inconsistencies = self._detect_header_inconsistencies(cells, header_levels, dimensions)
        
        complexity = 0.0
        
        # Score based on number of header levels
        if header_levels > 2:
            complexity += min((header_levels - 2) * 0.1, 0.15)
        
        # Add complexity for inconsistencies
        complexity += min(inconsistencies * 0.1, 0.1)
        
        return min(complexity, 1.0)
    
    def _analyze_data_sparsity(self, sheet_data: Dict[str, Any]) -> float:
        """
        Analyze data sparsity complexity.
        
        Returns score 0.0-1.0 based on:
        - Ratio of empty to total addressable cells
        - Clustering vs scattered data patterns
        - Large empty regions
        """
        dimensions = sheet_data.get('dimensions', {})
        cells = sheet_data.get('cells', {})
        
        if not dimensions or not cells:
            return 0.0
        
        # Calculate addressable space
        total_addressable = (
            (dimensions.get('max_row', 1) - dimensions.get('min_row', 1) + 1) *
            (dimensions.get('max_col', 1) - dimensions.get('min_col', 1) + 1)
        )
        
        if total_addressable == 0:
            return 0.0
        
        # Calculate actual data cells
        data_cells = len([cell for cell in cells.values() 
                         if cell.get('value') is not None and str(cell.get('value', '')).strip()])
        
        sparsity_ratio = 1.0 - (data_cells / total_addressable)
        
        complexity = 0.0
        
        # High sparsity contributes to complexity
        if sparsity_ratio > self.thresholds['sparsity_threshold']:
            excess_sparsity = sparsity_ratio - self.thresholds['sparsity_threshold']
            complexity += min(excess_sparsity * 0.4, 0.2)
        
        return min(complexity, 1.0)
    
    def _analyze_column_structure(self, sheet_data: Dict[str, Any]) -> float:
        """
        Analyze column structure consistency.
        
        Returns score 0.0-1.0 based on:
        - Variance in column data types
        - Inconsistent column patterns across rows
        - Mixed content types in columns
        """
        cells = sheet_data.get('cells', {})
        dimensions = sheet_data.get('dimensions', {})
        
        if not cells or not dimensions:
            return 0.0
        
        # Analyze each column for consistency
        min_col = dimensions.get('min_col', 1)
        max_col = dimensions.get('max_col', 1)
        min_row = dimensions.get('min_row', 1)
        max_row = dimensions.get('max_row', 1)
        
        column_inconsistencies = []
        
        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            col_values = []
            
            for row in range(min_row, max_row + 1):
                coord = f"{col_letter}{row}"
                if coord in cells:
                    value = cells[coord].get('value')
                    if value is not None:
                        col_values.append(value)
            
            if len(col_values) > 1:
                inconsistency = self._calculate_column_inconsistency(col_values)
                column_inconsistencies.append(inconsistency)
        
        if not column_inconsistencies:
            return 0.0
        
        # Average inconsistency across columns
        avg_inconsistency = sum(column_inconsistencies) / len(column_inconsistencies)
        return min(avg_inconsistency, 1.0)
    
    def _analyze_formula_patterns(self, sheet_data: Dict[str, Any]) -> float:
        """
        Analyze formula complexity.
        
        Returns score 0.0-1.0 based on:
        - Ratio of formula cells to total cells
        - Complexity of individual formulas
        - Dependencies and dynamic ranges
        """
        cells = sheet_data.get('cells', {})
        
        if not cells:
            return 0.0
        
        formula_count = 0
        complex_formula_count = 0
        total_cells = len(cells)
        
        for cell in cells.values():
            formula = cell.get('formula')
            if formula:
                formula_count += 1
                
                # Check for complex formula patterns
                if self._is_complex_formula(formula):
                    complex_formula_count += 1
        
        if total_cells == 0:
            return 0.0
        
        # Base complexity from formula ratio
        formula_ratio = formula_count / total_cells
        complexity = min(formula_ratio * 0.5, 0.05)
        
        # Additional complexity for complex formulas
        if formula_count > 0:
            complex_ratio = complex_formula_count / formula_count
            complexity += min(complex_ratio * 0.05, 0.05)
        
        return min(complexity, 1.0)
    
    def _analyze_with_metadata(self, sheet_data: Dict[str, Any], 
                              complexity_metadata: Dict[str, Any]) -> Dict[str, float]:
        """
        Analyze complexity using preserved metadata for more accurate results.
        
        This method can provide complexity analysis even when the original cell data
        has been compressed or filtered.
        """
        metrics = {}
        
        # 1. Merged Cell Complexity from metadata
        merged_meta = complexity_metadata.get('merged_cells', {})
        total_merges = merged_meta.get('count', 0)
        complex_merges = merged_meta.get('complex_merges', 0)
        cell_count = complexity_metadata.get('cell_count', 1)
        
        if cell_count > 0 and total_merges > 0:
            merge_ratio = total_merges / cell_count
            complexity_score = min(merge_ratio * 3, 0.3)
            
            # Add complexity for complex merge patterns
            if total_merges > 0:
                complex_ratio = complex_merges / total_merges
                complexity_score += complex_ratio * 0.2
            
            metrics['merged_cell_complexity'] = min(complexity_score, 1.0)
        else:
            metrics['merged_cell_complexity'] = 0.0
        
        # 2. Header Complexity from metadata
        header_meta = complexity_metadata.get('header_structure', {})
        header_levels = header_meta.get('detected_levels', 1)
        inconsistency = header_meta.get('inconsistency_score', 0.0)
        
        complexity = 0.0
        if header_levels > 2:
            complexity += min((header_levels - 2) * 0.1, 0.15)
        complexity += min(inconsistency * 0.1, 0.1)
        
        metrics['header_complexity'] = min(complexity, 1.0)
        
        # 3. Sparsity Complexity from metadata
        distribution_meta = complexity_metadata.get('data_distribution', {})
        sparsity = distribution_meta.get('sparsity', 0.0)
        is_clustered = distribution_meta.get('clustered', False)
        
        complexity = 0.0
        if sparsity > self.thresholds['sparsity_threshold']:
            excess_sparsity = sparsity - self.thresholds['sparsity_threshold']
            complexity += min(excess_sparsity * 0.4, 0.2)
        
        # Additional complexity for clustered data
        if is_clustered:
            complexity += 0.05
        
        metrics['sparsity_complexity'] = min(complexity, 1.0)
        
        # 4. Column Inconsistency from metadata
        column_meta = complexity_metadata.get('columns', {})
        avg_inconsistency = column_meta.get('average_inconsistency', 0.0)
        high_inconsistency_cols = column_meta.get('high_inconsistency_columns', 0)
        
        # Base inconsistency score
        complexity = avg_inconsistency
        
        # Additional penalty for many inconsistent columns
        if high_inconsistency_cols > 0:
            complexity += min(high_inconsistency_cols * 0.02, 0.05)
        
        metrics['column_inconsistency'] = min(complexity, 1.0)
        
        # 5. Formula Complexity from metadata
        formula_meta = complexity_metadata.get('formulas', {})
        formula_ratio = formula_meta.get('formula_ratio', 0.0)
        complex_ratio = formula_meta.get('complex_ratio', 0.0)
        
        # Base complexity from formula ratio
        complexity = min(formula_ratio * 0.5, 0.05)
        
        # Additional complexity for complex formulas
        complexity += min(complex_ratio * 0.05, 0.05)
        
        metrics['formula_complexity'] = min(complexity, 1.0)
        
        return metrics
    
    def _calculate_complexity_score(self, metrics: Dict[str, float]) -> float:
        """Calculate overall complexity score from individual metrics."""
        return min(sum(metrics.values()), 1.0)
    
    def _get_complexity_level(self, score: float) -> str:
        """Get complexity level string from score."""
        if score < self.thresholds['simple_threshold']:
            return 'simple'
        elif score < self.thresholds['complex_threshold']:
            return 'moderate'
        else:
            return 'complex'
    
    def _get_processing_recommendation(self, score: float) -> str:
        """Get processing recommendation from complexity score."""
        if score < self.thresholds['simple_threshold']:
            return 'traditional'
        elif score < self.thresholds['complex_threshold']:
            return 'dual'  # Both traditional and AI for comparison
        else:
            return 'ai_first'
    
    def _identify_failure_indicators(self, metrics: Dict[str, float], 
                                   sheet_data: Dict[str, Any]) -> List[str]:
        """Identify specific indicators that suggest traditional methods will fail."""
        indicators = []
        
        # Merged cell indicators
        if metrics['merged_cell_complexity'] > 0.15:
            indicators.append('complex_merged_cells')
        
        # Header complexity indicators  
        if metrics['header_complexity'] > 0.15:
            indicators.append('multi_level_headers')
        
        # Sparsity indicators
        if metrics['sparsity_complexity'] > 0.1:
            indicators.append('high_data_sparsity')
        
        # Column structure indicators
        if metrics['column_inconsistency'] > 0.1:
            indicators.append('inconsistent_column_structure')
        
        # Formula complexity indicators
        if metrics['formula_complexity'] > 0.05:
            indicators.append('complex_formulas')
        
        # Check for specific problematic patterns
        merged_cells = sheet_data.get('merged_cells', [])
        if len(merged_cells) > 10:
            indicators.append('extensive_merging')
        
        return indicators
    
    def _estimate_data_cell_count(self, sheet_data: Dict[str, Any]) -> int:
        """Estimate the number of data cells in the sheet."""
        cells = sheet_data.get('cells', {})
        return len([cell for cell in cells.values() 
                   if cell.get('value') is not None and str(cell.get('value', '')).strip()])
    
    def _detect_header_levels(self, cells: Dict[str, Any], max_row: int) -> int:
        """Detect the number of header levels in the first few rows."""
        # This is a simplified detection - can be enhanced later
        header_levels = 1
        
        # Look for patterns that suggest multiple header levels
        for row in range(1, min(max_row + 1, 6)):  # Check first 5 rows
            row_cells = [cell for coord, cell in cells.items() 
                        if coord.endswith(str(row))]
            
            if len(row_cells) > 0:
                # Check if this looks like a header row
                text_cells = sum(1 for cell in row_cells 
                               if isinstance(cell.get('value'), str))
                
                if text_cells / len(row_cells) > 0.7:  # Mostly text = likely header
                    header_levels = max(header_levels, row)
        
        return min(header_levels, 5)  # Cap at 5 levels
    
    def _detect_header_inconsistencies(self, cells: Dict[str, Any], 
                                     header_levels: int, dimensions: Dict[str, Any]) -> float:
        """Detect inconsistencies in header patterns."""
        # Simplified inconsistency detection
        inconsistencies = 0.0
        
        # Check for gaps in header rows
        min_col = dimensions.get('min_col', 1)
        max_col = dimensions.get('max_col', 1)
        
        for row in range(1, header_levels + 1):
            gaps = 0
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                if coord not in cells or not cells[coord].get('value'):
                    gaps += 1
            
            if gaps > 0:
                gap_ratio = gaps / (max_col - min_col + 1)
                inconsistencies += gap_ratio * 0.1
        
        return min(inconsistencies, 1.0)
    
    def _calculate_column_inconsistency(self, col_values: List[Any]) -> float:
        """Calculate inconsistency score for a column's values."""
        if len(col_values) <= 1:
            return 0.0
        
        # Analyze data types
        type_counts = {}
        for value in col_values:
            value_type = type(value).__name__
            type_counts[value_type] = type_counts.get(value_type, 0) + 1
        
        # Calculate inconsistency based on type diversity
        if len(type_counts) == 1:
            return 0.0  # All same type = consistent
        
        # More types = more inconsistency
        inconsistency = (len(type_counts) - 1) / len(col_values)
        return min(inconsistency, 1.0)
    
    def _is_complex_formula(self, formula: str) -> bool:
        """Check if a formula is complex based on patterns."""
        if not formula:
            return False
        
        # Complex formula indicators
        complex_patterns = [
            r'IF.*IF',      # Nested IFs
            r'VLOOKUP',     # Lookup functions
            r'INDEX.*MATCH', # Index/Match combinations
            r'SUMPRODUCT',  # Array functions
            r'\$[A-Z]+\$\d+:\$[A-Z]+\$\d+', # Absolute ranges
        ]
        
        for pattern in complex_patterns:
            if re.search(pattern, formula, re.IGNORECASE):
                return True
        
        return False
    
    def _get_timestamp(self) -> str:
        """Get current timestamp for metadata."""
        from datetime import datetime
        return datetime.now().isoformat()
