import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


class ExcelProcessor:
    """Comprehensive Excel to JSON converter using openpyxl"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """
        Process an Excel file and return comprehensive JSON representation
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing the complete Excel structure
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension not in self.supported_extensions:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
        try:
            # First pass: get all data including formulas (data_only=False)
            workbook_with_formulas = load_workbook(file_path, data_only=False, keep_vba=True)
            
            # Second pass: get calculated values (data_only=True)
            workbook_with_values = load_workbook(file_path, data_only=True, keep_vba=True)
            
            return self._process_workbook(workbook_with_formulas, workbook_with_values, file_path)
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def _process_workbook(self, workbook_with_formulas: Workbook, workbook_with_values: Workbook, file_path: str) -> Dict[str, Any]:
        """Process the entire workbook with both formula and value data"""
        file_stats = os.stat(file_path)
        
        result = {
            "workbook": {
                "metadata": self._extract_workbook_metadata(workbook_with_formulas, file_path, file_stats),
                "sheets": [],
                "properties": self._extract_workbook_properties(workbook_with_formulas)
            }
        }
        
        # Process each worksheet
        for sheet_name in workbook_with_formulas.sheetnames:
            worksheet_with_formulas = workbook_with_formulas[sheet_name]
            worksheet_with_values = workbook_with_values[sheet_name]
            
            # Check if it's a worksheet (not a chart sheet)
            if hasattr(worksheet_with_formulas, 'title') and not hasattr(worksheet_with_formulas, 'chart'):
                sheet_data = self._process_worksheet(worksheet_with_formulas, worksheet_with_values)
                result["workbook"]["sheets"].append(sheet_data)
        
        # Final safeguard: ensure the entire result is JSON serializable
        return self._safe(result)
    
    def _extract_workbook_metadata(self, workbook: Workbook, file_path: str, file_stats) -> Dict[str, Any]:
        """Extract workbook-level metadata"""
        props = workbook.properties
        
        metadata = {
            "filename": os.path.basename(file_path),
            "file_size": file_stats.st_size,
            "created_date": self._format_datetime(props.created) if props.created else None,
            "modified_date": self._format_datetime(props.modified) if props.modified else None,
            "creator": getattr(props, 'creator', '') or "",
            "last_modified_by": getattr(props, 'lastModifiedBy', '') or "",
            "title": getattr(props, 'title', '') or "",
            "subject": getattr(props, 'subject', '') or "",
            "keywords": getattr(props, 'keywords', '') or "",
            "category": getattr(props, 'category', '') or "",
            "description": getattr(props, 'description', '') or "",
            "language": getattr(props, 'language', '') or "",
            "revision": getattr(props, 'revision', 0) or 0,
            "version": getattr(props, 'version', '') or "",
            "application": getattr(props, 'application', '') or "",
            "app_version": getattr(props, 'appName', '') or "",
            "company": getattr(props, 'company', '') or "",
            "manager": getattr(props, 'manager', '') or "",
            "hyperlink_base": getattr(props, 'hyperlinkBase', '') or "",
            "template": getattr(props, 'template', '') or "",
            "status": getattr(props, 'status', '') or "",
            "total_editing_time": getattr(props, 'totalTime', 0) or 0,
            "pages": getattr(props, 'pages', 0) or 0,
            "words": getattr(props, 'words', 0) or 0,
            "characters": getattr(props, 'characters', 0) or 0,
            "characters_with_spaces": getattr(props, 'charactersWithSpaces', 0) or 0,
            "application_name": getattr(props, 'application', '') or "",
            "security": {
                "workbook_password": bool(workbook.security),
                "revision_password": False,  # openpyxl doesn't expose this directly
                "lock_structure": workbook.security.lockStructure if workbook.security else False,
                "lock_windows": workbook.security.lockWindows if workbook.security else False
            }
        }
        
        return metadata
    
    def _extract_workbook_properties(self, workbook: Workbook) -> Dict[str, Any]:
        """Extract workbook properties"""
        return {
            "code_name": getattr(workbook, 'code_name', '') or "",
            "date_1904": getattr(workbook, 'excel_base_date', datetime(1900, 1, 1)) == datetime(1904, 1, 1),
            "filter_mode": getattr(workbook, 'filter_mode', False),
            "hide_pivot_chart_list": getattr(workbook, 'hide_pivot_chart_list', False),
            "published": getattr(workbook, 'published', False),
            "refresh_all_connections": getattr(workbook, 'refresh_all_connections', False),
            "save_external_link_values": getattr(workbook, 'save_external_link_values', False),
            "update_links": str(getattr(workbook, 'update_links', '')) if getattr(workbook, 'update_links', None) else "",
            "check_compatibility": getattr(workbook, 'check_compatibility', False),
            "auto_compress_pictures": getattr(workbook, 'auto_compress_pictures', False),
            "backup_file": getattr(workbook, 'backup_file', False),
            "check_calculations": getattr(workbook, 'check_calculations', False),
            "create_backup": getattr(workbook, 'create_backup', False),
            "crash_save": getattr(workbook, 'crash_save', False),
            "data_extract_load": getattr(workbook, 'data_extract_load', False),
            "default_theme_version": getattr(workbook, 'default_theme_version', 0),
            "delete_inactive_worksheets": getattr(workbook, 'delete_inactive_worksheets', False),
            "display_ink_annotations": getattr(workbook, 'display_ink_annotations', False),
            "first_sheet": getattr(workbook, 'first_sheet', 0),
            "minimized": getattr(workbook, 'minimized', False),
            "prevent_update": getattr(workbook, 'prevent_update', False),
            "show_ink_annotations": getattr(workbook, 'show_ink_annotations', False),
            "show_pivot_chart_filter": getattr(workbook, 'show_pivot_chart_filter', False),
            "update_remote_references": getattr(workbook, 'update_remote_references', False),
            "window_height": getattr(workbook, 'window_height', 600),
            "window_width": getattr(workbook, 'window_width', 800),
            "window_minimized": getattr(workbook, 'window_minimized', False),
            "window_maximized": getattr(workbook, 'window_maximized', False),
            "window_x": getattr(workbook, 'window_x', 0),
            "window_y": getattr(workbook, 'window_y', 0),
            "active_sheet_index": getattr(workbook, 'active', 0),
            "first_visible_sheet": getattr(workbook, 'first_visible_sheet', 0),
            "tab_ratio": getattr(workbook, 'tab_ratio', 600),
            "visibility": getattr(workbook, 'visibility', 'visible')
        }
    
    def _process_worksheet(self, worksheet_with_formulas: Worksheet, worksheet_with_values: Worksheet) -> Dict[str, Any]:
        """Process a single worksheet with both formula and value data"""
        sheet_data = {
            "name": worksheet_with_formulas.title,
            "index": getattr(worksheet_with_formulas.sheet_properties.tabColor, 'index', 0) if worksheet_with_formulas.sheet_properties.tabColor else 0,
            "sheet_id": getattr(worksheet_with_formulas.sheet_properties, 'sheetId', ''),
            "state": getattr(worksheet_with_formulas, 'sheet_state', 'visible'),
            "sheet_type": "worksheet",
            "properties": self._extract_sheet_properties(worksheet_with_formulas),
            "dimensions": self._extract_dimensions(worksheet_with_formulas),
            "frozen_panes": self._extract_frozen_panes(worksheet_with_formulas),
            "views": self._extract_sheet_views(worksheet_with_formulas),
            "protection": self._extract_sheet_protection(worksheet_with_formulas),
            "rows": self._extract_row_properties(worksheet_with_formulas),
            "columns": self._extract_column_properties(worksheet_with_formulas),
            "cells": self._extract_cell_data(worksheet_with_formulas, worksheet_with_values),
            "merged_cells": self._extract_merged_cells(worksheet_with_formulas),
            "data_validations": self._extract_data_validations(worksheet_with_formulas),
            "conditional_formatting": self._extract_conditional_formatting(worksheet_with_formulas),
            "charts": self._extract_charts(worksheet_with_formulas),
            "images": self._extract_images(worksheet_with_formulas),
            "comments": self._extract_comments(worksheet_with_formulas)
        }
        
        return sheet_data
    
    def _extract_sheet_properties(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract sheet properties"""
        props = worksheet.sheet_properties
        return {
            "code_name": props.codeName or "",
            "enable_format_conditions_calculation": props.enableFormatConditionsCalculation,
            "filter_mode": props.filterMode,
            "published": props.published,
            "sync_horizontal": props.syncHorizontal,
            "sync_ref": props.syncRef or "",
            "sync_vertical": props.syncVertical,
            "transition_evaluation": props.transitionEvaluation,
            "transition_entry": props.transitionEntry,
            "tab_color": self._extract_color(props.tabColor) if props.tabColor else None
        }
    
    def _extract_dimensions(self, worksheet: Worksheet) -> Dict[str, int]:
        """Extract worksheet dimensions"""
        return {
            "min_row": worksheet.min_row or 1,
            "max_row": worksheet.max_row or 1,
            "min_col": worksheet.min_column or 1,
            "max_col": worksheet.max_column or 1
        }
    
    def _extract_frozen_panes(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract frozen pane information"""
        if not worksheet.freeze_panes:
            return {
                "frozen_rows": 0,
                "frozen_cols": 0,
                "top_left_cell": None
            }
        
        frozen_cell = worksheet.freeze_panes
        if isinstance(frozen_cell, str):
            # Handle string format like "A2"
            try:
                if ':' in frozen_cell:
                    parts = frozen_cell.split(':')
                    if len(parts) == 2:
                        col_part = parts[0]
                        row_part = parts[1]
                        col_idx = column_index_from_string(col_part)
                        row_idx = int(row_part)
                    else:
                        col_idx = 1
                        row_idx = 1
                else:
                    # Extract column and row from cell reference like "A2"
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', frozen_cell)
                    if match:
                        col_letter, row_num = match.groups()
                        col_idx = column_index_from_string(col_letter)
                        row_idx = int(row_num)
                    else:
                        # Try to parse as column-only reference
                        col_idx = column_index_from_string(frozen_cell)
                        row_idx = 1
            except:
                col_idx = 1
                row_idx = 1
        else:
            col_idx, row_idx = frozen_cell
        
        return {
            "frozen_rows": row_idx - 1 if row_idx > 0 else 0,
            "frozen_cols": col_idx - 1 if col_idx > 0 else 0,
            "top_left_cell": f"{get_column_letter(col_idx)}{row_idx}"
        }
    
    def _extract_sheet_views(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract sheet view information"""
        views = []
        for view in worksheet.views:
            view_data = {
                "state": view.state,
                "zoom_scale": view.zoomScale,
                "zoom_scale_normal": view.zoomScaleNormal,
                "zoom_scale_page_layout_view": view.zoomScalePageLayoutView,
                "zoom_scale_sheet_layout_view": view.zoomScaleSheetLayoutView,
                "tab_selected": view.tabSelected,
                "show_grid_lines": view.showGridLines,
                "show_row_col_headers": view.showRowColHeaders,
                "show_ruler": view.showRuler,
                "show_zeros": view.showZeros,
                "right_to_left": view.rightToLeft,
                "show_outline_symbols": view.showOutlineSymbols,
                "default_grid_color": view.defaultGridColor,
                "show_white_space": view.showWhiteSpace,
                "view": view.view,
                "window_protection": view.windowProtection,
                "show_formulas": view.showFormulas,
                "show_vertical_scroll_bar": view.showVerticalScrollBar,
                "show_horizontal_scroll_bar": view.showHorizontalScrollBar,
                "show_sheet_tabs": view.showSheetTabs,
                "auto_filter_date_grouping": view.autoFilterDateGrouping,
                "sheet_state": view.sheetState
            }
            views.append(view_data)
        
        return views
    
    def _extract_sheet_protection(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract sheet protection information"""
        protection = worksheet.protection
        if not protection:
            return {
                "sheet": False,
                "objects": False,
                "scenarios": False,
                "format_cells": False,
                "format_columns": False,
                "format_rows": False,
                "insert_columns": False,
                "insert_rows": False,
                "insert_hyperlinks": False,
                "delete_columns": False,
                "delete_rows": False,
                "select_locked_cells": False,
                "sort": False,
                "auto_filter": False,
                "pivot_tables": False,
                "password": ""
            }
        
        return {
            "sheet": protection.sheet,
            "objects": protection.objects,
            "scenarios": protection.scenarios,
            "format_cells": protection.formatCells,
            "format_columns": protection.formatColumns,
            "format_rows": protection.formatRows,
            "insert_columns": protection.insertColumns,
            "insert_rows": protection.insertRows,
            "insert_hyperlinks": protection.insertHyperlinks,
            "delete_columns": protection.deleteColumns,
            "delete_rows": protection.deleteRows,
            "select_locked_cells": protection.selectLockedCells,
            "sort": protection.sort,
            "auto_filter": protection.autoFilter,
            "pivot_tables": protection.pivotTables,
            "password": protection.password or ""
        }
    
    def _extract_row_properties(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract row properties"""
        rows = []
        for row in worksheet.iter_rows():
            row_props = row[0].parent.row_dimensions.get(row[0].row)
            if row_props:
                row_data = {
                    "row_number": self._safe(getattr(row_props, 'index', None)),
                    "height": self._safe(getattr(row_props, 'height', None)),
                    "hidden": self._safe(getattr(row_props, 'hidden', False)),
                    "outline_level": self._safe(getattr(row_props, 'outline_level', 0)),
                    "collapsed": self._safe(getattr(row_props, 'collapsed', False)),
                    "style": self._safe(str(getattr(row_props, 'style', '')) if getattr(row_props, 'style', None) else ""),
                    "custom_height": self._safe(getattr(row_props, 'customHeight', False)),
                    "custom_format": self._safe(getattr(row_props, 'customFormat', False)),
                    "thick_top": self._safe(getattr(row_props, 'thickTop', False)),
                    "thick_bottom": self._safe(getattr(row_props, 'thickBottom', False)),
                    "ht": self._safe(getattr(row_props, 'ht', None)),
                    "s": self._safe(getattr(row_props, 's', 0)),
                    "custom_font": self._safe(getattr(row_props, 'customFont', False)),
                    "custom_border": self._safe(getattr(row_props, 'customBorder', False)),
                    "custom_pattern": self._safe(getattr(row_props, 'customPattern', False)),
                    "custom_protection": self._safe(getattr(row_props, 'customProtection', False)),
                    "custom_alignment": self._safe(getattr(row_props, 'customAlignment', False))
                }
                rows.append(row_data)
        return rows
    
    def _extract_column_properties(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract column properties"""
        columns = []
        for col_letter, col_props in worksheet.column_dimensions.items():
            if col_props:
                col_data = {
                    "column_letter": col_letter,
                    "column_number": column_index_from_string(col_letter),
                    "width": getattr(col_props, 'width', None),
                    "hidden": getattr(col_props, 'hidden', False),
                    "outline_level": getattr(col_props, 'outline_level', 0),
                    "collapsed": getattr(col_props, 'collapsed', False),
                    "style": str(getattr(col_props, 'style', '')) if getattr(col_props, 'style', None) else "",
                    "custom_width": getattr(col_props, 'customWidth', False),
                    "custom_format": getattr(col_props, 'customFormat', False),
                    "best_fit": getattr(col_props, 'bestFit', False),
                    "auto_size": getattr(col_props, 'autoSize', False),
                    "s": getattr(col_props, 's', 0),
                    "custom_font": getattr(col_props, 'customFont', False),
                    "custom_border": getattr(col_props, 'customBorder', False),
                    "custom_pattern": getattr(col_props, 'customPattern', False),
                    "custom_protection": getattr(col_props, 'customProtection', False),
                    "custom_alignment": getattr(col_props, 'customAlignment', False)
                }
                columns.append(col_data)
        return columns
    
    def _extract_cell_data(self, worksheet_with_formulas: Worksheet, worksheet_with_values: Worksheet) -> Dict[str, Any]:
        """Extract all cell data with formatting, formulas, and calculated values"""
        cells = {}
        
        for row in worksheet_with_formulas.iter_rows():
            for cell in row:
                # Get formula from the worksheet with formulas
                formula = cell.value if getattr(cell, 'data_type', None) == 'f' else None
                
                # Get calculated value from the worksheet with values
                value_cell = worksheet_with_values[cell.coordinate]
                calculated_value = self._extract_cell_value(value_cell) if value_cell.value is not None else None
                
                if calculated_value is not None or formula or cell.comment:
                    cell_data = {
                        "coordinate": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "column_letter": cell.column_letter,
                        "value": calculated_value,
                        "data_type": self._get_cell_data_type(value_cell),
                        "formula": formula,
                        "formula_type": self._get_formula_type(cell, formula),
                        "shared_formula": None,  # openpyxl doesn't expose this directly
                        "array_formula": None,   # openpyxl doesn't expose this directly
                        "comment": self._extract_cell_comment(cell),
                        "hyperlink": self._extract_cell_hyperlink(cell),
                        "style": self._extract_cell_style(cell),
                        "is_date": value_cell.is_date,
                        "is_time": getattr(value_cell, 'is_time', False),
                        "is_datetime": value_cell.is_date and not getattr(value_cell, 'is_time', False),
                        "is_number": isinstance(calculated_value, (int, float)),
                        "is_string": isinstance(calculated_value, str),
                        "is_boolean": isinstance(calculated_value, bool),
                        "is_error": value_cell.data_type == 'e',
                        "is_empty": calculated_value is None
                    }
                    cells[cell.coordinate] = cell_data
        
        return cells
    
    def _extract_cell_value(self, cell) -> Any:
        """Extract cell value with proper formatting"""
        if cell.value is None:
            return None
        
        if cell.is_date:
            return self._format_datetime(cell.value)
        
        return cell.value
    
    def _get_cell_data_type(self, cell) -> str:
        """Get cell data type"""
        if cell.value is None:
            return "null"
        elif isinstance(cell.value, str):
            return "string"
        elif isinstance(cell.value, (int, float)):
            return "number"
        elif isinstance(cell.value, bool):
            return "boolean"
        elif cell.is_date:
            return "date"
        elif cell.data_type == 'e':
            return "error"
        else:
            return "string"
    
    def _get_formula_type(self, cell, formula) -> Optional[str]:
        """Get formula type"""
        if not formula:
            return None
        if str(formula).startswith('{') and str(formula).endswith('}'):
            return "array"
        else:
            return "normal"
    
    def _extract_cell_comment(self, cell) -> Optional[Dict[str, Any]]:
        """Extract cell comment"""
        if not cell.comment:
            return None
        
        return {
            "author": cell.comment.author or "",
            "text": cell.comment.text or "",
            "height": cell.comment.height if hasattr(cell.comment, 'height') else None,
            "width": cell.comment.width if hasattr(cell.comment, 'width') else None
        }
    
    def _extract_cell_hyperlink(self, cell) -> Optional[Dict[str, Any]]:
        """Extract cell hyperlink"""
        if not cell.hyperlink:
            return None
        
        return {
            "target": cell.hyperlink.target or "",
            "location": cell.hyperlink.location or "",
            "tooltip": cell.hyperlink.tooltip or "",
            "display": cell.hyperlink.display or ""
        }
    
    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """Extract comprehensive cell styling"""
        style = {
            "font": self._extract_font_style(cell.font),
            "fill": self._extract_fill_style(cell.fill),
            "border": self._extract_border_style(cell.border),
            "alignment": self._extract_alignment_style(cell.alignment),
            "number_format": cell.number_format or "",
            "protection": self._extract_protection_style(cell.protection)
        }
        return style
    
    def _extract_font_style(self, font: Font) -> Dict[str, Any]:
        """Extract font styling"""
        if not font:
            return {}
        
        return {
            "name": font.name or "",
            "size": font.size,
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "strike": font.strike,
            "color": self._extract_color(font.color),
            "scheme": font.scheme or ""
        }
    
    def _extract_fill_style(self, fill: Fill) -> Dict[str, Any]:
        """Extract fill styling"""
        if not fill:
            return {}
        
        return {
            "fill_type": fill.fill_type or "",
            "start_color": self._extract_color(fill.start_color),
            "end_color": self._extract_color(fill.end_color)
        }
    
    def _extract_border_style(self, border: Border) -> Dict[str, Any]:
        """Extract border styling"""
        if not border:
            return {}
        
        return {
            "left": self._extract_side_style(border.left),
            "right": self._extract_side_style(border.right),
            "top": self._extract_side_style(border.top),
            "bottom": self._extract_side_style(border.bottom),
            "diagonal": self._extract_side_style(border.diagonal),
            "diagonal_direction": border.diagonal_direction or ""
        }
    
    def _extract_side_style(self, side) -> Dict[str, Any]:
        """Extract border side styling"""
        if not side:
            return {}
        
        return {
            "style": side.style or "",
            "color": self._extract_color(side.color)
        }
    
    def _extract_alignment_style(self, alignment: Alignment) -> Dict[str, Any]:
        """Extract alignment styling"""
        if not alignment:
            return {}
        
        return {
            "horizontal": getattr(alignment, 'horizontal', None),
            "vertical": getattr(alignment, 'vertical', None),
            "text_rotation": getattr(alignment, 'text_rotation', 0),
            "wrap_text": getattr(alignment, 'wrap_text', False),
            "shrink_to_fit": getattr(alignment, 'shrink_to_fit', False),
            "indent": getattr(alignment, 'indent', 0),
            "relative_indent": getattr(alignment, 'relative_indent', 0),
            "justify_last_line": getattr(alignment, 'justify_last_line', False),
            "reading_order": getattr(alignment, 'reading_order', 0)
        }
    
    def _extract_protection_style(self, protection: Protection) -> Dict[str, Any]:
        """Extract protection styling"""
        if not protection:
            return {}
        
        return {
            "locked": protection.locked,
            "hidden": protection.hidden
        }
    
    def _safe(self, val):
        # Convert openpyxl types and other non-serializable types to serializable Python types
        if hasattr(val, 'value'):
            return val.value
        if isinstance(val, (int, float, str, bool, type(None))):
            return val
        if isinstance(val, (list, tuple)):
            return [self._safe(v) for v in val]
        if isinstance(val, dict):
            return {self._safe(k): self._safe(v) for k, v in val.items()}
        try:
            return str(val)
        except Exception:
            return None

    def _extract_color(self, color) -> Optional[Dict[str, Any]]:
        if not color:
            return None
        return {
            "rgb": self._safe(getattr(color, 'rgb', None)),
            "theme": self._safe(getattr(color, 'theme', None)),
            "tint": self._safe(getattr(color, 'tint', None))
        }

    def _extract_merged_cells(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            start_cell = getattr(merged_range, 'start_cell', None)
            end_cell = getattr(merged_range, 'end_cell', None)
            merged_data = {
                "range": str(merged_range),
                "start_cell": self._safe(getattr(start_cell, 'coordinate', start_cell)),
                "end_cell": self._safe(getattr(end_cell, 'coordinate', end_cell)),
                "start_row": self._safe(getattr(start_cell, 'row', None)),
                "start_column": self._safe(getattr(start_cell, 'column', None)),
                "end_row": self._safe(getattr(end_cell, 'row', None)),
                "end_column": self._safe(getattr(end_cell, 'column', None))
            }
            merged_cells.append(merged_data)
        return merged_cells

    def _extract_data_validations(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        validations = []
        for validation in getattr(worksheet.data_validations, 'dataValidation', []):
            validation_data = {
                "range": self._safe(getattr(validation, 'range', '')),
                "type": self._safe(getattr(validation, 'type', '')),
                "operator": self._safe(getattr(validation, 'operator', '')),
                "formula1": self._safe(getattr(validation, 'formula1', '')),
                "formula2": self._safe(getattr(validation, 'formula2', '')),
                "allow_blank": self._safe(getattr(validation, 'allowBlank', False)),
                "show_error_message": self._safe(getattr(validation, 'showErrorMessage', False)),
                "error_title": self._safe(getattr(validation, 'errorTitle', '')),
                "error_message": self._safe(getattr(validation, 'error', '')),
                "show_input_message": self._safe(getattr(validation, 'showInputMessage', False)),
                "input_title": self._safe(getattr(validation, 'promptTitle', '')),
                "input_message": self._safe(getattr(validation, 'prompt', '')),
                "prompt_title": self._safe(getattr(validation, 'promptTitle', '')),
                "prompt_message": self._safe(getattr(validation, 'prompt', ''))
            }
            validations.append(validation_data)
        return validations

    def _extract_conditional_formatting(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        conditional_formats = []
        for cf in getattr(worksheet, 'conditional_formatting', []):
            # Handle MultiCellRange object properly
            cell_range = ""
            try:
                cells_attr = getattr(cf, 'cells', None)
                if cells_attr is not None:
                    cell_range = str(cells_attr)
            except:
                cell_range = ""
            
            cf_data = {
                "range": self._safe(cell_range),
                "priority": self._safe(getattr(cf, 'priority', None)),
                "stop_if_true": self._safe(getattr(cf, 'stopIfTrue', None)),
                "type": self._safe(getattr(getattr(cf, 'cfRule', [{}])[0], 'type', '') if getattr(cf, 'cfRule', None) else ''),
                "color_scale": {},
                "data_bar": {},
                "icon_set": {},
                "dxf": {}
            }
            conditional_formats.append(cf_data)
        return conditional_formats

    def _extract_charts(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        charts = []
        for chart in getattr(worksheet, '_charts', []):
            chart_data = {
                "chart_type": self._safe(getattr(chart, 'chart_type', '')),
                "title": self._safe(getattr(chart, 'title', '')),
                "x_axis": {},
                "y_axis": {},
                "series": [],
                "position": {}
            }
            charts.append(chart_data)
        return charts

    def _extract_images(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        images = []
        for image in getattr(worksheet, '_images', []):
            image_data = {
                "image_type": "unknown",
                "position": {},
                "anchor": "",
                "filename": "",
                "data": ""
            }
            images.append(image_data)
        return images

    def _extract_comments(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        comments = []
        for row in worksheet.iter_rows():
            for cell_obj in row:
                if cell_obj.comment:
                    comment_data = {
                        "cell": self._safe(cell_obj.coordinate),
                        "author": self._safe(getattr(cell_obj.comment, 'author', '')),
                        "text": self._safe(getattr(cell_obj.comment, 'text', '')),
                        "position": {},
                        "visible": True
                    }
                    comments.append(comment_data)
        return comments
    
    def _format_datetime(self, dt) -> Optional[str]:
        """Format datetime objects to ISO 8601 string"""
        if dt is None:
            return None
        
        if isinstance(dt, datetime):
            return dt.isoformat()
        
        return str(dt) 