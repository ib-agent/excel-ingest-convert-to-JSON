import os
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.comments import Comment


def write_verbose_json_to_excel(excel_json: Dict[str, Any], output_path: str) -> None:
    """
    Create an Excel file at output_path from the verbose Excel JSON schema.

    Notes and constraints:
    - Cached formula values cannot be embedded reliably; formulas are written, but
      Excel must calculate them. For round-trip comparisons, prefer comparing
      formula strings and non-formula values.
    - Only commonly supported workbook/sheet properties are restored.
    - This function is intentionally generic and does not hardcode any document-specific logic.
    """
    workbook = Workbook()

    # Remove default sheet if we'll add our own
    if workbook.active and excel_json.get("workbook", {}).get("sheets"):
        workbook.remove(workbook.active)

    workbook_data = excel_json.get("workbook", {})

    # Restore workbook document properties when possible
    metadata = workbook_data.get("metadata", {})
    _apply_workbook_metadata(workbook, metadata)

    # Basic workbook properties (limited support)
    # Many of the fields in the verbose schema are not exposed by openpyxl
    # and will not be round-trippable. We avoid setting unsupported fields.

    # Create sheets
    for sheet in workbook_data.get("sheets", []):
        title = sheet.get("name") or "Sheet"
        ws = workbook.create_sheet(title=title)

        # Frozen panes
        frozen = sheet.get("frozen_panes", {})
        top_left = frozen.get("top_left_cell")
        if top_left:
            # openpyxl accepts an A1 reference
            ws.freeze_panes = top_left

        # Column properties
        for col in sheet.get("columns", []) or []:
            col_letter = col.get("column_letter")
            if not col_letter:
                continue
            dim = ws.column_dimensions[col_letter]
            if col.get("width") is not None:
                dim.width = col.get("width")
            if col.get("hidden") is not None:
                dim.hidden = bool(col.get("hidden"))
            if col.get("outline_level") is not None:
                dim.outline_level = int(col.get("outline_level"))
            if col.get("collapsed") is not None:
                dim.collapsed = bool(col.get("collapsed"))

        # Row properties
        for rowp in sheet.get("rows", []) or []:
            row_number = rowp.get("row_number")
            if not row_number:
                continue
            rd = ws.row_dimensions[row_number]
            if rowp.get("height") is not None:
                rd.height = rowp.get("height")
            if rowp.get("hidden") is not None:
                rd.hidden = bool(rowp.get("hidden"))
            if rowp.get("outline_level") is not None:
                rd.outline_level = int(rowp.get("outline_level"))
            if rowp.get("collapsed") is not None:
                rd.collapsed = bool(rowp.get("collapsed"))

        # Cells
        for coord, cell in (sheet.get("cells") or {}).items():
            ws_cell = ws[coord]

            # Value and formula
            formula = cell.get("formula")
            if formula:
                ws_cell.value = str(formula)
            else:
                ws_cell.value = cell.get("value")

            # Number format
            num_fmt = (cell.get("style") or {}).get("number_format")
            if num_fmt:
                ws_cell.number_format = num_fmt

            # Style
            _apply_cell_style(ws_cell, cell.get("style") or {})

            # Hyperlink
            hyperlink = cell.get("hyperlink")
            if hyperlink and hyperlink.get("target"):
                ws_cell.hyperlink = hyperlink.get("target")
                if hyperlink.get("tooltip"):
                    ws_cell.hyperlink.tooltip = hyperlink.get("tooltip")

            # Comment
            comment = cell.get("comment")
            if comment and (comment.get("author") or comment.get("text")):
                ws_cell.comment = Comment(text=comment.get("text") or "", author=comment.get("author") or "")

        # Merged cells
        for m in sheet.get("merged_cells", []) or []:
            cell_range = m.get("range")
            if cell_range:
                try:
                    ws.merge_cells(cell_range)
                except Exception:
                    # Skip invalid ranges quietly to keep generic behavior
                    pass

        # Data validations not reconstructed due to API complexity and fidelity constraints
        # Conditional formatting not reconstructed for same reason

    # Save workbook
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    workbook.save(output_path)


def _apply_workbook_metadata(workbook: Workbook, metadata: Dict[str, Any]) -> None:
    """Map verbose metadata fields to openpyxl workbook.properties when possible."""
    props = workbook.properties

    # Best-effort mapping; many metadata fields are not supported by openpyxl
    if metadata.get("creator"):
        props.creator = metadata.get("creator")
    if metadata.get("last_modified_by"):
        props.lastModifiedBy = metadata.get("last_modified_by")
    if metadata.get("title"):
        props.title = metadata.get("title")
    if metadata.get("subject"):
        props.subject = metadata.get("subject")
    if metadata.get("keywords"):
        props.keywords = metadata.get("keywords")
    if metadata.get("category"):
        props.category = metadata.get("category")
    if metadata.get("description"):
        props.description = metadata.get("description")
    if metadata.get("language"):
        props.language = metadata.get("language")
    if metadata.get("revision") is not None:
        try:
            props.revision = int(metadata.get("revision"))
        except Exception:
            pass
    if metadata.get("version"):
        props.version = metadata.get("version")
    if metadata.get("application"):
        props.application = metadata.get("application")
    if metadata.get("app_version"):
        props.appName = metadata.get("app_version")
    if metadata.get("company"):
        props.company = metadata.get("company")
    if metadata.get("manager"):
        props.manager = metadata.get("manager")
    if metadata.get("hyperlink_base"):
        props.hyperlinkBase = metadata.get("hyperlink_base")
    if metadata.get("template"):
        props.template = metadata.get("template")
    if metadata.get("status"):
        props.status = metadata.get("status")

    # Dates
    created = _parse_iso_datetime(metadata.get("created_date"))
    if created:
        props.created = created
    modified = _parse_iso_datetime(metadata.get("modified_date"))
    if modified:
        props.modified = modified


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def _apply_cell_style(ws_cell, style: Dict[str, Any]) -> None:
    # Font
    font_dict = style.get("font") or {}
    if font_dict:
        ws_cell.font = Font(
            name=font_dict.get("name"),
            size=font_dict.get("size"),
            bold=font_dict.get("bold"),
            italic=font_dict.get("italic"),
            underline=font_dict.get("underline"),
            strike=font_dict.get("strike"),
        )

    # Fill
    fill_dict = style.get("fill") or {}
    if fill_dict:
        # openpyxl uses PatternFill for solid/gradient fills; we map basic solid
        fill_type = fill_dict.get("fill_type") or fill_dict.get("type")
        start_color = (fill_dict.get("start_color") or {}).get("rgb") or None
        if fill_type:
            ws_cell.fill = PatternFill(fill_type=fill_type, start_color=start_color, end_color=start_color)

    # Border
    border_dict = style.get("border") or {}
    if border_dict:
        def side_from(name: str) -> Side:
            side_style = None
            side_color_dict = None
            side_entry = border_dict.get(name)
            if isinstance(side_entry, dict):
                side_style = side_entry.get("style")
                side_color_dict = (side_entry.get("color") or {})
            elif isinstance(side_entry, str):
                side_style = side_entry
                side_color_dict = None
            color_rgb = (side_color_dict or {}).get("rgb") if side_color_dict else None
            return Side(style=side_style, color=color_rgb)

        ws_cell.border = Border(
            left=side_from("left"),
            right=side_from("right"),
            top=side_from("top"),
            bottom=side_from("bottom"),
            diagonal=side_from("diagonal"),
        )

    # Alignment
    align_dict = style.get("alignment") or {}
    if align_dict:
        # Only pass supported kwargs and correct types for this project's openpyxl version
        kwargs = {}
        if align_dict.get("horizontal") is not None:
            kwargs["horizontal"] = align_dict.get("horizontal")
        if align_dict.get("vertical") is not None:
            kwargs["vertical"] = align_dict.get("vertical")
        if align_dict.get("wrap_text") is not None:
            kwargs["wrap_text"] = bool(align_dict.get("wrap_text"))
        if align_dict.get("shrink_to_fit") is not None:
            kwargs["shrink_to_fit"] = bool(align_dict.get("shrink_to_fit"))
        # text_rotation should be an int in valid range
        tr = align_dict.get("text_rotation")
        if isinstance(tr, (int, float)):
            kwargs["text_rotation"] = int(tr)
        # indent should be an int
        ind = align_dict.get("indent")
        if isinstance(ind, (int, float)):
            kwargs["indent"] = int(ind)
        ws_cell.alignment = Alignment(**kwargs)

    # Protection
    protection_dict = style.get("protection") or {}
    if protection_dict:
        ws_cell.protection = Protection(
            locked=protection_dict.get("locked"),
            hidden=protection_dict.get("hidden"),
        )


