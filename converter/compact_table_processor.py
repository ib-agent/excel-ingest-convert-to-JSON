import json
from typing import Dict, Any, List, Optional, Tuple
from openpyxl.utils import get_column_letter, column_index_from_string
from .table_detector import TableDetector, HeaderResolver


class CompactTableProcessor:
    """
    Transforms compact Excel JSON schema into compact table-oriented representation
    while eliminating data duplication and maintaining all table structure information
    """
    
    def __init__(self):
        self.table_detection_rules = []
        # Add new simplified detection system
        self.detector = TableDetector()
        self.header_resolver = HeaderResolver()
    
    def transform_to_compact_table_format(self, compact_excel_json: dict, options: dict = None) -> dict:
        """
        Transform compact Excel JSON schema to compact table-oriented format
        
        Args:
            compact_excel_json: The compact Excel JSON schema output
            options: Configuration options for table detection and processing
            
        Returns:
            Dictionary with compact table-oriented structure
        """
        if not compact_excel_json or 'workbook' not in compact_excel_json:
            raise ValueError("Invalid compact Excel JSON format")
        
        # Start with the original structure
        result = compact_excel_json.copy()
        
        # Process each sheet to add table information
        for sheet in result['workbook']['sheets']:
            if 'rows' in sheet:
                sheet['tables'] = self._detect_and_process_compact_tables(sheet, options or {})
        
        return result
    
    def _detect_and_process_compact_tables(self, sheet_data: dict, options: dict) -> List[Dict[str, Any]]:
        """
        Detect tables in a sheet and process them into compact table format
        
        Args:
            sheet_data: Sheet data from the compact JSON
            options: Processing options
            
        Returns:
            List of compact table objects
        """
        tables = []
        
        # Get sheet data in compact format
        rows = sheet_data.get('rows', [])
        dimensions = sheet_data.get('dimensions', [1, 1, 1, 1])
        
        # Prepare options for new detection system
        detection_options = options.copy()
        detection_options['sheet_data'] = sheet_data
        
        # Use new simplified detection system with compact data
        table_regions = self.detector.detect_tables(rows, dimensions, detection_options)
        
        # Process each detected table region
        for i, region in enumerate(table_regions):
            table = self._process_compact_table_region(
                sheet_data, region, i, detection_options
            )
            if table:
                tables.append(table)
        
        # If no tables detected, create a default table for the entire sheet
        if not tables and rows:
            default_table = self._create_compact_default_table(sheet_data, options)
            if default_table:
                tables.append(default_table)
        
        return tables
    
    def _build_cell_map_from_compact_rows(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Convert compact row format back to cell map for table detection"""
        cell_map = {}
        
        for row_data in rows:
            row_num = row_data.get('r', 1)
            cells = row_data.get('cells', [])
            
            for cell_array in cells:
                if len(cell_array) >= 2:
                    col_num = cell_array[0]
                    value = cell_array[1]
                    
                    cell_key = f"{get_column_letter(col_num)}{row_num}"
                    cell_map[cell_key] = {
                        'value': value,
                        'row': row_num,
                        'column': col_num
                    }
        
        return cell_map
    

    

    

    
    def _process_compact_table_region(self, sheet_data: dict, region: dict, 
                                    table_index: int, options: dict) -> Optional[Dict[str, Any]]:
        """
        Process a table region into compact table format
        
        Args:
            sheet_data: Original sheet data
            region: Table region coordinates
            table_index: Index of the table
            options: Processing options
            
        Returns:
            Compact table object
        """
        rows = sheet_data.get('rows', [])
        if not rows:
            return None
        
        # Determine header information using new resolver
        # Convert compact rows to normalized cells for header resolution
        cell_map = self._build_cell_map_from_compact_rows(rows)
        normalized_cells = self.detector._normalize_cell_data(cell_map)
        header_info = self.header_resolver.resolve_headers(normalized_cells, region, options)
        
        # Detect table title (row above the detected region)
        title_info = self._detect_table_title(rows, region, options)
        
        # Adjust region to exclude title row if found
        adjusted_region = region.copy()
        if title_info['title_row'] is not None:
            # If title row is the first row of the region, adjust start to exclude it
            if adjusted_region['start_row'] == title_info['title_row']:
                adjusted_region['start_row'] = title_info['title_row'] + 1
            # If title row is above the region, no adjustment needed
            elif title_info['title_row'] < adjusted_region['start_row']:
                pass
        
        # Create compact table structure
        table = {
            'id': f"t{table_index + 1}",
            'name': title_info['title'] or f"Table {table_index + 1}",
            'title': title_info['title'],
            'title_row': title_info['title_row'],
            'region': [
                adjusted_region['start_row'],
                adjusted_region['start_col'],
                adjusted_region['end_row'],
                adjusted_region['end_col']
            ],
            'headers': {
                'rows': header_info['header_rows'],
                'cols': header_info['header_columns'],
                'data_start': [header_info['data_start_row'], header_info['data_start_col']]
            },
            'labels': {
                'cols': self._create_compact_column_labels(rows, header_info, adjusted_region),
                'rows': self._create_compact_row_labels(rows, header_info, adjusted_region)
            },
            'meta': {
                'method': region.get('detection_method', 'unknown'),
                'cells': self._count_cells_in_region(rows, adjusted_region),
                'numeric_cells': self._count_numeric_cells_in_region(rows, adjusted_region),
                'merged': len(sheet_data.get('merged', [])) > 0,
                'original_region': [region['start_row'], region['start_col'], region['end_row'], region['end_col']]
            }
        }
        
        return table
    
    def _detect_table_title(self, rows: List[Dict[str, Any]], region: dict, options: dict) -> Dict[str, Any]:
        """
        Detect table title in the row immediately above the table region.
        
        Args:
            rows: Sheet rows data
            region: Table region coordinates
            options: Processing options
            
        Returns:
            Dictionary with title information: {'title': str, 'title_row': int}
        """
        start_row = region['start_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        # Check if the first row of the region could be a title
        # or check the row immediately above the table
        potential_title_rows = [start_row - 1, start_row]
        
        # Build row lookup for quick access
        row_lookup = {}
        for row_data in rows:
            row_num = row_data.get('r', 0)
            row_lookup[row_num] = row_data
        
        for potential_title_row in potential_title_rows:
            if potential_title_row in row_lookup:
                row_data = row_lookup[potential_title_row]
                cells = row_data.get('cells', [])
                
                # Look for a single cell in the leftmost position that could be a title
                title_candidates = []
                data_candidates = []
                
                for cell_array in cells:
                    if len(cell_array) >= 2:
                        col_num = cell_array[0]
                        value = cell_array[1]
                        
                        # Check if this is in the table's column range and has text content
                        if start_col <= col_num <= end_col and value is not None and str(value).strip():
                            if isinstance(value, str) and len(value.strip()) > 0:
                                title_candidates.append((col_num, value.strip()))
                            else:
                                data_candidates.append((col_num, value))
                
                # If we have exactly one text value in the leftmost column and no other data, it's likely a title
                if (len(title_candidates) == 1 and title_candidates[0][0] == start_col and 
                    len(data_candidates) == 0):
                    title_text = title_candidates[0][1]
                    
                    # Validate that this looks like a title (not just random text)
                    if self._is_likely_table_title(title_text):
                        return {
                            'title': title_text,
                            'title_row': potential_title_row
                        }
        
        return {
            'title': None,
            'title_row': None
        }
    
    def _is_likely_table_title(self, text: str) -> bool:
        """
        Determine if text is likely a table title.
        
        Args:
            text: Text to evaluate
            
        Returns:
            True if text looks like a table title
        """
        # Basic heuristics for table titles
        text = text.strip()
        
        # Must be reasonable length
        if len(text) < 3 or len(text) > 100:
            return False
            
        # Should not be purely numeric
        if text.replace('.', '').replace('-', '').replace('/', '').isdigit():
            return False
            
        # Should not be a date (basic check)
        if any(date_indicator in text.lower() for date_indicator in ['2024', '2025', '2026', 'jan', 'feb', 'mar']):
            return False
            
        # Should contain some alphabetic characters
        if not any(c.isalpha() for c in text):
            return False
            
        # Common title patterns
        title_indicators = ['table', 'summary', 'report', 'data', 'cost', 'revenue', 'analysis']
        if any(indicator in text.lower() for indicator in title_indicators):
            return True
            
        # If it passes basic checks and isn't obviously not a title, assume it is
        return True
    
    def _determine_compact_headers(self, rows: List[Dict[str, Any]], region: dict, options: dict) -> Dict[str, Any]:
        """Determine header rows and columns for the compact table"""
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        # Get frozen panes information from sheet data
        sheet_data = options.get('sheet_data', {})
        frozen_panes = sheet_data.get('frozen', [0, 0])
        frozen_rows, frozen_cols = frozen_panes if len(frozen_panes) >= 2 else [0, 0]
        
        # Determine header rows
        header_rows = []
        if frozen_rows > 0:
            for row in range(start_row, start_row + frozen_rows):
                if row <= end_row:
                    header_rows.append(row)
        
        if not header_rows and start_row < end_row:
            header_rows.append(start_row)
        
        # Determine header columns
        header_cols = []
        if frozen_cols > 0:
            for col in range(start_col, start_col + frozen_cols):
                if col <= end_col:
                    header_cols.append(col)
        
        if not header_cols and start_col < end_col:
            header_cols.append(start_col)
        
        return {
            'header_rows': header_rows,
            'header_columns': header_cols,
            'data_start_row': start_row + len(header_rows),
            'data_start_col': start_col + len(header_cols)
        }
    
    def _create_compact_column_labels(self, rows: List[Dict[str, Any]], 
                                    header_info: dict, region: dict) -> List[str]:
        """Create compact column labels array"""
        labels = []
        start_col = region['start_col']
        end_col = region['end_col']
        header_rows = header_info['header_rows']
        
        # Create row lookup for easier access
        row_lookup = {row_data['r']: row_data for row_data in rows}
        
        for col in range(start_col, end_col + 1):
            col_labels = []
            
            for header_row in header_rows:
                if header_row in row_lookup:
                    row_data = row_lookup[header_row]
                    # Find cell for this column
                    for cell_array in row_data.get('cells', []):
                        if len(cell_array) >= 2 and cell_array[0] == col:
                            value = cell_array[1]
                            if value is not None:
                                col_labels.append(str(value))
                            break
            
            label = " | ".join(col_labels) if col_labels else f"Column {get_column_letter(col)}"
            labels.append(label)
        
        return labels
    
    def _create_compact_row_labels(self, rows: List[Dict[str, Any]], 
                                 header_info: dict, region: dict) -> List[str]:
        """Create compact row labels array for data rows only"""
        labels = []
        start_row = region['start_row']
        end_row = region['end_row']
        header_cols = header_info['header_columns']
        data_start_row = header_info['data_start_row']
        
        # Create row lookup for easier access
        row_lookup = {row_data['r']: row_data for row_data in rows}
        
        for row in range(data_start_row, end_row + 1):
            if row in row_lookup:
                row_data = row_lookup[row]
                row_labels = []
                
                for header_col in header_cols:
                    # Find cell for this header column
                    for cell_array in row_data.get('cells', []):
                        if len(cell_array) >= 2 and cell_array[0] == header_col:
                            value = cell_array[1]
                            if value is not None:
                                row_labels.append(str(value))
                            break
                
                label = " | ".join(row_labels) if row_labels else f"Row {row}"
                labels.append(label)
        
        return labels
    
    def _count_cells_in_region(self, rows: List[Dict[str, Any]], region: dict) -> int:
        """Count non-empty cells in the table region"""
        count = 0
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        for row_data in rows:
            row_num = row_data.get('r', 1)
            if start_row <= row_num <= end_row:
                for cell_array in row_data.get('cells', []):
                    if len(cell_array) >= 2:
                        col_num = cell_array[0]
                        value = cell_array[1]
                        if start_col <= col_num <= end_col and value is not None:
                            # Account for RLE cells by run length
                            if self._is_rle_cell(cell_array):
                                count += max(int(cell_array[-1]), 0)
                            else:
                                count += 1
        
        return count

    def _count_numeric_cells_in_region(self, rows: List[Dict[str, Any]], region: dict) -> int:
        """Count numeric (int/float, excluding bool) cells in the region (accounts for RLE)."""
        count = 0
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']
        
        for row_data in rows:
            row_num = row_data.get('r', 1)
            if start_row <= row_num <= end_row:
                for cell_array in row_data.get('cells', []):
                    if len(cell_array) >= 2:
                        col_num = cell_array[0]
                        value = cell_array[1]
                        if start_col <= col_num <= end_col and self._is_numeric(value):
                            if self._is_rle_cell(cell_array):
                                count += max(int(cell_array[-1]), 0)
                            else:
                                count += 1
        return count

    def _is_rle_cell(self, cell: List[Any]) -> bool:
        """Detect if a compact-format cell array uses RLE encoding."""
        return isinstance(cell, list) and len(cell) >= 5 and isinstance(cell[-1], int) and cell[-1] > 1

    def _is_numeric(self, value: Any) -> bool:
        """Return True if value is a number (int/float) but not a bool."""
        if isinstance(value, bool):
            return False
        return isinstance(value, (int, float))
    
    def _create_compact_default_table(self, sheet_data: dict, options: dict) -> Optional[Dict[str, Any]]:
        """Create a default compact table for the entire sheet if no tables detected"""
        rows = sheet_data.get('rows', [])
        if not rows:
            return None
        
        dimensions = sheet_data.get('dimensions', {'min_row': 1, 'max_row': 1, 'min_col': 1, 'max_col': 1})
        region = {
            'start_row': dimensions.get('min_row', 1),
            'end_row': dimensions.get('max_row', 1),
            'start_col': dimensions.get('min_col', 1),
            'end_col': dimensions.get('max_col', 1),
            'detection_method': 'default'
        }
        
        return self._process_compact_table_region(sheet_data, region, 0, options) 